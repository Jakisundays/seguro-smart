import streamlit as st
import pandas as pd
from io import BytesIO
import base64
from typing import List, Optional
from typing import TypedDict, Dict, List, Optional, Literal
import os
import certifi
import aiohttp
import ssl
from polizas_tools import tools as tools_standard
import asyncio
from typing import Optional, List, TypedDict
import uuid
from dotenv import load_dotenv
import logging
from pathlib import Path
from dataclasses import dataclass, field
import json


@dataclass
class Prima:
    """
    Representa el valor de una prima individual con desglose de IVA.
    """

    prima_sin_iva: float
    iva: float
    prima_con_iva: float


@dataclass
class DetalleCoberturaItem:
    """
    Representa un interés asegurado con su valor asegurado.
    """

    interes_asegurado: str  # Ejemplo: "Edificio", "Maquinaria"
    valor_asegurado: float  # Valor monetario asegurado


@dataclass
class Cobertura:
    """
    Representa el detalle completo de la cobertura.
    """

    detalle_cobertura: List[DetalleCoberturaItem] = field(default_factory=list)
    total_valores_asegurados: float = 0.0

    def calcular_total(self) -> float:
        """
        Calcula y actualiza la suma total de los valores asegurados.
        """
        self.total_valores_asegurados = sum(
            item.valor_asegurado for item in self.detalle_cobertura
        )
        return self.total_valores_asegurados


load_dotenv()

# Configurar logger
app_logger = logging.getLogger(__name__)
app_logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)
app_logger.addHandler(handler)


class QueueItem(TypedDict):
    file_name: str
    file_extension: str
    file_path: str
    media_type: str
    process_id: str
    doc_type: Literal["actual", "renovacion", "adicional"]


# Funciones auxiliares para manejo de archivos
def save_uploaded_file(uploaded_file, doc_type: str) -> Optional[QueueItem]:
    """Guarda un archivo subido directamente en la carpeta downloads y crea un QueueItem"""
    if uploaded_file is None:
        return None

    try:
        # Crear directorio downloads si no existe
        downloads_dir = os.path.join(os.getcwd(), "downloads")
        os.makedirs(downloads_dir, exist_ok=True)

        # Generar nombre único para el archivo
        file_extension = uploaded_file.name.split(".")[-1].lower()
        timestamp = str(int(uuid.uuid4().int))[0:8]
        unique_filename = f"{timestamp}_{uploaded_file.name}"
        file_path = os.path.join(downloads_dir, unique_filename)

        # Guardar el archivo directamente en downloads
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # Determinar media type
        media_type_map = {
            "pdf": "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "txt": "text/plain",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
        }

        media_type = media_type_map.get(file_extension, "application/octet-stream")

        # Crear QueueItem
        queue_item: QueueItem = {
            "file_name": uploaded_file.name,
            "file_extension": file_extension,
            "file_path": file_path,
            "media_type": media_type,
            "process_id": str(uuid.uuid4()),
            "doc_type": doc_type,
        }

        return queue_item

    except Exception as e:
        st.error(f"Error al guardar el archivo {uploaded_file.name}: {str(e)}")
        return None


def get_file_info(file_path: str, file_name: str) -> str:
    """Obtiene información del archivo descargado"""
    try:
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            return f"✅ **{file_name}** descargado ({file_size_mb:.2f} MB)\n📁 Ubicación: `{file_path}`"
        else:
            return f"❌ Error: No se encontró el archivo {file_name}"
    except Exception as e:
        return f"❌ Error al obtener información de {file_name}: {str(e)}"


def process_files_and_create_queue(
    archivo_actual, archivo_renovacion, archivos_multiples
) -> List[QueueItem]:
    """Procesa todos los archivos y crea la cola de procesamiento"""
    queue_items = []

    # Procesar póliza actual
    if archivo_actual:
        item = save_uploaded_file(archivo_actual, "actual")
        if item:
            queue_items.append(item)

    # Procesar póliza de renovación
    if archivo_renovacion:
        item = save_uploaded_file(archivo_renovacion, "renovacion")
        if item:
            queue_items.append(item)

    # Procesar documentos adicionales
    if archivos_multiples:
        for archivo in archivos_multiples:
            item = save_uploaded_file(archivo, "adicional")
            if item:
                queue_items.append(item)

    return queue_items


class QueueItem(TypedDict):
    file_name: str
    file_extension: str
    file_path: str
    media_type: str
    process_id: str
    doc_type: Literal["actual", "renovacion", "adicional"]


# Configuración de la página
st.set_page_config(
    page_title="Análisis de pólizas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Título principal
st.title("📊 Análisis de pólizas")
st.markdown("---")

# Área principal - Mostrar estado de la cola si existe
if "queue_items" in st.session_state and st.session_state.queue_items:
    st.header("📊 Estado del Sistema")

    # Métricas generales
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total de Archivos", len(st.session_state.queue_items))

    with col2:
        actual_count = sum(
            1 for item in st.session_state.queue_items if item["doc_type"] == "actual"
        )
        st.metric("Pólizas Actuales", actual_count)

    with col3:
        renovacion_count = sum(
            1
            for item in st.session_state.queue_items
            if item["doc_type"] == "renovacion"
        )
        st.metric("Pólizas Renovación", renovacion_count)

    with col4:
        adicional_count = sum(
            1
            for item in st.session_state.queue_items
            if item["doc_type"] == "adicional"
        )
        st.metric("Documentos Adicionales", adicional_count)

    st.markdown("---")

    # Tabla detallada de la cola
    st.subheader("📋 Cola de Procesamiento Actual")

    queue_display_data = []
    for i, item in enumerate(st.session_state.queue_items, 1):
        queue_display_data.append(
            {
                "#": i,
                "Archivo": item["file_name"],
                "Tipo de Documento": item["doc_type"].title(),
                "Extensión": item["file_extension"].upper(),
                "Media Type": item["media_type"],
                "Process ID": item["process_id"],
                "Estado": "🟢 En Cola",
            }
        )

    df_detailed = pd.DataFrame(queue_display_data)
    st.dataframe(df_detailed, use_container_width=True)

    # Botón para limpiar la cola
    if st.button("🗑️ Limpiar Cola", type="secondary"):
        # Limpiar archivos descargados
        deleted_count = 0
        for item in st.session_state.queue_items:
            try:
                if os.path.exists(item["file_path"]):
                    os.remove(item["file_path"])
                    deleted_count += 1
            except Exception as e:
                st.warning(f"No se pudo eliminar el archivo {item['file_name']}: {e}")

        # Limpiar session state
        st.session_state.queue_items = []
        st.success(
            f"✅ Cola limpiada exitosamente. Se eliminaron {deleted_count} archivo(s) de la carpeta downloads."
        )
        st.rerun()

    st.markdown("---")
else:
    st.info(
        "👈 Utiliza la barra lateral para cargar archivos y comenzar el procesamiento."
    )
    st.markdown("---")


class InvoiceOrchestrator:
    def __init__(
        self,
        api_key: str,
        model: str,
    ):
        self.api_key = api_key
        self.model = model

    # Hace requests a la API con reintentos
    async def make_api_request(
        self, url: str, headers: Dict, data: Dict, process_id: str, retries: int = 5
    ) -> Optional[Dict]:
        ssl_context = ssl.create_default_context(cafile=certifi.where())
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        async with aiohttp.ClientSession(connector=connector) as session:
            for i in range(retries):
                try:
                    async with session.post(
                        url, headers=headers, json=data
                    ) as response:
                        if response.status == 200:
                            return await response.json()
                        elif response.status in [429, 529, 503]:
                            sleep_time = 15 * (i + 1)  # Espera incremental en segundos
                            app_logger.warning(
                                f"API request failed with status {response.status}. Retrying in {sleep_time} seconds..."
                            )
                            await asyncio.sleep(sleep_time)
                        else:
                            app_logger.error(
                                f"API request failed with status {response.status} - {await response.text()}"
                            )
                            raise ValueError(
                                f"Request failed with status {response.status}"
                            )
                except aiohttp.ClientError as e:
                    raise ValueError(f"Request error: {str(e)}")
        raise ValueError("Max retries exceeded.")

    async def handle_pdf(self, item: QueueItem):
        file_path = Path(item["file_path"])
        encoded_pdf = base64.b64encode(file_path.read_bytes()).decode()

        prompt = (
            tools_standard[1]["prompt"]
            if item["doc_type"] == "adicional"
            else tools_standard[0]["prompt"]
        )
        responseSchema = (
            tools_standard[1]["data"]
            if item["doc_type"] == "adicional"
            else tools_standard[0]["data"]
        )
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"
        headers = {"x-goog-api-key": self.api_key, "Content-Type": "application/json"}
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "inline_data": {
                                "mime_type": "application/pdf",
                                "data": encoded_pdf,
                            }
                        },
                        {"text": prompt},
                    ]
                }
            ],
            "generationConfig": {
                "responseMimeType": "application/json",
                "responseSchema": responseSchema,
            },
        }
        response = await self.make_api_request(
            url, headers, payload, item["process_id"]
        )
        item["data"] = response
        return item


orchestrator = InvoiceOrchestrator(
    api_key=os.getenv("GEMINI_API_KEY"),
    model="gemini-2.5-flash",
)


# Barra lateral
with st.sidebar:

    # Campo 1: Póliza Actual
    st.subheader("1️⃣ Póliza Actual")
    archivo_poliza_actual = st.file_uploader(
        "Cargar archivo de póliza actual",
        type=["pdf", "docx", "txt", "jpg", "png"],
        key="poliza_actual",
        help="Sube el archivo de tu póliza actual para extraer sus datos",
    )

    st.markdown("---")

    # Campo 2: Póliza de Renovación
    st.subheader("2️⃣ Póliza de Renovación")
    archivo_poliza_renovacion = st.file_uploader(
        "Cargar archivo de póliza de renovación",
        type=["pdf", "docx", "txt", "jpg", "png"],
        key="poliza_renovacion",
        help="Sube el archivo de la póliza de renovación para comparar",
    )

    st.markdown("---")

    # Campo 3: Múltiples Documentos
    st.subheader("3️⃣ Documentos Adicionales")
    archivos_multiples = st.file_uploader(
        "Cargar múltiples documentos",
        type=["pdf", "docx", "txt", "jpg", "png"],
        accept_multiple_files=True,
        key="documentos_multiples",
        help="Sube múltiples documentos para extraer las primas de cada uno",
    )

    st.markdown("---")

    # Botón de procesamiento
if st.sidebar.button(
    "🚀 Iniciar Proceso",
    type="primary",
    use_container_width=True,
    help="Haz clic para iniciar el procesamiento de todos los archivos cargados",
):
    # Verificar que al menos un archivo esté cargado
    if (
        not archivo_poliza_actual
        and not archivo_poliza_renovacion
        and not archivos_multiples
    ):
        st.error("⚠️ Por favor, carga al menos un archivo antes de iniciar el proceso.")
    else:
        # Crear contenedor para mostrar el progreso
        with st.spinner("Procesando archivos..."):
            # Procesar archivos y crear cola
            queue_items = process_files_and_create_queue(
                archivo_poliza_actual, archivo_poliza_renovacion, archivos_multiples
            )

            # st.write(queue_items)

            if queue_items:
                st.success(
                    f"✅ Se procesaron {len(queue_items)} archivo(s) exitosamente."
                )

                # Mostrar información de la cola
                st.subheader("📋 Cola de Procesamiento")

                # Crear DataFrame para mostrar la información
                queue_data = []
                for item in queue_items:
                    queue_data.append(
                        {
                            "Archivo": item["file_name"],
                            "Tipo": item["doc_type"].title(),
                            "Extensión": item["file_extension"].upper(),
                            "Process ID": item["process_id"][:8] + "...",
                            "Estado": "✅ Listo para procesar",
                        }
                    )

                df_queue = pd.DataFrame(queue_data)
                st.dataframe(df_queue, use_container_width=True)

                # Sección de archivos descargados
                # st.subheader("📁 Archivos Descargados")

                results = []

                for item in queue_items:
                    if item["media_type"] == "application/pdf":
                        result = asyncio.run(orchestrator.handle_pdf(item))
                        results.append(result)
                        # st.write(result)

                # Procesar y mostrar resultados
                if results:
                    st.subheader("📊 Resultados del Análisis")
                    with st.expander("Results"):
                        st.write(results)

                    # Función para limpiar archivos automáticamente después del procesamiento
                    def cleanup_processed_files(queue_items_list):
                        """Elimina automáticamente los archivos después del procesamiento exitoso"""
                        deleted_files = []
                        failed_deletions = []

                        for item in queue_items_list:
                            try:
                                if os.path.exists(item["file_path"]):
                                    os.remove(item["file_path"])
                                    deleted_files.append(item["file_name"])
                                    app_logger.info(
                                        f"Archivo eliminado automáticamente: {item['file_name']}"
                                    )
                            except Exception as e:
                                failed_deletions.append((item["file_name"], str(e)))
                                app_logger.error(
                                    f"Error al eliminar archivo {item['file_name']}: {str(e)}"
                                )

                        return deleted_files, failed_deletions

                    # Función para convertir JSON a clases dataclass
                    def procesar_resultados(results_list):
                        polizas_data = []
                        primas_data = []

                        for result in results_list:
                            if result and "data" in result:
                                try:
                                    # Extraer información del archivo
                                    file_name = result.get(
                                        "file_name", "Archivo desconocido"
                                    )
                                    doc_type = result.get("doc_type", "desconocido")

                                    # Extraer el texto JSON del resultado
                                    candidates = result["data"].get("candidates", [])
                                    if candidates and len(candidates) > 0:
                                        content = candidates[0].get("content", {})
                                        parts = content.get("parts", [])
                                        if parts and len(parts) > 0:
                                            json_text = parts[0].get("text", "")

                                            # Parsear el JSON
                                            data = json.loads(json_text)

                                            # Procesar según el tipo de documento
                                            if (
                                                doc_type in ["actual", "renovacion"]
                                                and "detalle_cobertura" in data
                                            ):
                                                # Crear instancia de Cobertura
                                                cobertura_items = []
                                                for item_data in data[
                                                    "detalle_cobertura"
                                                ]:
                                                    cobertura_items.append(
                                                        DetalleCoberturaItem(
                                                            interes_asegurado=item_data[
                                                                "interes_asegurado"
                                                            ],
                                                            valor_asegurado=item_data[
                                                                "valor_asegurado"
                                                            ],
                                                        )
                                                    )

                                                cobertura = Cobertura(
                                                    detalle_cobertura=cobertura_items,
                                                    total_valores_asegurados=data.get(
                                                        "total_valores_asegurados", 0
                                                    ),
                                                )

                                                # Agregar a datos de pólizas
                                                for (
                                                    item_cob
                                                ) in cobertura.detalle_cobertura:
                                                    polizas_data.append(
                                                        {
                                                            "Archivo": file_name,
                                                            "Tipo de Documento": doc_type.title(),
                                                            "Interés Asegurado": item_cob.interes_asegurado,
                                                            "Valor Asegurado": item_cob.valor_asegurado,
                                                            "Total Póliza": cobertura.total_valores_asegurados,
                                                        }
                                                    )

                                            elif (
                                                doc_type == "adicional"
                                                and "prima_sin_iva" in data
                                                and "iva" in data
                                                and "prima_con_iva" in data
                                            ):
                                                # Crear instancia de Prima
                                                prima = Prima(
                                                    prima_sin_iva=data["prima_sin_iva"],
                                                    iva=data["iva"],
                                                    prima_con_iva=data["prima_con_iva"],
                                                )

                                                # Agregar a datos de primas
                                                primas_data.append(
                                                    {
                                                        "Archivo": file_name,
                                                        "Tipo de Documento": doc_type.title(),
                                                        "Prima Sin IVA": prima.prima_sin_iva,
                                                        "IVA": prima.iva,
                                                        "Prima Con IVA": prima.prima_con_iva,
                                                    }
                                                )

                                except Exception as e:
                                    st.error(
                                        f"Error procesando {result.get('file_name', 'archivo')}: {str(e)}"
                                    )

                        return polizas_data, primas_data

                    # Procesar todos los resultados
                    polizas_data, primas_data = procesar_resultados(results)

                    with st.expander("Ver Polizas"):
                        st.write(polizas_data)
                    with st.expander("Ver Primas"):
                        st.write(primas_data)

                    amparos_y_riesgos_text = (
                        results[0]
                        .get("data", {})
                        .get("candidates", [{}])[0]
                        .get("content", {})
                        .get("parts", [{}])[0]
                        .get("text", "{}")
                    )
                    if amparos_y_riesgos_text:
                        amparos_y_riesgos = json.loads(amparos_y_riesgos_text)
                        amparos = amparos_y_riesgos["amparos"]
                        riesgos = amparos_y_riesgos["riesgos"]
                        with st.expander("Ver Amparos"):
                            st.write(amparos)
                        with st.expander("Ver Riesgos"):
                            st.write(riesgos)

                    # Separar datos por tipo de documento
                    if polizas_data:
                        solo_intereses = [
                            {
                                "interes_asegurado": item["Interés Asegurado"],
                                "valor_asegurado": item["Valor Asegurado"],
                            }
                            for item in polizas_data
                            if item["Tipo de Documento"].lower() == "actual"
                        ]
                        solo_renovacion = [
                            {
                                "interes_asegurado": item["Interés Asegurado"],
                                "valor_asegurado": item["Valor Asegurado"],
                            }
                            for item in polizas_data
                            if item["Tipo de Documento"].lower() == "renovacion"
                        ]
                        df_polizas = pd.DataFrame(polizas_data)

                        # Filtrar por tipo de documento
                        polizas_actuales = df_polizas[
                            df_polizas["Tipo de Documento"] == "Actual"
                        ]
                        polizas_renovacion = df_polizas[
                            df_polizas["Tipo de Documento"] == "Renovacion"
                        ]

                        # Tabla para Pólizas Actuales
                        if not polizas_actuales.empty:
                            st.subheader("📋 Pólizas Actuales")
                            df_actuales_display = polizas_actuales.copy()
                            df_actuales_display["Valor Asegurado"] = (
                                df_actuales_display["Valor Asegurado"].apply(
                                    lambda x: f"${x:,.0f}"
                                )
                            )
                            df_actuales_display["Total Póliza"] = df_actuales_display[
                                "Total Póliza"
                            ].apply(lambda x: f"${x:,.0f}")
                            st.dataframe(solo_intereses, use_container_width=True)

                            # Métrica del total
                            total_actual = (
                                polizas_actuales["Total Póliza"].iloc[0]
                                if len(polizas_actuales) > 0
                                else 0
                            )
                            st.metric("💰 Total Póliza Actual", f"${total_actual:,.0f}")
                            st.markdown("---")

                        # Tabla para Pólizas de Renovación
                        if not polizas_renovacion.empty:
                            # with st.expander("Poliza Renovacion"):
                            #     st.write(
                            #         polizas_renovacion.drop_duplicates(
                            #             subset=["Archivo", "Tipo de Documento"]
                            #         )
                            #     )
                            st.subheader("🔄 Pólizas de Renovación")
                            df_renovacion_display = polizas_renovacion.copy()
                            df_renovacion_display["Valor Asegurado"] = (
                                df_renovacion_display["Valor Asegurado"].apply(
                                    lambda x: f"${x:,.0f}"
                                )
                            )
                            df_renovacion_display["Total Póliza"] = (
                                df_renovacion_display["Total Póliza"].apply(
                                    lambda x: f"${x:,.0f}"
                                )
                            )
                            st.dataframe(solo_renovacion, use_container_width=True)

                            # Métrica del total
                            total_renovacion = (
                                polizas_renovacion["Total Póliza"].iloc[0]
                                if len(polizas_renovacion) > 0
                                else 0
                            )
                            st.metric(
                                "💰 Total Póliza Renovación",
                                f"${total_renovacion:,.0f}",
                            )

                            # Comparación si hay ambas pólizas
                            if (
                                not polizas_actuales.empty
                                and not polizas_renovacion.empty
                            ):
                                diferencia = total_renovacion - total_actual
                                porcentaje_cambio = (
                                    ((diferencia / total_actual) * 100)
                                    if total_actual > 0
                                    else 0
                                )
                                # st.metric(
                                #     "📊 Diferencia (Renovación vs Actual)",
                                #     f"${diferencia:,.0f}",
                                #     f"{porcentaje_cambio:+.1f}%",
                                # )
                            st.markdown("---")

                    # Tabla para Documentos Adicionales (Primas)
                    if primas_data:
                        st.subheader("📄 Documentos Adicionales (Primas)")
                        solo_primas = [
                            {
                                "Archivo": item["Archivo"],
                                "Prima Sin IVA": item["Prima Sin IVA"],
                                "IVA": item["IVA"],
                                "Prima Con IVA": item["Prima Con IVA"],
                            }
                            for item in primas_data
                        ]
                        df_primas = pd.DataFrame(primas_data)

                        # Formatear valores monetarios para visualización
                        df_primas_display = df_primas.copy()
                        df_primas_display["Prima Sin IVA"] = df_primas_display[
                            "Prima Sin IVA"
                        ].apply(lambda x: f"${x:,.0f}")
                        df_primas_display["IVA"] = df_primas_display["IVA"].apply(
                            lambda x: f"${x:,.0f}"
                        )
                        df_primas_display["Prima Con IVA"] = df_primas_display[
                            "Prima Con IVA"
                        ].apply(lambda x: f"${x:,.0f}")

                        st.dataframe(solo_primas, use_container_width=True)

                        # Mostrar totales de primas
                        total_prima_sin_iva = df_primas["Prima Sin IVA"].sum()
                        total_iva = df_primas["IVA"].sum()
                        total_prima_con_iva = df_primas["Prima Con IVA"].sum()

                        st.markdown("---")

                    if amparos:
                        # Sección de Amparos con diseño profesional y simple
                        st.markdown(
                            """
                        <div style="background: linear-gradient(90deg, #2c5aa0 0%, #1e3a8a 100%); 
                                    padding: 1rem; border-radius: 8px; margin: 1rem 0;">
                            <h3 style="color: white; margin: 0; text-align: center;">🛡️ Amparos de la Póliza</h3>
                        </div>
                        """,
                            unsafe_allow_html=True,
                        )

                        # Crear DataFrame para mostrar los amparos de forma profesional
                        amparos_display = []
                        for i, amparo_item in enumerate(amparos, 1):
                            amparos_display.append(
                                {
                                    # "N°": i,
                                    "Tipo de Amparo": amparo_item.get(
                                        "amparo", "No especificado"
                                    ),
                                    "Deducible": amparo_item.get(
                                        "deducible", "No especificado"
                                    ),
                                }
                            )

                        if amparos_display:
                            df_amparos = pd.DataFrame(amparos_display)

                            # Mostrar métricas rápidas
                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Total de Amparos", len(amparos_display))
                            with col2:
                                amparos_con_deducible = sum(
                                    1
                                    for item in amparos_display
                                    if item["Deducible"] != "No especificado"
                                )
                                st.metric(
                                    "Con Deducible Definido", amparos_con_deducible
                                )

                            # Tabla profesional
                            st.dataframe(
                                df_amparos,
                                use_container_width=True,
                                hide_index=True,
                                column_config={
                                    "N°": st.column_config.NumberColumn(
                                        "N°", width="small"
                                    ),
                                    "Tipo de Amparo": st.column_config.TextColumn(
                                        "Tipo de Amparo", width="large"
                                    ),
                                    "Deducible": st.column_config.TextColumn(
                                        "Deducible", width="medium"
                                    ),
                                },
                            )
                        else:
                            st.info("No se encontraron amparos en el análisis.")

                        st.markdown("---")

                    if riesgos:
                        # Encabezado con diseño profesional
                        st.markdown(
                            """
                        <div style="background: linear-gradient(90deg, #dc2626 0%, #b91c1c 100%); 
                                    padding: 1rem; border-radius: 8px; margin: 1rem 0;">
                            <h3 style="color: white; margin: 0; text-align: center;">⚠️ Análisis de Riesgos por Ubicación</h3>
                        </div>
                        """,
                            unsafe_allow_html=True,
                        )

                        try:
                            # Métricas rápidas
                            total_ubicaciones = (
                                len(riesgos) if isinstance(riesgos, list) else 0
                            )
                            total_coberturas = 0
                            total_valor_asegurado = 0

                            if isinstance(riesgos, list):
                                for ubicacion in riesgos:
                                    if (
                                        isinstance(ubicacion, dict)
                                        and "detalle_cobertura" in ubicacion
                                    ):
                                        detalle = ubicacion["detalle_cobertura"]
                                        if isinstance(detalle, list):
                                            total_coberturas += len(detalle)
                                            for item in detalle:
                                                if (
                                                    isinstance(item, dict)
                                                    and "valor_asegurado" in item
                                                ):
                                                    try:
                                                        total_valor_asegurado += float(
                                                            item["valor_asegurado"]
                                                        )
                                                    except (ValueError, TypeError):
                                                        pass

                            # Mostrar métricas
                            # col1, col2, col3 = st.columns(3)
                            # with col1:
                            #     st.metric("🏢 Total Ubicaciones", total_ubicaciones)
                            # with col2:
                            #     st.metric("📋 Total Coberturas", total_coberturas)
                            # with col3:
                            #     st.metric(
                            #         "💰 Valor Total Asegurado",
                            #         f"${total_valor_asegurado:,.0f}",
                            #     )

                            # st.markdown("---")

                            # Mostrar cada ubicación con su detalle
                            if isinstance(riesgos, list):
                                for i, ubicacion_data in enumerate(riesgos, 1):
                                    if isinstance(ubicacion_data, dict):
                                        ubicacion = ubicacion_data.get(
                                            "ubicacion", f"Ubicación {i}"
                                        )
                                        detalle_cobertura = ubicacion_data.get(
                                            "detalle_cobertura", []
                                        )

                                        # Card para cada ubicación
                                        with st.expander(
                                            f"📍 {ubicacion}", expanded=True
                                        ):
                                            # st.markdown(
                                            #     f"""
                                            # <div style="background: #f8fafc; padding: 1rem; border-radius: 6px; border-left: 4px solid #dc2626;">
                                            #     <h4 style="margin: 0; color: #1f2937;">🏢 {ubicacion}</h4>
                                            # </div>
                                            # """,
                                            #     unsafe_allow_html=True,
                                            # )

                                            if (
                                                isinstance(detalle_cobertura, list)
                                                and detalle_cobertura
                                            ):
                                                # Crear DataFrame para la tabla
                                                tabla_data = []
                                                for j, item in enumerate(
                                                    detalle_cobertura, 1
                                                ):
                                                    if isinstance(item, dict):
                                                        interes = item.get(
                                                            "interes_asegurado",
                                                            "No especificado",
                                                        )
                                                        valor = item.get(
                                                            "valor_asegurado", 0
                                                        )
                                                        riesgos_list = item.get(
                                                            "riesgos", []
                                                        )

                                                        # Formatear la lista de riesgos
                                                        if isinstance(
                                                            riesgos_list, list
                                                        ):
                                                            riesgos_str = (
                                                                ", ".join(riesgos_list)
                                                                if riesgos_list
                                                                else "No especificados"
                                                            )
                                                        else:
                                                            riesgos_str = (
                                                                str(riesgos_list)
                                                                if riesgos_list
                                                                else "No especificados"
                                                            )

                                                        tabla_data.append(
                                                            {
                                                                # "N°": j,
                                                                "Interés Asegurado": interes,
                                                                "Valor Asegurado": (
                                                                    f"${float(valor):,.0f}"
                                                                    if valor
                                                                    else "$0"
                                                                ),
                                                                # "Riesgos Cubiertos": riesgos_str,
                                                            }
                                                        )

                                                if tabla_data:
                                                    df_riesgos = pd.DataFrame(
                                                        tabla_data
                                                    )
                                                    st.dataframe(
                                                        df_riesgos,
                                                        use_container_width=True,
                                                        hide_index=True,
                                                        column_config={
                                                            # "N°": st.column_config.NumberColumn(
                                                            #     "N°", width="small"
                                                            # ),
                                                            "Interés Asegurado": st.column_config.TextColumn(
                                                                "Interés Asegurado",
                                                                width="medium",
                                                            ),
                                                            "Valor Asegurado": st.column_config.TextColumn(
                                                                "Valor Asegurado",
                                                                width="medium",
                                                            ),
                                                            # "Riesgos Cubiertos": st.column_config.TextColumn(
                                                            #     "Riesgos Cubiertos",
                                                            #     width="large",
                                                            # ),
                                                        },
                                                    )

                                                    # Resumen de la ubicación
                                                    # total_ubicacion = sum(
                                                    #     float(item.get("valor_asegurado", 0))
                                                    #     for item in detalle_cobertura
                                                    #     if isinstance(item, dict)
                                                    # )
                                                    # st.info(f"💼 **Total asegurado en esta ubicación:** ${total_ubicacion:,.0f}")
                                                else:
                                                    st.warning(
                                                        "No se encontraron detalles de cobertura válidos para esta ubicación."
                                                    )
                                            else:
                                                st.warning(
                                                    "No se encontraron detalles de cobertura para esta ubicación."
                                                )

                                            # st.markdown("---")
                            else:
                                st.warning(
                                    "Los datos de riesgos no tienen el formato esperado (lista de ubicaciones)."
                                )

                        except json.JSONDecodeError:
                            st.error(
                                "Error al procesar los datos de riesgos: formato JSON inválido."
                            )
                        except KeyError as e:
                            st.error(
                                f"Error al acceder a los datos de riesgos: clave faltante {e}"
                            )
                        except Exception as e:
                            st.error(f"Error inesperado al procesar riesgos: {str(e)}")
                            st.info("Mostrando datos sin procesar:")
                            st.write(riesgos)

                        st.markdown("---")

                    # Generar archivo Excel para descarga
                    if polizas_data or primas_data:
                        st.subheader("📥 Descargar Resultados")

                        # Crear archivo Excel en memoria con datos filtrados y formato profesional
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine="openpyxl") as writer:
                            from openpyxl.styles import (
                                Font,
                                PatternFill,
                                Border,
                                Side,
                                Alignment,
                            )
                            from openpyxl.utils.dataframe import dataframe_to_rows
                            from openpyxl.utils import get_column_letter

                            # Definir estilos
                            header_font = Font(
                                name="Calibri", size=12, bold=True, color="FFFFFF"
                            )
                            header_fill = PatternFill(
                                start_color="2E75B6",
                                end_color="2E75B6",
                                fill_type="solid",
                            )
                            data_font = Font(name="Calibri", size=11)
                            border = Border(
                                left=Side(style="thin", color="D0D0D0"),
                                right=Side(style="thin", color="D0D0D0"),
                                top=Side(style="thin", color="D0D0D0"),
                                bottom=Side(style="thin", color="D0D0D0"),
                            )
                            center_alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                            currency_alignment = Alignment(
                                horizontal="right", vertical="center"
                            )
                            left_alignment = Alignment(
                                horizontal="left", vertical="center"
                            )

                            def format_worksheet(ws, df, sheet_type):
                                # Insertar fila para el título
                                ws.insert_rows(1, 2)  # Insertar 2 filas al inicio

                                # Agregar título en la primera fila
                                title_cell = ws.cell(row=1, column=1)
                                title_cell.value = (
                                    "CENTRO DE DIAGNOSTICO AUTOMOTOR DE PALMIRA"
                                )

                                # Formato del título
                                from openpyxl.styles import Font, Alignment

                                title_font = Font(
                                    name="Arial", size=16, bold=True, color="FFFFFF"
                                )
                                title_fill = PatternFill(
                                    start_color="1F4E79",
                                    end_color="1F4E79",
                                    fill_type="solid",
                                )
                                title_alignment = Alignment(
                                    horizontal="center", vertical="center"
                                )

                                title_cell.font = title_font
                                title_cell.fill = title_fill
                                title_cell.alignment = title_alignment

                                # Combinar celdas para el título (toda la fila)
                                ws.merge_cells(
                                    start_row=1,
                                    start_column=1,
                                    end_row=1,
                                    end_column=len(df.columns),
                                )

                                # Formato especial para etiquetas de pólizas consolidadas
                                if sheet_type == "polizas":
                                    label_font = Font(
                                        name="Arial", size=14, bold=True, color="FFFFFF"
                                    )
                                    label_fill = PatternFill(
                                        start_color="4472C4",
                                        end_color="4472C4",
                                        fill_type="solid",
                                    )

                                    # Buscar y formatear etiquetas de pólizas
                                    for row_num in range(3, ws.max_row + 1):
                                        cell_value = ws.cell(
                                            row=row_num, column=1
                                        ).value
                                        if cell_value and (
                                            "PÓLIZAS ACTUALES" in str(cell_value)
                                            or "PÓLIZAS DE RENOVACIÓN"
                                            in str(cell_value)
                                        ):
                                            # Aplicar formato a la etiqueta
                                            label_cell = ws.cell(row=row_num, column=1)
                                            label_cell.font = label_font
                                            label_cell.fill = label_fill
                                            label_cell.alignment = title_alignment

                                            # Combinar celdas para la etiqueta
                                            ws.merge_cells(
                                                start_row=row_num,
                                                start_column=1,
                                                end_row=row_num,
                                                end_column=len(df.columns),
                                            )

                                # Aplicar formato a encabezados (ahora en la fila 3)
                                for col_num in range(1, len(df.columns) + 1):
                                    cell = ws.cell(row=3, column=col_num)
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.border = border
                                    cell.alignment = center_alignment

                                # Aplicar formato a datos (ahora empezando desde la fila 4)
                                for row_num in range(4, len(df) + 4):
                                    for col_num in range(1, len(df.columns) + 1):
                                        cell = ws.cell(row=row_num, column=col_num)
                                        cell.font = data_font
                                        cell.border = border

                                        # Aplicar alineación según el tipo de columna
                                        col_name = df.columns[col_num - 1]
                                        if (
                                            "valor" in col_name.lower()
                                            or "prima" in col_name.lower()
                                            or "iva" in col_name.lower()
                                        ):
                                            cell.alignment = currency_alignment
                                            # Formatear como moneda
                                            if isinstance(cell.value, (int, float)):
                                                cell.number_format = "$#,##0"
                                        elif col_name.lower() == "coberturas":
                                            cell.alignment = left_alignment
                                        else:
                                            cell.alignment = center_alignment

                                # Ajustar ancho de columnas
                                for col_num in range(1, len(df.columns) + 1):
                                    column_letter = get_column_letter(col_num)
                                    col_name = df.columns[col_num - 1]

                                    # Calcular ancho basado en el contenido (incluyendo el título)
                                    max_length = max(
                                        len(str(col_name)),
                                        len(
                                            "CENTRO DE DIAGNOSTICO AUTOMOTOR DE PALMIRA"
                                        )
                                        // len(df.columns),
                                    )
                                    for row_num in range(4, len(df) + 4):
                                        cell_value = str(
                                            ws.cell(row=row_num, column=col_num).value
                                            or ""
                                        )
                                        max_length = max(max_length, len(cell_value))

                                    # Establecer ancho específico para la columna Coberturas
                                    if col_name.lower() == "coberturas":
                                        width = 45  # Ancho específico para coberturas
                                    else:
                                        # Establecer ancho mínimo y máximo para otras columnas
                                        width = min(max(max_length + 2, 15), 30)

                                    ws.column_dimensions[column_letter].width = width

                                # Aplicar filtros automáticos
                                ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

                                # Congelar primera fila
                                ws.freeze_panes = "A2"

                            # Crear primero la hoja "Análisis_Estructurado" para que sea la primera pestaña
                            if solo_intereses and solo_renovacion and primas_data:
                                # Crear estructura organizada con 4 columnas específicas

                                # Obtener datos únicos de intereses asegurados
                                intereses_unicos = list(
                                    set(
                                        [
                                            item["interes_asegurado"]
                                            for item in solo_intereses
                                        ]
                                        + [
                                            item["interes_asegurado"]
                                            for item in solo_renovacion
                                        ]
                                    )
                                )

                                # Crear diccionarios para búsqueda rápida
                                valores_actuales = {
                                    item["interes_asegurado"]: item["valor_asegurado"]
                                    for item in solo_intereses
                                }
                                valores_renovacion = {
                                    item["interes_asegurado"]: item["valor_asegurado"]
                                    for item in solo_renovacion
                                }

                                # Calcular totales
                                total_actual = sum(valores_actuales.values())
                                total_renovacion = sum(valores_renovacion.values())

                                # Crear datos estructurados
                                datos_estructurados = []

                                # Agregar filas de intereses asegurados
                                for interes in sorted(intereses_unicos):
                                    datos_estructurados.append(
                                        {
                                            "Interés Asegurado": interes,
                                            "Valor Asegurado Actual": valores_actuales.get(
                                                interes, 0
                                            ),
                                            "Valor Asegurado Renovado": valores_renovacion.get(
                                                interes, 0
                                            ),
                                            "Primas": "",
                                        }
                                    )

                                # Agregar fila de totales
                                datos_estructurados.append(
                                    {
                                        "Interés Asegurado": "TOTAL",
                                        "Valor Asegurado Actual": total_actual,
                                        "Valor Asegurado Renovado": total_renovacion,
                                        "Primas": "",
                                    }
                                )

                                # Crear estructura base con columnas dinámicas para primas
                                # Definir lista de coberturas
                                coberturas = [
                                    "Incendio y/o Rayo Edificios",
                                    "Explosión Mejoras Locativas",
                                    "Terremoto, temblor Muebles y Enseres",
                                    "Asonada, motin, conm. Civil/popular huelga Mercancias Fijas",
                                    "Extension de amparo",
                                    "Daños por agua, Anegación Dineros",
                                    "Incendio y/o Rayo en aparatos electricos Equipo Electronico",
                                ]

                                columnas_base = [
                                    "Coberturas",
                                    "Interés Asegurado",
                                    "Valor Asegurado Actual",
                                    "Valor Asegurado Renovado",
                                ]

                                # Agregar columnas para cada archivo de prima
                                columnas_primas = []
                                if solo_primas:
                                    for prima in solo_primas:
                                        nombre_archivo = prima["Archivo"]
                                        if nombre_archivo not in columnas_primas:
                                            columnas_primas.append(nombre_archivo)

                                # Crear todas las columnas
                                todas_columnas = columnas_base + columnas_primas

                                # Inicializar datos estructurados con todas las columnas
                                datos_estructurados = []

                                # Crear mapas de valores de prima por archivo
                                valores_prima_por_archivo = {}
                                if solo_primas:
                                    for prima in solo_primas:
                                        nombre_archivo = prima["Archivo"]
                                        valores_prima_por_archivo[nombre_archivo] = [
                                            f"${prima['Prima Sin IVA']:,.0f}",
                                        ]

                                # Determinar el número máximo de filas necesarias
                                intereses_ordenados = sorted(intereses_unicos)
                                max_filas = max(
                                    len(intereses_ordenados), len(coberturas)
                                )

                                # Agregar filas con coberturas e intereses asegurados
                                for i in range(max_filas):
                                    fila = {
                                        "Coberturas": (
                                            coberturas[i] if i < len(coberturas) else ""
                                        ),
                                        "Interés Asegurado": (
                                            intereses_ordenados[i]
                                            if i < len(intereses_ordenados)
                                            else ""
                                        ),
                                        "Valor Asegurado Actual": (
                                            valores_actuales.get(
                                                intereses_ordenados[i], 0
                                            )
                                            if i < len(intereses_ordenados)
                                            else ""
                                        ),
                                        "Valor Asegurado Renovado": (
                                            valores_renovacion.get(
                                                intereses_ordenados[i], 0
                                            )
                                            if i < len(intereses_ordenados)
                                            else ""
                                        ),
                                    }

                                    # Agregar valores de prima en las primeras filas
                                    for col_prima in columnas_primas:
                                        if (
                                            col_prima in valores_prima_por_archivo
                                            and i < 1
                                        ):
                                            fila[col_prima] = valores_prima_por_archivo[
                                                col_prima
                                            ][i]
                                        else:
                                            fila[col_prima] = ""

                                    datos_estructurados.append(fila)

                                # Agregar fila de totales
                                fila_total = {
                                    "Coberturas": "",
                                    "Interés Asegurado": "TOTAL",
                                    "Valor Asegurado Actual": total_actual,
                                    "Valor Asegurado Renovado": total_renovacion,
                                }
                                # Inicializar columnas de primas vacías para totales
                                for col_prima in columnas_primas:
                                    fila_total[col_prima] = ""
                                datos_estructurados.append(fila_total)

                                # Crear DataFrame estructurado con columnas ordenadas
                                df_estructurado = pd.DataFrame(
                                    datos_estructurados, columns=todas_columnas
                                )

                                # Validación de integridad de datos
                                if (
                                    abs(
                                        total_actual
                                        - sum(
                                            item["valor_asegurado"]
                                            for item in solo_intereses
                                        )
                                    )
                                    > 0.01
                                ):
                                    st.warning(
                                        "⚠️ Advertencia: Discrepancia detectada en totales actuales"
                                    )

                                if (
                                    abs(
                                        total_renovacion
                                        - sum(
                                            item["valor_asegurado"]
                                            for item in solo_renovacion
                                        )
                                    )
                                    > 0.01
                                ):
                                    st.warning(
                                        "⚠️ Advertencia: Discrepancia detectada en totales de renovación"
                                    )

                                # Exportar a Excel
                                df_estructurado.to_excel(
                                    writer,
                                    sheet_name="Análisis_Estructurado",
                                    index=False,
                                )
                                format_worksheet(
                                    writer.sheets["Análisis_Estructurado"],
                                    df_estructurado,
                                    "polizas",
                                )

                            # Crear las otras hojas después de Análisis_Estructurado
                            if polizas_data:
                                df_polizas = pd.DataFrame(polizas_data)

                                # Separar por tipo de documento
                                df_actuales = df_polizas[
                                    df_polizas["Tipo de Documento"].str.lower()
                                    == "actual"
                                ].copy()
                                df_renovacion = df_polizas[
                                    df_polizas["Tipo de Documento"].str.lower()
                                    == "renovacion"
                                ].copy()

                                # Función para crear tabla transpuesta consolidada
                                def crear_hoja_consolidada(df_actuales, df_renovacion):
                                    if df_actuales.empty and df_renovacion.empty:
                                        return

                                    # Función auxiliar para procesar cada tipo de póliza
                                    def procesar_poliza(df, tipo_poliza):
                                        if df.empty:
                                            return None, None

                                        # Obtener intereses únicos y sus valores
                                        intereses_valores = {}
                                        total_poliza = 0

                                        for _, row in df.iterrows():
                                            interes = row["Interés Asegurado"]
                                            valor = row["Valor Asegurado"]
                                            total_poliza = row["Total Póliza"]
                                            intereses_valores[interes] = valor

                                        # Crear estructura transpuesta
                                        columnas = list(intereses_valores.keys()) + [
                                            "Total Póliza"
                                        ]
                                        valores = [
                                            f"${valor:,.0f}"
                                            for valor in intereses_valores.values()
                                        ] + [f"${total_poliza:,.0f}"]

                                        return columnas, valores

                                    # Procesar ambos tipos de pólizas
                                    columnas_actuales, valores_actuales = (
                                        procesar_poliza(df_actuales, "Actual")
                                    )
                                    columnas_renovacion, valores_renovacion = (
                                        procesar_poliza(df_renovacion, "Renovación")
                                    )

                                    # Crear DataFrame consolidado
                                    datos_consolidados = []

                                    # Agregar etiqueta y datos de pólizas actuales
                                    if columnas_actuales and valores_actuales:
                                        # Fila de etiqueta para pólizas actuales
                                        fila_etiqueta_actual = ["PÓLIZAS ACTUALES"] + [
                                            ""
                                        ] * (len(columnas_actuales) - 1)
                                        datos_consolidados.append(fila_etiqueta_actual)

                                        # Fila de encabezados
                                        datos_consolidados.append(columnas_actuales)

                                        # Fila de valores
                                        datos_consolidados.append(valores_actuales)

                                        # Fila de separación
                                        datos_consolidados.append(
                                            [""] * len(columnas_actuales)
                                        )

                                    # Agregar etiqueta y datos de pólizas de renovación
                                    if columnas_renovacion and valores_renovacion:
                                        # Ajustar longitud de columnas para que coincidan
                                        max_cols = max(
                                            (
                                                len(columnas_actuales)
                                                if columnas_actuales
                                                else 0
                                            ),
                                            len(columnas_renovacion),
                                        )

                                        # Fila de etiqueta para pólizas de renovación
                                        fila_etiqueta_renovacion = [
                                            "PÓLIZAS DE RENOVACIÓN"
                                        ] + [""] * (max_cols - 1)
                                        datos_consolidados.append(
                                            fila_etiqueta_renovacion
                                        )

                                        # Ajustar columnas de renovación si es necesario
                                        columnas_renovacion_ajustadas = (
                                            columnas_renovacion
                                            + [""]
                                            * (max_cols - len(columnas_renovacion))
                                        )
                                        valores_renovacion_ajustados = (
                                            valores_renovacion
                                            + [""]
                                            * (max_cols - len(valores_renovacion))
                                        )

                                        # Fila de encabezados
                                        datos_consolidados.append(
                                            columnas_renovacion_ajustadas
                                        )

                                        # Fila de valores
                                        datos_consolidados.append(
                                            valores_renovacion_ajustados
                                        )

                                    # Crear DataFrame final
                                    if datos_consolidados:
                                        max_cols = max(
                                            len(fila) for fila in datos_consolidados
                                        )

                                        # Ajustar todas las filas a la misma longitud
                                        for i, fila in enumerate(datos_consolidados):
                                            if len(fila) < max_cols:
                                                datos_consolidados[i] = fila + [""] * (
                                                    max_cols - len(fila)
                                                )

                                        # Crear columnas genéricas
                                        columnas_genericas = [
                                            f"Columna_{i+1}" for i in range(max_cols)
                                        ]

                                        df_consolidado = pd.DataFrame(
                                            datos_consolidados,
                                            columns=columnas_genericas,
                                        )

                                        # Exportar a Excel
                                        df_consolidado.to_excel(
                                            writer,
                                            sheet_name="Polizas_Consolidadas",
                                            index=False,
                                            header=False,
                                        )
                                        format_worksheet(
                                            writer.sheets["Polizas_Consolidadas"],
                                            df_consolidado,
                                            "polizas",
                                        )

                                # Crear hoja consolidada
                                crear_hoja_consolidada(df_actuales, df_renovacion)

                            # Hoja de primas eliminada según solicitud del usuario

                        excel_data = output.getvalue()

                        # Botón de descarga
                        # Crear enlace de descarga que abre en nueva ventana
                        b64_excel = base64.b64encode(excel_data).decode()
                        href = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}"

                        # Botón con enlace que abre en nueva ventana y descarga automáticamente
                        st.markdown(
                            f"""
                            <a href="{href}" download="analisis_polizas.xlsx" target="_blank" 
                               style="display: inline-block; padding: 0.5rem 1rem; background-color: #ff4b4b; 
                                      color: white; text-decoration: none; border-radius: 0.25rem; 
                                      font-weight: 600; border: none; cursor: pointer;"
                               onclick="setTimeout(function(){{window.close();}}, 1000);">
                                📊 Descargar Excel
                            </a>
                            """,
                            unsafe_allow_html=True,
                        )

                        # Limpieza automática de archivos procesados
                        cleanup_processed_files(queue_items)

                else:
                    st.warning("No se encontraron resultados para mostrar.")
                    # Limpieza automática incluso si no hay resultados
                    cleanup_processed_files(queue_items)
                    st.info(
                        "🧹 Archivos eliminados automáticamente del directorio downloads."
                    )

            else:
                st.error(
                    "❌ No se pudieron procesar los archivos. Verifica que los archivos sean válidos."
                )
