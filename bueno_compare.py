import streamlit as st
from typing import TypedDict, Dict, List, Optional, Union
import os
import base64
import certifi
import aiohttp
import ssl
from tools_bueno import tools as tools_standard
import asyncio
import fitz  # PyMuPDF
from PIL import Image
import io
from jsonschema import validate, ValidationError
import json
import pandas as pd
from typing import Optional, List, TypedDict, cast
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


load_dotenv()


class ValoresAsegurables(TypedDict, total=False):
    # Valor asegurable de Edificios en caso de incendio, si está indicado.
    edificios: str

    # Valor asegurable de Muebles y Enseres en caso de incendio, si está indicado.
    muebles_y_enseres: str

    # Valor asegurable de Equipos Eléctricos y Electrónicos en caso de incendio, si está indicado.
    equipos_electricos_y_electronicos: str

    # Valor asegurable de Equipo Eléctrico y Electrónico Móvil y Portátil en caso de incendio, si está indicado.
    equipo_movil_y_portatil: str

    # Valor asegurable de Maquinaria y Equipo en caso de incendio, si está indicado.
    maquinaria_y_equipo: str

    # Valor asegurable de Mercancías Fijas en caso de incendio, si está indicado.
    mercancias_fijas: str

    # Valor asegurable de Dineros en caso de incendio, si está indicado.
    dineros: str

    # Valor asegurable de Obras de Arte en caso de incendio, si está indicado.
    obras_de_arte: str

    # Valor asegurable correspondiente a Asistencia en caso de incendio, si está indicado.
    asistencia: Optional[str]

    # Valor total asegurable indicado en el texto.
    total_valor_asegurable: str

    # TASA de DAÑOS MATERIALES indicada en el texto. Si no está presente, indicar explícitamente que no se encuentra.
    tasa_danos_materiales: str

    # PRIMA TOTAL indicada en el texto. Si no está presente, indicar explícitamente que no se encuentra.
    prima: str

    # LÍMITE POR DESPACHO indicado en el contexto de TRANSPORTE DE VALORES. Si no está presente, indicar explícitamente que no se encuentra.
    limite_por_despacho: str

    # PRESUPUESTO indicado en el contexto de TRANSPORTE DE VALORES. Si no está presente, indicar explícitamente que no se encuentra.
    presupuesto: str


class CondicionesCoberturasEspeciales(TypedDict, total=False):
    # Lista de condiciones y coberturas explícitas relacionadas con sustracción en general.
    # Incluye información sobre porcentajes, ubicación del bien, requisitos de seguridad, etc.
    # Si no hay información, indicar: 'No se encuentra información explícita sobre sustracción en general.'
    condiciones_sustraccion: List[str]

    # Condiciones específicas, coberturas u observaciones relacionadas con EQUIPO ELECTRÓNICO.
    # Incluye cualquier mención explícita sobre cobertura, exclusión o condición particular.
    # Si no hay información, indicar: 'No se encuentra información específica sobre equipo electrónico.'
    condiciones_equipo_electronico: List[str]

    # Observaciones o condiciones explícitas relacionadas con ROTURA DE MAQUINARIA.
    # Incluye excepciones, notas, comentarios relevantes.
    # Si no hay información, indicar: 'No se encuentra información específica sobre rotura de maquinaria.'
    observaciones_rotura_maquinaria: List[str]


class InfoArchivo(TypedDict):
    filename: str
    valores_asegurables: ValoresAsegurables
    condiciones_sustraccion: CondicionesCoberturasEspeciales


class QueueItem(TypedDict):
    file_name: str
    file_extension: str
    file_path: str
    media_type: str


class SeguroOrchestrator:
    def __init__(
        self,
        api_key: str,  # API key de Anthropic
        model: str,  # Modelo de Claude a usar
    ):
        self.api_key = api_key
        self.model = model

    def format_docs(self, respuestas: List[dict]) -> List[InfoArchivo]:
        result: List[InfoArchivo] = []

        for r in respuestas:
            filename = r.get("file_name", "desconocido.pdf")
            valores_asegurables = {}
            condiciones_sustraccion = {}

            for entry in r.get("data", []):
                for item in entry.get("content", []):
                    name = item.get("name")
                    if not name:
                        continue

                    if name == "valores_asegurables_incendio":
                        valores_asegurables = item.get("input", {})
                    elif name == "condiciones_sustraccion_y_observaciones":
                        input_data = item.get("input", {})
                        condiciones_sustraccion = {
                            "condiciones_equipo_electronico": input_data.get(
                                "condiciones_equipo_electronico", []
                            ),
                            "condiciones_sustraccion": input_data.get(
                                "condiciones_sustraccion", []
                            ),
                            "observaciones_rotura_maquinaria": input_data.get(
                                "observaciones_rotura_maquinaria", []
                            ),
                        }

            result.append(
                InfoArchivo(
                    filename=filename,
                    valores_asegurables=valores_asegurables,
                    condiciones_sustraccion=condiciones_sustraccion,
                )
            )

        return result

    def generate_excel_table(
        self,
        docs: List[InfoArchivo],
        title: str = "Arango Bueno",
        id: str = "CAMARA DE COMERCIO DE PALMIRA",
        subtitle: str = "CUADRO COMPARATIVO DE SEGUROS",
    ):
        # Orden deseado para valores asegurables
        valores_order = [
            "edificios",
            "muebles_y_enseres",
            "equipos_electricos_y_electronicos",
            "equipo_movil_y_portatil",
            "maquinaria_y_equipo",
            "mercancias_fijas",
            "dineros",
            "obras_de_arte",
            "asistencia",
            "total_valor_asegurable",
            "tasa_danos_materiales",
            "prima",
            "limite_por_despacho",
            "presupuesto",
        ]

        # Orden deseado para condiciones
        condiciones_order = [
            "condiciones_sustraccion",
            "condiciones_equipo_electronico",
            "observaciones_rotura_maquinaria",
        ]

        def format_value(value):
            if isinstance(value, list):
                return "\n".join(str(v) for v in value) if value else "—"
            elif value and str(value).strip():
                return str(value)
            else:
                return "—"

        data = {}
        for doc in docs:
            combined_fields = {}
            # Agregar valores asegurables
            combined_fields.update(doc.get("valores_asegurables", {}))
            # Agregar condiciones de sustracción
            combined_fields.update(doc.get("condiciones_sustraccion", {}))
            data[doc["filename"]] = combined_fields

        header = ["Concepto"] + list(data.keys())
        print("Header: ", header)
        table_data = []

        # Agregar sección de valores asegurables
        table_data.append(["VALORES ASEGURABLES"] + [""] * (len(header) - 1))
        for key in valores_order:
            if any(key in doc_data for doc_data in data.values()):
                row = [key.replace("_", " ").title()]
                for filename in data:
                    row.append(format_value(data[filename].get(key)))
                table_data.append(row)

        # Agregar sección de condiciones
        table_data.append(["CONDICIONES Y COBERTURAS"] + [""] * (len(header) - 1))
        for key in condiciones_order:
            if any(key in doc_data for doc_data in data.values()):
                row = [key.replace("_", " ").title()]
                for filename in data:
                    row.append(format_value(data[filename].get(key)))
                table_data.append(row)

        df = pd.DataFrame(table_data, columns=header)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Create a dummy DataFrame to write the header information
            header_df = pd.DataFrame([[""], [""], [""]])
            header_df.to_excel(
                writer, index=False, header=False, sheet_name="Comparador"
            )
            worksheet = writer.sheets["Comparador"]

            # Write title and ID
            worksheet["A1"] = title
            worksheet["B1"] = id
            # Write subtitle
            worksheet["A2"] = subtitle

            # Write the main DataFrame starting from row 4 (0-indexed, so row 3 in Excel)
            df.to_excel(writer, startrow=3, index=False, sheet_name="Comparador")

            # Ajustar ancho columnas y wrap text
            max_col_width = 50
            for col_idx, col in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                # Ancho máximo de contenido en la columna
                max_length = max(df[col].astype(str).map(len).max(), len(str(col)))
                width = min(max_length + 2, max_col_width)
                worksheet.column_dimensions[col_letter].width = width

                # Aplicar wrap text para toda la columna (incluyendo header)
                for row_idx in range(
                    1, len(df) + 2
                ):  # +1 para encabezado y otro +1 porque Excel empieza en 1
                    cell = worksheet[f"{col_letter}{row_idx}"]
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Ajustar altura de filas para que se vean saltos de línea
            for row_idx in range(1, len(df) + 2):
                # La altura base puede variar, aquí se multiplica por cantidad de líneas más 1 para espacio extra
                max_lines = 1
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet[f"{get_column_letter(col_idx)}{row_idx}"]
                    if cell.value:
                        # Contar líneas por saltos de línea
                        lines = str(cell.value).count("\n") + 1
                        if lines > max_lines:
                            max_lines = lines
                worksheet.row_dimensions[row_idx].height = max(15, max_lines * 15)

        output.seek(0)
        return output.getvalue()

    def show_excel_download(
        self, title: str, id: str, subtitle: str, docs: List[InfoArchivo]
    ):
        excel_bytes = self.generate_excel_table(
            docs,
            title,
            id,
            subtitle,
        )
        b64_excel = base64.b64encode(excel_bytes).decode("utf-8")
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}" download="comparador_table.xlsx" target="_blank">📄 Descargar Excel</a>'
        st.markdown(href, unsafe_allow_html=True)

    async def make_api_request(
        self, url: str, headers: dict, data: dict, retries: int = 5
    ) -> Optional[Dict]:
        ssl_context = ssl.create_default_context(cafile=certifi.where())
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        async with aiohttp.ClientSession(connector=connector) as session:
            for i in range(retries):
                try:
                    async with session.post(
                        url, headers=headers, json=data
                    ) as response:
                        if response.status == 200:
                            return await response.json()
                        elif response.status in [429, 529, 503]:
                            print(response)
                            x = await response.json()
                            print(x)
                            sleep_time = 15 * (i + 1)  # Espera incremental en segundos
                            print(
                                f"API request failed with status {response.status}. Retrying in {sleep_time} seconds..."
                            )
                            await asyncio.sleep(sleep_time)
                        else:
                            print(
                                f"API request failed with status {response.status} - {await response.text()}"
                            )
                            raise ValueError(
                                f"Request failed with status {response.status}"
                            )
                except aiohttp.ClientError as e:
                    raise ValueError(f"Request error: {str(e)}")
        raise ValueError("Max retries exceeded.")

    async def tool_handler(
        self,
        tools: list,
        messages: list,
        tool_name: str,
        url: str = "https://generativelanguage.googleapis.com/v1beta/openai/chat/completions",
        model: str = "gemini-2.5-flash",
        max_retries: int = 6,
    ):
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {self.api_key}",
        }
        data = {
            "model": model,
            "messages": messages,
            "tools": tools,
            "tool_choice": {
                "type": "function",
                "function": {"name": tool_name},
            },
        }
        schema = next(
            (
                tool["function"]["parameters"]
                for tool in tools
                if tool["function"]["name"] == tool_name
            ),
            None,
        )
        tool_output = None
        for attempt in range(0, max_retries):
            try:
                response = await self.make_api_request(
                    url=url,
                    headers=headers,
                    data=data,
                )

                if response["choices"][0]["message"]["tool_calls"][0]["function"][
                    "arguments"
                ]:
                    tool_output = json.loads(
                        response["choices"][0]["message"]["tool_calls"][0]["function"][
                            "arguments"
                        ]
                    )
                else:
                    raise ValueError("No tool output")

                usage = response["usage"]

                # print(f"Response for {tool_name}: {response}")

                validate(instance=tool_output, schema=schema)
                print("✅ Validation passed.")
                return {
                    "content": [
                        {
                            "name": tool_name,
                            "input": tool_output,
                        }
                    ],
                    "usage": {
                        "input_tokens": usage["prompt_tokens"],
                        "cache_creation_input_tokens": 0,
                        "cache_read_input_tokens": 0,
                        "output_tokens": usage["completion_tokens"],
                        "service_tier": "standard",
                    },
                }
            except ValidationError as e:
                # Notifica error de validación
                print(f"❌ Validation error for '{tool_name}': {e.message}")
                if attempt < max_retries:
                    print("🔄 Retrying...")
                    continue
                else:
                    print("❌ Max retries exceeded.")
                    raise ValueError(
                        f"Max retries exceeded for '{tool_name}'. Last error: {e.message}"
                    )
            except Exception as e:
                # Notifica error general
                print(f"❌ Unexpected error: {e}")
                if attempt < max_retries:
                    print("🔄 Retrying...")
                    continue
                else:
                    print("❌ Max retries exceeded.")
                    raise ValueError(
                        f"Max retries exceeded for '{tool_name}'. Last error: {e.message}"
                    )

    async def run_pdf_toolchain(
        self,
        item: QueueItem,
    ):
        print("PDF TOOLCHAIN")
        doc = fitz.open(item["file_path"])
        base64_images = []
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pix = page.get_pixmap(dpi=150)  # convertimos a imagen

            # Convertimos el pixmap a imagen PIL
            img = Image.open(io.BytesIO(pix.tobytes("png")))

            # Guardamos en memoria y convertimos a base64
            buffer = io.BytesIO()
            img.save(buffer, format="PNG")
            img_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            base64_images.append(img_base64)

        image_messages = [
            {
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{img_b64}"},
            }
            for img_b64 in base64_images
        ]

        response = await self.tool_handler(
            tools=[tool["data"] for tool in tools_standard],
            messages=[
                {
                    "role": "user",
                    "content": [
                        *image_messages,
                        {"type": "text", "text": tools_standard[0]["prompt"]},
                    ],
                }
            ],
            tool_name=tools_standard[0]["data"]["function"]["name"],
        )

        # Procesa con resto de herramientas en paralelo
        tasks = []
        for tool in tools_standard[1:]:
            print(tool["data"]["function"]["name"])
            tool_res = self.tool_handler(
                tools=[tool["data"] for tool in tools_standard],
                messages=[
                    {
                        "role": "user",
                        "content": [
                            *image_messages,
                            {"type": "text", "text": tool["prompt"]},
                        ],
                    }
                ],
                tool_name=tool["data"]["function"]["name"],
            )
            tasks.append(tool_res)
        results = await asyncio.gather(*tasks)
        respuestas = [response] + results
        item["data"] = respuestas
        return item

    def pdf_to_base64(self, file_path: str) -> Union[str, None]:
        try:
            with open(file_path, "rb") as pdf_file:
                binary_data = pdf_file.read()
                base_64_encoded_data = base64.b64encode(binary_data)
                base64_string = base_64_encoded_data.decode("utf-8")
            return base64_string
        except Exception as e:
            print(f"An error occurred: {e}")
            return None

    def save_uploaded_file(self, uploaded_file, save_path):
        # Create the full file path
        file_path = os.path.join(save_path, uploaded_file.name)

        # Ensure the directory exists
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # Save the file
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return file_path


async def main():
    # Configuración de la página
    st.set_page_config(page_title="AseguraScan", page_icon="📄", layout="wide")

    # Título y descripción
    st.sidebar.title("AseguraScan 📄")
    st.sidebar.write(
        "AseguraScan es una aplicación de inteligencia artificial diseñada para comparar pólizas de seguros y ayudarte a entender sus diferencias clave."
    )

    st.sidebar.subheader("Cargar Archivos PDF")
    uploaded_files = st.sidebar.file_uploader(
        "Arrastra y suelta tus archivos PDF aquí o haz clic para seleccionarlos",
        type=["pdf"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        st.sidebar.write("Archivos cargados:")
        for file in uploaded_files:
            st.sidebar.write(f"- {file.name}")

        if len(uploaded_files) < 2:
            st.sidebar.error(
                "Por favor, selecciona dos o más archivos PDF para comparar."
            )
        else:
            if st.sidebar.button("Iniciar Proceso", use_container_width=True):
                try:
                    comparador = SeguroOrchestrator(
                        api_key=os.getenv("GOOGLE_API_KEY"),
                        model="gemini-2.5-flash",
                    )

                    with st.spinner("Procesando archivos..."):
                        tasks = []
                        file_paths = []

                        for file in uploaded_files:
                            saved_file_path = comparador.save_uploaded_file(
                                file, "./downloads"
                            )
                            file_paths.append(saved_file_path)
                            item = {
                                "file_name": file.name,
                                "file_extension": "pdf",
                                "file_path": saved_file_path,
                                "media_type": file.type,
                                "process_id": file.file_id,
                            }
                            task = comparador.run_pdf_toolchain(item)
                            tasks.append(task)

                        respuestas = await asyncio.gather(*tasks)
                        with st.expander("Respuestas Raw"):
                            st.write(respuestas)

                        docs = comparador.format_docs(respuestas)
                        with st.expander("Respuestas Formateadas"):
                            st.write(docs)
                        # title = "CLUB DE PATINAJE ORION"
                        # excel_id = "NIT 891900262-8"
                        # subtitle = "POLIZA COLECTIVA DE ACCIDENTES PERSONALES"
                        # show_comparador_table(docs)
                        comparador.show_excel_download(
                            "Arango Bueno",
                            "CAMARA DE COMERCIO DE PALMIRA",
                            "CUADRO COMPARATIVO DE SEGUROS",
                            docs,
                        )

                except Exception as e:
                    st.error(f"Ocurrió un error: {e}")
                finally:
                    # Clean up downloaded files
                    for file_path in file_paths:
                        try:
                            if os.path.exists(file_path):
                                os.remove(file_path)
                        except Exception as e:
                            print(f"Error deleting file {file_path}: {e}")


if __name__ == "__main__":
    asyncio.run(main())
