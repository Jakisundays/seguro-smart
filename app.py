# Standard library imports
import os
import ssl
import uuid
import json
import base64
import logging
from io import BytesIO
from pathlib import Path
from typing import List, Optional, TypedDict, Dict, Literal
from dataclasses import dataclass, field

# Third party imports
import aiohttp
import certifi
import asyncio
import streamlit as st
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import copy

# Local imports
from polizas_tools import tools as tools_standard


class PolizaDict(TypedDict):
    interes_asegurado: str
    valor_asegurado: int


class DocAdicionalDict(TypedDict):
    archivo: str
    prima_sin_iva: int
    iva: int
    prima_con_iva: int


class DetalleCobertura(TypedDict):
    interes_asegurado: str
    valor_asegurado: int
    tipo: List[str]


class RiesgoDict(TypedDict):
    ubicacion: str
    detalle_cobertura: List[DetalleCobertura]


class Amparo(TypedDict):
    amparo: str
    deducible: str
    tipo: List[str]


class AmparosDict(TypedDict):
    archivo: str
    amparos: List[Amparo]


@dataclass
class Prima:
    """
    Representa el valor de una prima individual con desglose de IVA.
    """

    prima_sin_iva: float
    iva: float
    prima_con_iva: float


@dataclass
class DetalleCoberturaItem:
    """
    Representa un interés asegurado con su valor asegurado.
    """

    interes_asegurado: str  # Ejemplo: "Edificio", "Maquinaria"
    valor_asegurado: float  # Valor monetario asegurado


@dataclass
class Cobertura:
    """
    Representa el detalle completo de la cobertura.
    """

    detalle_cobertura: List[DetalleCoberturaItem] = field(default_factory=list)
    total_valores_asegurados: float = 0.0

    def calcular_total(self) -> float:
        """
        Calcula y actualiza la suma total de los valores asegurados.
        """
        self.total_valores_asegurados = sum(
            item.valor_asegurado for item in self.detalle_cobertura
        )
        return self.total_valores_asegurados


@dataclass
class QueueItem(TypedDict):
    file_name: str
    file_extension: str
    file_path: str
    media_type: str
    process_id: str
    doc_type: Literal["actual", "renovacion", "adicional", "conjunto"]


load_dotenv()

# Configurar logger
app_logger = logging.getLogger(__name__)
app_logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)
app_logger.addHandler(handler)


def mostrar_riesgos(riesgos: list, titulo: str):
    """
    Muestra en Streamlit una tabla con los riesgos actuales.

    Args:
        riesgos (list): Lista de riesgos con ubicacion y detalle_cobertura.
        titulo (str): Título a mostrar arriba de la tabla.
    """
    # Normalizar datos en filas planas
    rows = []
    for riesgo in riesgos:
        for cobertura in riesgo["detalle_cobertura"]:
            rows.append(
                {
                    "Ubicación": riesgo["ubicacion"],
                    "Interés Asegurado": cobertura["interes_asegurado"],
                    "Valor Asegurado": cobertura["valor_asegurado"],
                    "Tipo": cobertura["tipo"],
                }
            )

    df = pd.DataFrame(rows)

    # Formatear valores asegurados con separador de miles
    df["Valor Asegurado"] = df["Valor Asegurado"].apply(lambda x: f"{x:,.0f}")

    # Mostrar tabla
    st.title(titulo)
    st.dataframe(df, use_container_width=True)


def mostrar_amparos(amparos_input, titulo: str = "📄 Amparos"):
    """
    Muestra en Streamlit una tabla de amparos sin modificar el input original.
    """
    # Hacemos una copia profunda para no tocar el input
    amparos = copy.deepcopy(amparos_input)

    rows = []

    # Si es un dict con un solo archivo
    if isinstance(amparos, dict):
        archivo = amparos.get("archivo", "")
        for a in amparos.get("amparos", []):
            rows.append(
                {
                    "Archivo": archivo,
                    "Amparo": a["amparo"],
                    "Deducible": a["deducible"],
                    "Tipo": a["tipo"],
                }
            )
    # Si es una lista de dicts (varios archivos)
    elif isinstance(amparos, list):
        for archivo_data in amparos:
            archivo = archivo_data.get("archivo", "")
            for a in archivo_data.get("amparos", []):
                rows.append(
                    {
                        "Archivo": archivo,
                        "Amparo": a["amparo"],
                        "Deducible": a["deducible"],
                        "Tipo": a["tipo"],
                    }
                )

    df = pd.DataFrame(rows)
    st.title(titulo)
    st.dataframe(df, use_container_width=True)


def mostrar_amparos_adicionales(
    amparos_adicionales: list, titulo: str = "📂 Amparos Adicionales"
):
    """
    Muestra en Streamlit una tabla con los amparos adicionales de varios archivos.

    Args:
        amparos_adicionales (list): Lista de diccionarios con clave 'archivo' y lista de 'amparos'.
        titulo (str): Título a mostrar arriba de la tabla.
    """
    rows = []
    for archivo_data in amparos_adicionales:
        archivo = archivo_data.get("archivo", "")
        for a in archivo_data.get("amparos", []):
            rows.append(
                {
                    "Archivo": archivo,
                    "Amparo": a["amparo"],
                    "Deducible": a["deducible"],
                    "Tipo": a["tipo"],
                }
            )

    df = pd.DataFrame(rows)

    st.title(titulo)
    st.dataframe(df, use_container_width=True)


def obtener_nombres_archivos(archivos):
    """
    Recibe una lista de UploadedFile (Streamlit) y devuelve
    un string con todos los nombres de archivo separados por comas.
    """
    if not archivos:
        return ""
    return ", ".join([archivo.name for archivo in archivos])


def generar_excel_analisis_polizas(
    poliza_actual: List[PolizaDict],
    poliza_renovacion: List[PolizaDict],
    riesgos_actuales: List[RiesgoDict],
    riesgos_renovacion: List[RiesgoDict],
    amparos_actuales: AmparosDict,
    amparos_renovacion: AmparosDict,
    amparos_adicionales: List[AmparosDict],
    output_path="reporte_polizas_riesgos.xlsx",
):
    """
    Genera un archivo Excel con análisis de pólizas replicando la funcionalidad
    del código original de app_structure.py usando los datos de x.py

    Args:
        output_path (str): Ruta donde se guardará el archivo Excel

    Returns:
        str: Ruta del archivo generado
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ===== estilos =====
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"
        )
        data_font = Font(name="Calibri", size=11)
        border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True
        )
        currency_alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True, shrink_to_fit=True
        )
        left_alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=True
        )

        def format_worksheet(ws, df, sheet_type):
            ws.insert_rows(1, 2)
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = "CENTRO DE DIAGNOSTICO AUTOMOTOR DE PALMIRA"
            title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(
                start_color="1F4E79", end_color="1F4E79", fill_type="solid"
            )
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=len(df.columns)
            )

            if sheet_type == "polizas":
                label_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
                label_fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                for row_num in range(3, ws.max_row + 1):
                    cell_value = ws.cell(row=row_num, column=1).value
                    if cell_value and (
                        "PÓLIZAS ACTUALES" in str(cell_value)
                        or "PÓLIZAS DE RENOVACIÓN" in str(cell_value)
                    ):
                        label_cell = ws.cell(row=row_num, column=1)
                        label_cell.font = label_font
                        label_cell.fill = label_fill
                        label_cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        ws.merge_cells(
                            start_row=row_num,
                            start_column=1,
                            end_row=row_num,
                            end_column=len(df.columns),
                        )

            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=3, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment

            for row_num in range(4, len(df) + 4):
                for col_num in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.border = border
                    col_name = df.columns[col_num - 1]
                    if any(x in col_name.lower() for x in ["valor", "prima", "iva"]):
                        cell.alignment = currency_alignment
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "$#,##0"
                    elif col_name.lower() == "coberturas":
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = center_alignment

            for col_num in range(1, len(df.columns) + 1):
                column_letter = get_column_letter(col_num)
                col_name = df.columns[col_num - 1]
                max_length = max(
                    len(str(col_name)),
                    len("CENTRO DE DIAGNOSTICO AUTOMOTOR DE PALMIRA")
                    // len(df.columns),
                )
                for row_num in range(4, len(df) + 4):
                    cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
                    max_length = max(max_length, len(cell_value))
                if col_name.lower() == "coberturas":
                    width = 45
                else:
                    width = min(max(max_length + 2, 15), 40)
                ws.column_dimensions[column_letter].width = width

            ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
            ws.freeze_panes = "A2"

        # ===== hoja Análisis_Estructurado =====
        intereses_unicos = list(
            set(
                [item["Interés Asegurado"] for item in poliza_actual]
                + [item["Interés Asegurado"] for item in poliza_renovacion]
            )
        )
        valores_actuales = {
            item["Interés Asegurado"]: item["Valor Asegurado"] for item in poliza_actual
        }
        valores_renovacion = {
            item["Interés Asegurado"]: item["Valor Asegurado"]
            for item in poliza_renovacion
        }
        total_actual = sum(valores_actuales.values())
        total_renovacion = sum(valores_renovacion.values())

        coberturas = [
            "Incendio y/o Rayo Edificios",
            "Explosión Mejoras Locativas",
            "Terremoto, temblor Muebles y Enseres",
            "Asonada, motin, conm. Civil/popular huelga Mercancias Fijas",
            "Extension de amparo",
            "Daños por agua, Anegación Dineros",
            "Incendio y/o Rayo en aparatos electricos Equipo Electronico",
        ]

        columnas_analisis = [
            "Coberturas",
            "Interés Asegurado",
            "Valor Asegurado Actual",
            "Valor Asegurado Renovado",
        ]
        intereses_ordenados = sorted(intereses_unicos)
        max_filas = max(len(intereses_ordenados), len(coberturas))

        datos_estructurados = []
        for i in range(max_filas):
            fila = {
                "Coberturas": coberturas[i] if i < len(coberturas) else "",
                "Interés Asegurado": (
                    intereses_ordenados[i] if i < len(intereses_ordenados) else ""
                ),
                "Valor Asegurado Actual": (
                    valores_actuales.get(intereses_ordenados[i], 0)
                    if i < len(intereses_ordenados)
                    else ""
                ),
                "Valor Asegurado Renovado": (
                    valores_renovacion.get(intereses_ordenados[i], 0)
                    if i < len(intereses_ordenados)
                    else ""
                ),
            }
            datos_estructurados.append(fila)

        datos_estructurados.append(
            {
                "Coberturas": "",
                "Interés Asegurado": "TOTAL",
                "Valor Asegurado Actual": total_actual,
                "Valor Asegurado Renovado": total_renovacion,
            }
        )

        # forzar solo columnas definidas
        df_estructurado = pd.DataFrame(datos_estructurados)[columnas_analisis]

        df_estructurado.to_excel(
            writer, sheet_name="Análisis_Estructurado", index=False
        )
        format_worksheet(
            writer.sheets["Análisis_Estructurado"], df_estructurado, "polizas"
        )

        # ===== hoja Polizas_Consolidadas =====
        def crear_hoja_consolidada():
            def procesar_poliza(poliza_data):
                if not poliza_data:
                    return None, None
                intereses_valores = {}
                total_poliza = sum(item["Valor Asegurado"] for item in poliza_data)
                for item in poliza_data:
                    intereses_valores[item["Interés Asegurado"]] = item[
                        "Valor Asegurado"
                    ]
                columnas = list(intereses_valores.keys()) + ["Total Póliza"]
                valores = [f"${valor:,.0f}" for valor in intereses_valores.values()] + [
                    f"${total_poliza:,.0f}"
                ]
                return columnas, valores

            columnas_actuales, valores_actuales = procesar_poliza(poliza_actual)
            columnas_renovacion, valores_renovacion = procesar_poliza(poliza_renovacion)

            datos_consolidados = []
            if columnas_actuales and valores_actuales:
                datos_consolidados.append(
                    ["PÓLIZAS ACTUALES"] + [""] * (len(columnas_actuales) - 1)
                )
                datos_consolidados.append(columnas_actuales)
                datos_consolidados.append(valores_actuales)
                datos_consolidados.append([""] * len(columnas_actuales))

            if columnas_renovacion and valores_renovacion:
                max_cols = max(
                    (len(columnas_actuales) if columnas_actuales else 0),
                    len(columnas_renovacion),
                )
                datos_consolidados.append(
                    ["PÓLIZAS DE RENOVACIÓN"] + [""] * (max_cols - 1)
                )
                columnas_renovacion_ajustadas = columnas_renovacion + [""] * (
                    max_cols - len(columnas_renovacion)
                )
                valores_renovacion_ajustados = valores_renovacion + [""] * (
                    max_cols - len(valores_renovacion)
                )
                datos_consolidados.append(columnas_renovacion_ajustadas)
                datos_consolidados.append(valores_renovacion_ajustados)

            if datos_consolidados:
                max_cols = max(len(fila) for fila in datos_consolidados)
                for i, fila in enumerate(datos_consolidados):
                    if len(fila) < max_cols:
                        datos_consolidados[i] = fila + [""] * (max_cols - len(fila))
                columnas_genericas = [f"Columna_{i+1}" for i in range(max_cols)]
                df_consolidado = pd.DataFrame(
                    datos_consolidados, columns=columnas_genericas
                )
                df_consolidado.to_excel(
                    writer, sheet_name="Polizas_Consolidadas", index=False, header=False
                )
                format_worksheet(
                    writer.sheets["Polizas_Consolidadas"], df_consolidado, "polizas"
                )

        crear_hoja_consolidada()

        # ===== hoja Amparos (tercera hoja) =====
        def crear_hoja_amparos():
            # Generar la tabla unificada de amparos basada en la imagen de referencia
            datos_amparos = []

            # Agregar secciones basadas en tipo de amparo
            tipos_encontrados = set()

            # Procesar amparos actuales
            for amp in amparos_actuales["amparos"]:
                tipos = amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                for t in tipos:
                    tipos_encontrados.add(t)

            # Procesar amparos renovación
            for amp in amparos_renovacion["amparos"]:
                tipos = amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                for t in tipos:
                    tipos_encontrados.add(t)

            # Procesar amparos adicionales
            for doc in amparos_adicionales:
                for amp in doc["amparos"]:
                    tipos = (
                        amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                    )
                    for t in tipos:
                        tipos_encontrados.add(t)

            # Ordenar tipos para crear secciones consistentes
            tipos_ordenados = sorted(tipos_encontrados)

            # Crear estructura basada en la imagen de referencia
            amparos_unicos = []
            amparos_vistos = set()

            # Recopilar todos los amparos únicos
            for amp in amparos_actuales["amparos"]:
                if amp["amparo"] not in amparos_vistos:
                    amparos_unicos.append(amp)
                    amparos_vistos.add(amp["amparo"])

            for amp in amparos_renovacion["amparos"]:
                if amp["amparo"] not in amparos_vistos:
                    amparos_unicos.append(amp)
                    amparos_vistos.add(amp["amparo"])

            for doc in amparos_adicionales:
                for amp in doc["amparos"]:
                    if amp["amparo"] not in amparos_vistos:
                        amparos_unicos.append(amp)
                        amparos_vistos.add(amp["amparo"])

            # Organizar por tipo (cada amparo puede pertenecer a múltiples tipos)
            amparos_por_tipo = {}
            for amp in amparos_unicos:
                tipos_amp = (
                    amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                )
                for t in tipos_amp:
                    amparos_por_tipo.setdefault(t, []).append(amp)

            # Crear estructura de datos para Excel
            for tipo in tipos_ordenados:
                # Añadir fila de sección (encabezado del tipo)
                datos_amparos.append(
                    {
                        "RAMO": tipo.upper(),
                        "CONDICIONES ACTUALES": "",
                        "ZURICH": "",
                        "AXA": "",
                        "BBVA": "",
                        "_es_seccion": True,
                    }
                )

                # Añadir amparos de este tipo
                if tipo in amparos_por_tipo:
                    for amp in amparos_por_tipo[tipo]:
                        # Buscar deducibles en actuales, renovación y adicionales
                        ded_actual = ""
                        ded_renovacion = ""
                        ded_adicional = ""

                        # Buscar en actuales
                        for a in amparos_actuales["amparos"]:
                            if a["amparo"] == amp["amparo"]:
                                ded_actual = a["deducible"]
                                break

                        # Buscar en renovación
                        for a in amparos_renovacion["amparos"]:
                            if a["amparo"] == amp["amparo"]:
                                ded_renovacion = a["deducible"]
                                break

                        # Buscar en adicionales
                        for doc in amparos_adicionales:
                            for a in doc["amparos"]:
                                if a["amparo"] == amp["amparo"]:
                                    ded_adicional = a["deducible"]
                                    break

                        datos_amparos.append(
                            {
                                "RAMO": amp["amparo"],
                                "CONDICIONES ACTUALES": ded_actual,
                                "CONDICIONES DE RENOVACIÓN": ded_renovacion,
                                **{
                                    doc["archivo"]: (
                                        next(
                                            (
                                                a["deducible"]
                                                for a in doc["amparos"]
                                                if a["amparo"] == amp["amparo"]
                                            ),
                                            "",
                                        )
                                    )
                                    for doc in amparos_adicionales
                                },
                                "_es_seccion": False,
                            }
                        )

            # Crear DataFrame
            # Columnas dinámicas: RAMO, actuales, renovación y un campo por cada archivo adicional
            columnas = [
                "RAMO",
                "CONDICIONES ACTUALES",
                "CONDICIONES DE RENOVACIÓN",
            ] + [doc["archivo"] for doc in amparos_adicionales]
            df_amparos = pd.DataFrame(datos_amparos)[columnas]

            # Escribir a Excel
            df_amparos.to_excel(writer, sheet_name="Amparos", index=False)
            ws = writer.sheets["Amparos"]

            # Formateo especial para la hoja de amparos

            # 1. Insertar filas de título y preparar cabeceras multinivel
            ws.insert_rows(1, 2)

            # 2. Título principal
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = "DEDUCIBLES"
            title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(
                start_color="1F4E79", end_color="1F4E79", fill_type="solid"
            )
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=len(columnas)
            )

            # 3. Encabezados de grupo (fila 2) y subencabezados (fila 3)
            # RAMO ocupa dos filas
            ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
            a2 = ws.cell(row=2, column=1)
            a2.value = "RAMO"
            a2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            a2.fill = PatternFill(
                start_color="305496", end_color="305496", fill_type="solid"
            )
            a2.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            a2.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            # CONDICIONES ACTUALES (columna B) con subencabezado dinámico del archivo de actuales
            b2 = ws.cell(row=2, column=2)
            b2.value = "CONDICIONES ACTUALES"
            b2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            b2.fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )
            b2.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            b2.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
            ws.cell(row=3, column=2).value = amparos_actuales.get(
                "archivo", "CONDICIONES ACTUALES"
            )

            # CONDICIONES DE RENOVACIÓN (columna C)
            c2 = ws.cell(row=2, column=3)
            c2.value = "CONDICIONES DE RENOVACIÓN"
            c2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c2.fill = PatternFill(
                start_color="2F75B5", end_color="2F75B5", fill_type="solid"
            )
            c2.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            c2.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
            ws.cell(row=3, column=3).value = amparos_renovacion.get(
                "archivo", "CONDICIONES DE RENOVACIÓN"
            )

            # COTIZACIONES adicionales (desde la columna 4 en adelante)
            if len(columnas) > 3:
                ws.merge_cells(
                    start_row=2, start_column=4, end_row=2, end_column=len(columnas)
                )
                d2 = ws.cell(row=2, column=4)
                d2.value = "ARCHIVOS ADICIONALES"
                d2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                d2.fill = PatternFill(
                    start_color="404040", end_color="404040", fill_type="solid"
                )
                d2.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                d2.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
                # Subencabezados con nombres reales de archivos
                for idx, nombre_archivo in enumerate(columnas[3:], start=4):
                    ws.cell(row=3, column=idx).value = nombre_archivo

            # 4. Formatear subencabezados de columna (fila 3)
            for col_num in range(1, len(columnas) + 1):
                cell = ws.cell(row=3, column=col_num)
                cell.font = Font(
                    name="Calibri",
                    size=11,
                    bold=True,
                    color="000000" if col_num >= 4 else "FFFFFF",
                )
                fill_map = {
                    1: "305496",  # RAMO
                    2: "A9D08E",  # Actual
                    3: "9DC3E6",  # Renovación
                }
                color = fill_map.get(col_num, "D9D9D9")
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # 4. Formatear datos
            for row_num in range(4, len(df_amparos) + 4):
                fila_datos = datos_amparos[row_num - 4]
                es_seccion = fila_datos.get("_es_seccion", False)

                for col_num in range(1, len(columnas) + 1):
                    cell = ws.cell(row=row_num, column=col_num)

                    if es_seccion:
                        # Formato para filas de sección (tipo de amparo)
                        cell.font = Font(
                            name="Calibri", size=11, bold=True, color="FFFFFF"
                        )
                        cell.fill = PatternFill(
                            start_color="4472C4", end_color="4472C4", fill_type="solid"
                        )
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

                        # Mergear toda la fila para la sección
                        if col_num == 1:
                            ws.merge_cells(
                                start_row=row_num,
                                start_column=1,
                                end_row=row_num,
                                end_column=len(columnas),
                            )
                    else:
                        # Formato para filas de datos normales
                        cell.font = Font(name="Calibri", size=10)
                        cell.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=True
                        )

                    # Bordes para todas las celdas
                    cell.border = Border(
                        left=Side(style="thin", color="000000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000"),
                    )

            # 5. Ajustar anchos de columna
            ws.column_dimensions["A"].width = 50  # RAMO - más ancho para nombres largos
            ws.column_dimensions["B"].width = 30  # CONDICIONES ACTUALES
            ws.column_dimensions["C"].width = 30  # ZURICH
            ws.column_dimensions["D"].width = 30  # AXA
            ws.column_dimensions["E"].width = 30  # BBVA

            # 6. Ajustar altura de filas para mejor legibilidad
            for row_num in range(1, len(df_amparos) + 4):
                ws.row_dimensions[row_num].height = 20

        crear_hoja_amparos()

        # ===== hoja Riesgos (cuarta hoja) =====
        def crear_hoja_riesgos():
            # 1) Normalización de intereses asegurados (para consistencia)
            def normalizar_interes(texto: str) -> str:
                t = (texto or "").strip().lower()
                reemplazos = {
                    "edificio": "Edificio",
                    "muebles y enseres": "Muebles y Enseres",
                    "maquinaria y equipo": "Maquinaria y Equipo",
                    "equipo electrico y electronico": "Equipo Eléctrico y Electrónico",
                    "equipo eléctrico y electrónico": "Equipo Eléctrico y Electrónico",
                    "mercancías": "Mercancías",
                    "mercancias": "Mercancías",
                    "dineros": "Dineros",
                    "equipo movil": "Equipo Móvil",
                    "equipo móvil": "Equipo Móvil",
                    "obras de arte": "Obras de Arte",
                }
                # caso combinado en renovación
                if "+" in t and "muebles" in t and "obras" in t:
                    return "Muebles y Enseres"  # asignamos al rubro principal
                return reemplazos.get(t, texto)

            # 2) Definir columnas a partir de riesgos_actuales (L36-375)
            intereses_cols = sorted(
                {
                    normalizar_interes(det["interes_asegurado"])  # noqa: E501
                    for r in riesgos_actuales
                    for det in r.get("detalle_cobertura", [])
                }
            )

            def pivot_riesgos(dataset):
                filas = []
                for r in dataset:
                    fila = {"Ubicación": r.get("ubicacion", "")}
                    for col in intereses_cols:
                        fila[col] = 0
                    for det in r.get("detalle_cobertura", []):
                        k = normalizar_interes(det.get("interes_asegurado", ""))
                        if k in fila:
                            fila[k] += det.get("valor_asegurado", 0)
                    filas.append(fila)
                # Totales
                tot = {"Ubicación": "TOTAL VALORES"}
                for col in intereses_cols:
                    tot[col] = sum(f[col] for f in filas)
                filas.append(tot)
                return pd.DataFrame(filas, columns=["Ubicación"] + intereses_cols)

            df_actual = pivot_riesgos(riesgos_actuales)
            df_renov = pivot_riesgos(riesgos_renovacion)

            # Escribir ambas tablas en la MISMA hoja, con los mismos encabezados
            start_row = 0
            sheet_name = "Riesgos"

            # Tabla 1: Póliza actual
            df_actual.to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=start_row + 1
            )
            ws = writer.sheets[sheet_name]
            ws.cell(row=start_row + 1, column=1, value="Póliza actual").font = Font(
                name="Arial", size=13, bold=True
            )
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=1,
                end_row=start_row + 1,
                end_column=len(df_actual.columns),
            )

            # Formato encabezados y celdas de la primera tabla
            header_row_1 = start_row + 2
            for col_idx in range(1, len(df_actual.columns) + 1):
                c = ws.cell(row=header_row_1, column=col_idx)
                c.font = header_font
                c.fill = header_fill
                c.border = border
                c.alignment = center_alignment
            for r in range(header_row_1 + 1, header_row_1 + 1 + len(df_actual)):
                for cidx in range(1, len(df_actual.columns) + 1):
                    c = ws.cell(row=r, column=cidx)
                    c.font = data_font
                    c.border = border
                    if cidx == 1:
                        c.alignment = left_alignment
                    else:
                        c.alignment = currency_alignment
                        if isinstance(c.value, (int, float)):
                            c.number_format = "$#,##0"

            # Tabla 2: Póliza de Renovación
            start_row_2 = header_row_1 + len(df_actual) + 2
            df_renov.to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=start_row_2 + 1
            )
            ws.cell(
                row=start_row_2 + 1, column=1, value="Póliza de Renovación"
            ).font = Font(name="Arial", size=13, bold=True)
            ws.merge_cells(
                start_row=start_row_2 + 1,
                start_column=1,
                end_row=start_row_2 + 1,
                end_column=len(df_renov.columns),
            )

            header_row_2 = start_row_2 + 2
            for col_idx in range(1, len(df_renov.columns) + 1):
                c = ws.cell(row=header_row_2, column=col_idx)
                c.font = header_font
                c.fill = header_fill
                c.border = border
                c.alignment = center_alignment
            for r in range(header_row_2 + 1, header_row_2 + 1 + len(df_renov)):
                for cidx in range(1, len(df_renov.columns) + 1):
                    c = ws.cell(row=r, column=cidx)
                    c.font = data_font
                    c.border = border
                    if cidx == 1:
                        c.alignment = left_alignment
                    else:
                        c.alignment = currency_alignment
                        if isinstance(c.value, (int, float)):
                            c.number_format = "$#,##0"

            # Ajuste de anchos para toda la hoja
            for cidx in range(1, len(df_actual.columns) + 1):
                letter = get_column_letter(cidx)
                if cidx == 1:
                    ws.column_dimensions[letter].width = 50
                else:
                    ws.column_dimensions[letter].width = 20

        crear_hoja_riesgos()

    with open(output_path, "wb") as f:
        f.write(output.getvalue())

    print(f"✅ Archivo Excel generado exitosamente: {output_path}")
    return output_path


def cleanup_processed_files(queue_items_list):
    """Elimina automáticamente los archivos después del procesamiento exitoso"""
    deleted_files = []
    failed_deletions = []

    for item in queue_items_list:
        try:
            if os.path.exists(item["file_path"]):
                os.remove(item["file_path"])
                deleted_files.append(item["file_name"])
                app_logger.info(
                    f"Archivo eliminado automáticamente: {item['file_name']}"
                )
        except Exception as e:
            failed_deletions.append((item["file_name"], str(e)))
            app_logger.error(f"Error al eliminar archivo {item['file_name']}: {str(e)}")

    return deleted_files, failed_deletions


# Funciones auxiliares para manejo de archivos
def save_uploaded_file(uploaded_file, doc_type: str) -> Optional[QueueItem]:
    """Guarda un archivo subido directamente en la carpeta downloads y crea un QueueItem"""
    if uploaded_file is None:
        return None

    try:
        # Crear directorio downloads si no existe
        downloads_dir = os.path.join(os.getcwd(), "downloads")
        os.makedirs(downloads_dir, exist_ok=True)

        # Generar nombre único para el archivo
        file_extension = uploaded_file.name.split(".")[-1].lower()
        timestamp = str(int(uuid.uuid4().int))[0:8]
        unique_filename = f"{timestamp}_{uploaded_file.name}"
        file_path = os.path.join(downloads_dir, unique_filename)

        # Guardar el archivo directamente en downloads
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        # Determinar media type
        media_type_map = {
            "pdf": "application/pdf",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "txt": "text/plain",
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
        }

        media_type = media_type_map.get(file_extension, "application/octet-stream")

        # Crear QueueItem
        queue_item: QueueItem = {
            "file_name": uploaded_file.name,
            "file_extension": file_extension,
            "file_path": file_path,
            "media_type": media_type,
            "process_id": str(uuid.uuid4()),
            "doc_type": doc_type,
        }

        return queue_item

    except Exception as e:
        st.error(f"Error al guardar el archivo {uploaded_file.name}: {str(e)}")
        return None


def get_file_info(file_path: str, file_name: str) -> str:
    """Obtiene información del archivo descargado"""
    try:
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            return f"✅ **{file_name}** descargado ({file_size_mb:.2f} MB)\n📁 Ubicación: `{file_path}`"
        else:
            return f"❌ Error: No se encontró el archivo {file_name}"
    except Exception as e:
        return f"❌ Error al obtener información de {file_name}: {str(e)}"


def process_files_and_create_queue(
    archivo_actual, archivo_renovacion, archivos_multiples, conjunto_documentos
) -> List[QueueItem]:
    """Procesa todos los archivos y crea la cola de procesamiento"""
    queue_items = []

    # Procesar póliza actual
    if archivo_actual:
        item = save_uploaded_file(archivo_actual, "actual")
        if item:
            queue_items.append(item)

    # Procesar póliza de renovación
    if archivo_renovacion:
        item = save_uploaded_file(archivo_renovacion, "renovacion")
        if item:
            queue_items.append(item)

    # Procesar documentos adicionales
    if archivos_multiples:
        for archivo in archivos_multiples:
            item = save_uploaded_file(archivo, "adicional")
            if item:
                queue_items.append(item)

    if conjunto_documentos:
        for doc in conjunto_documentos:
            pdf_bytes = doc.read()
            base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")
            document = {
                "file_name": doc.name,
                "file_extension": doc.name.split(".")[-1].lower(),
                "media_type": "application/pdf",
                "process_id": str(uuid.uuid4()),
                "doc_type": "conjunto",
                "base64": base64_pdf,
            }
            queue_items.append(document)

    return queue_items


def separar_archivos_por_tipo(archivos):
    conjuntos = []
    otros = []

    for archivo in archivos:
        if archivo.get("doc_type") == "conjunto":
            conjuntos.append(archivo)
        else:
            otros.append(archivo)

    return conjuntos, otros


def listar_valores_asegurados(poliza_json, tipo_documento):
    with st.expander("Detalles de la Póliza"):
        st.write(poliza_json)
    archivo = poliza_json.get("file_name", "Desconocido")
    total_poliza = poliza_json.get("total_valores_asegurados", 0)

    resultado = []
    for item in poliza_json.get("detalle_cobertura", []):
        resultado.append(
            {
                "Archivo": archivo,
                "Tipo de Documento": tipo_documento,
                "Interés Asegurado": item.get("interes_asegurado"),
                "Valor Asegurado": item.get("valor_asegurado"),
                "Total Póliza": total_poliza,
            }
        )

    return resultado


class InvoiceOrchestrator:
    def __init__(
        self,
        api_key: str,
        model: str,
    ):
        self.api_key = api_key
        self.model = model

    # Hace requests a la API con reintentos
    async def make_api_request(
        self, url: str, headers: Dict, data: Dict, process_id: str, retries: int = 5
    ) -> Optional[Dict]:
        ssl_context = ssl.create_default_context(cafile=certifi.where())
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        async with aiohttp.ClientSession(connector=connector) as session:
            for i in range(retries):
                try:
                    async with session.post(
                        url, headers=headers, json=data
                    ) as response:
                        if response.status == 200:
                            return await response.json()
                        elif response.status in [429, 529, 503]:
                            sleep_time = 15 * (i + 1)  # Espera incremental en segundos
                            app_logger.warning(
                                f"API request failed with status {response.status}. Retrying in {sleep_time} seconds..."
                            )
                            await asyncio.sleep(sleep_time)
                        else:
                            app_logger.error(
                                f"API request failed with status {response.status} - {await response.text()}"
                            )
                            raise ValueError(
                                f"Request failed with status {response.status}"
                            )
                except aiohttp.ClientError as e:
                    raise ValueError(f"Request error: {str(e)}")
        raise ValueError("Max retries exceeded.")

    async def handle_pdf(self, item: QueueItem):
        file_path = Path(item["file_path"])
        encoded_pdf = base64.b64encode(file_path.read_bytes()).decode()

        prompt = (
            tools_standard[1]["prompt"]
            if item["doc_type"] == "adicional"
            else tools_standard[0]["prompt"]
        )
        responseSchema = (
            tools_standard[1]["data"]
            if item["doc_type"] == "adicional"
            else tools_standard[0]["data"]
        )
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent"
        headers = {"x-goog-api-key": self.api_key, "Content-Type": "application/json"}
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "inline_data": {
                                "mime_type": "application/pdf",
                                "data": encoded_pdf,
                            }
                        },
                        {"text": prompt},
                    ]
                }
            ],
            "generationConfig": {
                "responseMimeType": "application/json",
                "responseSchema": responseSchema,
            },
        }
        response = await self.make_api_request(
            url, headers, payload, item["process_id"]
        )
        item["data"] = response
        return item

    async def batch_processor(self, input_files):
        if not input_files:
            return

        base64_docs = []

        for file in input_files:
            base64_docs.append(file["base64"])

        # # Construir payload
        files_metadata_b64 = [
            {"inline_data": {"mime_type": "application/pdf", "data": doc}}
            for doc in base64_docs
        ]

        prompt = tools_standard[1]["prompt"]
        responseSchema = tools_standard[1]["data"]

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={orchestrator.api_key}"

        headers = {"Content-Type": "application/json"}

        payload = {
            "contents": [{"parts": files_metadata_b64 + [{"text": prompt}]}],
            "generationConfig": {
                "responseMimeType": "application/json",
                "responseSchema": responseSchema,
            },
        }

        response = await self.make_api_request(
            url=url, headers=headers, data=payload, process_id="Dinardi"
        )

        response["doc_type"] = input_files[0]["doc_type"]

        return response


orchestrator = InvoiceOrchestrator(
    api_key=os.getenv("GEMINI_API_KEY"),
    model="gemini-2.5-flash",
)

# Configuración de la página
st.set_page_config(
    page_title="Análisis de pólizas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)


async def main():
    # Barra lateral
    with st.sidebar:
        # Campo 1: Póliza Actual
        st.subheader("1️⃣ Póliza Actual")
        archivo_poliza_actual = st.file_uploader(
            "Cargar archivo de póliza actual",
            type=["pdf", "docx", "txt", "jpg", "png"],
            key="poliza_actual",
            help="Sube el archivo de tu póliza actual para extraer sus datos",
        )

        st.markdown("---")

        # Campo 2: Póliza de Renovación
        st.subheader("2️⃣ Póliza de Renovación")
        archivo_poliza_renovacion = st.file_uploader(
            "Cargar archivo de póliza de renovación",
            type=["pdf", "docx", "txt", "jpg", "png"],
            key="poliza_renovacion",
            help="Sube el archivo de la póliza de renovación para comparar",
        )

        st.markdown("---")

        # Campo 3: Múltiples Documentos
        st.subheader("3️⃣ Documentos Adicionales")
        archivos_multiples = st.file_uploader(
            "Cargar múltiples documentos",
            type=["pdf", "docx", "txt", "jpg", "png"],
            accept_multiple_files=True,
            key="documentos_multiples",
            help="Sube múltiples documentos para extraer las primas de cada uno",
        )

        st.markdown("---")

        # campo 4: Conjunto de documentos
        st.subheader("4️⃣ Conjunto de Documentos adicionales")
        archivos_conjuntos = st.file_uploader(
            "Cargar conjunto de documentos",
            type=["pdf", "docx", "txt", "jpg", "png"],
            accept_multiple_files=True,
            key="archivos_conjuntos",
            help="Sube un conjunto de documentos para extraer información de todos ellos",
        )

        st.markdown("---")

    if st.sidebar.button(
        "🚀 Iniciar Proceso",
        type="primary",
        use_container_width=True,
        help="Haz clic para iniciar el procesamiento de todos los archivos cargados",
    ):
        if (
            archivo_poliza_actual
            or archivo_poliza_renovacion
            or archivos_multiples
            or archivos_conjuntos
        ):

            queue_items = process_files_and_create_queue(
                archivo_poliza_actual,
                archivo_poliza_renovacion,
                archivos_multiples,
                archivos_conjuntos,
            )

            conjuntos, otros = separar_archivos_por_tipo(queue_items)

            tasks = []

            if otros:
                for item in otros:
                    tasks.append(orchestrator.handle_pdf(item))

            if conjuntos:
                procesos_conjuntos = orchestrator.batch_processor(conjuntos)
                tasks.append(procesos_conjuntos)

            with st.spinner("Procesando..."):
                results = await asyncio.gather(*tasks)

            with st.expander("Resultados"):
                st.write(results)

            poliza_actual = None
            poliza_renovacion = None
            docs_adicionales = []
            doc_conjuntos = None

            amparos_adicionales = []

            with st.expander("Actual"):
                for item in results:
                    if item.get("doc_type") == "actual":
                        poliza_actual = json.loads(
                            item.get("data")
                            .get("candidates")[0]
                            .get("content", {})
                            .get("parts", [{}])[0]
                            .get("text", "")
                        )
                        poliza_actual["file_name"] = item.get("file_name")
                        st.write(poliza_actual)

            with st.expander("Renovación"):
                for item in results:
                    if item.get("doc_type") == "renovacion":
                        poliza_renovacion = json.loads(
                            item.get("data")
                            .get("candidates")[0]
                            .get("content", {})
                            .get("parts", [{}])[0]
                            .get("text", "")
                        )
                        poliza_renovacion["file_name"] = item.get("file_name")
                        st.write(poliza_renovacion)

            with st.expander("Adicional"):
                for item in results:
                    if item.get("doc_type") == "adicional":
                        doc_adicional = json.loads(
                            item.get("data")
                            .get("candidates")[0]
                            .get("content", {})
                            .get("parts", [{}])[0]
                            .get("text", "")
                        )
                        doc_adicional["file_name"] = item.get("file_name")
                        docs_adicionales.append(doc_adicional)
                        amparos_adicionales.append(
                            {
                                "archivo": doc_adicional.get("file_name"),
                                "amparos": doc_adicional.get("amparos"),
                            }
                        )
                        st.write(doc_adicional)

            with st.expander("Conjunto"):
                for item in results:
                    if item.get("doc_type") == "conjunto":
                        doc_conjuntos = json.loads(
                            item.get("candidates")[0]
                            .get("content", {})
                            .get("parts", [{}])[0]
                            .get("text", "")
                        )
                        doc_conjuntos["file_name"] = obtener_nombres_archivos(
                            archivos_conjuntos
                        )
                        amparos_adicionales.append(
                            {
                                "archivo": doc_conjuntos.get("file_name"),
                                "amparos": doc_conjuntos.get("amparos", ""),
                            }
                        )
                        st.write(doc_conjuntos)

            if poliza_actual:
                poliza_data_actual = listar_valores_asegurados(poliza_actual, "Actual")
            else:
                poliza_data_actual = []
            if poliza_renovacion:
                poliza_data_renovacion = listar_valores_asegurados(
                    poliza_renovacion, "Renovacion"
                )
            else:
                poliza_data_renovacion = []
            if docs_adicionales:
                docs_adicionales_data = [
                    {
                        "Archivo": doc.get("file_name"),
                        "Tipo de Documento": "Adicional",
                        "Prima Sin IVA": doc.get("prima_sin_iva", 0),
                        "IVA": doc.get("iva", 0),
                        "Prima Con IVA": doc.get("prima_con_iva", 0),
                    }
                    for doc in docs_adicionales
                ]
            else:
                docs_adicionales_data = []

            with st.expander("Poliza Data Actual"):
                st.write(poliza_data_actual)

            with st.expander("Poliza Data Renovacion"):
                st.write(poliza_data_renovacion)

            with st.expander("Docs Adicionales"):
                st.write(docs_adicionales_data)

            df_polizas = pd.DataFrame(poliza_data_actual + poliza_data_renovacion)

            polizas_actuales = df_polizas[df_polizas["Tipo de Documento"] == "Actual"]
            polizas_renovacion = df_polizas[
                df_polizas["Tipo de Documento"] == "Renovacion"
            ]

            if not polizas_actuales.empty:
                poliza_actual_cuadro = [
                    {
                        "Interés Asegurado": poliza["Interés Asegurado"],
                        "Valor Asegurado": poliza["Valor Asegurado"],
                    }
                    for poliza in poliza_data_actual
                ]
                with st.expander("poliza_actual_cuadro"):
                    st.write(poliza_actual_cuadro)
                st.subheader("📋 Pólizas Actuales")
                df_actuales_display = polizas_actuales.copy()
                df_actuales_display["Valor Asegurado"] = df_actuales_display[
                    "Valor Asegurado"
                ].apply(lambda x: f"${x:,.0f}")
                df_actuales_display["Total Póliza"] = df_actuales_display[
                    "Total Póliza"
                ].apply(lambda x: f"${x:,.0f}")
                st.dataframe(poliza_actual_cuadro, use_container_width=True)

                st.markdown("---")

            if not polizas_renovacion.empty:
                poliza_renovacion_cuadro = [
                    {
                        "Interés Asegurado": poliza["Interés Asegurado"],
                        "Valor Asegurado": poliza["Valor Asegurado"],
                    }
                    for poliza in poliza_data_renovacion
                ]
                with st.expander("poliza_renovacion_cuadro"):
                    st.write(poliza_renovacion_cuadro)
                st.subheader("🔄 Pólizas de Renovación")
                df_renovacion_display = polizas_renovacion.copy()
                df_renovacion_display["Valor Asegurado"] = df_renovacion_display[
                    "Valor Asegurado"
                ].apply(lambda x: f"${x:,.0f}")
                df_renovacion_display["Total Póliza"] = df_renovacion_display[
                    "Total Póliza"
                ].apply(lambda x: f"${x:,.0f}")
                st.dataframe(poliza_renovacion_cuadro, use_container_width=True)

                # Métrica del total
                total_renovacion = (
                    polizas_renovacion["Total Póliza"].iloc[0]
                    if len(polizas_renovacion) > 0
                    else 0
                )
                st.metric(
                    "💰 Total Póliza Renovación",
                    f"${total_renovacion:,.0f}",
                )

                st.markdown("---")

            if docs_adicionales:
                st.header("Primas")
                solo_primas = [
                    {
                        "Archivo": item["Archivo"],
                        "Prima Sin IVA": item["Prima Sin IVA"],
                        "IVA": item["IVA"],
                        "Prima Con IVA": item["Prima Con IVA"],
                    }
                    for item in docs_adicionales_data
                ]
                with st.expander("solo_primas"):
                    st.write(solo_primas)

                df_primas = pd.DataFrame(solo_primas)

                # Formatear valores monetarios para visualización
                df_primas_display = df_primas.copy()
                df_primas_display["Prima Sin IVA"] = df_primas_display[
                    "Prima Sin IVA"
                ].apply(lambda x: f"${x:,.0f}")
                df_primas_display["IVA"] = df_primas_display["IVA"].apply(
                    lambda x: f"${x:,.0f}"
                )
                df_primas_display["Prima Con IVA"] = df_primas_display[
                    "Prima Con IVA"
                ].apply(lambda x: f"${x:,.0f}")

                st.dataframe(solo_primas, use_container_width=True)

                st.markdown("---")

            riesgos_actuales = poliza_actual["riesgos"] if poliza_actual else []
            riesgos_renovacion = (
                poliza_renovacion["riesgos"] if poliza_renovacion else []
            )
            amparos_actuales = (
                {
                    "archivo": poliza_actual.get("file_name"),
                    "amparos": poliza_actual["amparos"],
                }
                if poliza_actual
                else {
                    "archivo": poliza_actual.get("file_name"),
                    "amparos": [],
                }
            )
            amparos_renovacion = (
                {
                    "archivo": poliza_renovacion.get("file_name"),
                    "amparos": poliza_renovacion["amparos"],
                }
                if poliza_renovacion
                else {
                    "archivo": poliza_renovacion.get("file_name"),
                    "amparos": [],
                }
            )

            mostrar_riesgos(riesgos_actuales, "📊 Tabla de Riesgos Actuales")
            mostrar_riesgos(riesgos_renovacion, "📊 Tabla de Riesgos de Renovación")

            mostrar_amparos(amparos_actuales, "📄 Amparos Actuales")
            mostrar_amparos(amparos_renovacion, "📄 Amparos de Renovación")

            mostrar_amparos_adicionales(amparos_adicionales, "📂 Amparos Adicionales")

            if poliza_data_actual or docs_adicionales_data:
                st.subheader("📥 Descargar Resultados")
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp_path = tmp.name
                tmp.close()

                try:
                    output_path = generar_excel_analisis_polizas(
                        poliza_actual=poliza_actual_cuadro,
                        poliza_renovacion=poliza_renovacion_cuadro,
                        riesgos_actuales=riesgos_actuales,
                        riesgos_renovacion=riesgos_renovacion,
                        amparos_actuales=amparos_actuales,
                        amparos_renovacion=amparos_renovacion,
                        amparos_adicionales=amparos_adicionales,
                        output_path=tmp_path,
                    )
                    with open(output_path, "rb") as f:
                        excel_bytes = f.read()
                        b64_excel = base64.b64encode(excel_bytes).decode()
                        href = (
                            "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                            + b64_excel
                        )
                        st.markdown(
                            f"""
                            <a class="download-btn" href="{href}" download="reporte_polizas_riesgos.xlsx" target="_blank" role="button" aria-label="Descargar Excel de análisis">
                            📥 <span>Descargar Excel</span>
                            </a>
                            """,
                            unsafe_allow_html=True,
                        )

                except Exception as e:
                    st.error(f"Error al generar el archivo Excel: {e}")
                finally:
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass

        else:
            st.warning("Por favor, sube al menos un archivo para procesar.")


if __name__ == "__main__":
    asyncio.run(main())
