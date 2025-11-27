# Standard library imports
import os
import ssl
from unittest import result
import uuid
import json
import base64
import logging
import shutil
from io import BytesIO
from pathlib import Path
from typing import List, Optional, TypedDict, Dict, Literal
from dataclasses import dataclass, field
from streamlit.runtime.uploaded_file_manager import UploadedFile  # si usas Streamlit
import unicodedata
import re
from copy import deepcopy
from itertools import cycle


# Third party imports
import aiohttp
import certifi
import asyncio
import streamlit as st
import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tempfile

load_dotenv()

# Local imports
from tools_actual import tools as tools_actuales
from tools_adicionales import tools as tools_adicionales
import json, pathlib
from rc import (
    clasificar_por_tipo,
    generar_tabla_excel_rc,
    integrar_hoja_en_libro,
)

# Configurar logger
app_logger = logging.getLogger(__name__)
app_logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
handler.setFormatter(formatter)
app_logger.addHandler(handler)


class QueueItem(TypedDict):
    file_name: str
    file_extension: str
    b64_str: str | list[str]
    media_type: str
    process_id: str
    doc_type: Literal["actual", "renovacion", "adicional", "conjunto"]


class PolizaDict(TypedDict):
    interes_asegurado: str
    valor_asegurado: int


class DocAdicionalDict(TypedDict):
    archivo: str
    prima_sin_iva: int
    # iva: int
    # prima_con_iva: int


class DetalleCobertura(TypedDict):
    interes_asegurado: str
    valor_asegurado: int
    tipo: List[str]


class RiesgoDict(TypedDict):
    ubicacion: str
    detalle_cobertura: List[DetalleCobertura]


class Amparo(TypedDict):
    amparo: str
    deducible: str
    tipo: List[str]


class AmparosDict(TypedDict):
    archivo: str
    amparos: List[Amparo]


def flatten_detalle_cobertura(data):
    """
    Aplana una lista de riesgos con 'ubicacion' y 'detalle_cobertura'
    en una sola lista de intereses asegurados.

    Args:
        data (list): Lista de objetos con 'ubicacion' y 'detalle_cobertura'.

    Returns:
        list: Lista plana de objetos con 'interes_asegurado', 'valor_asegurado',
              'tipo' y 'ubicacion'.
    """
    flat_list = []
    for item in data:
        # ubicacion = item.get("ubicacion")
        for detalle in item.get("detalle_cobertura", []):
            new_item = {
                "interes_asegurado": detalle.get("interes_asegurado"),
                "valor_asegurado": detalle.get("valor_asegurado"),
                "tipo": detalle.get("tipo", []),
                # "ubicacion": ubicacion,
            }
            flat_list.append(new_item)
    return flat_list


def transformar_amparos(data, tipo_filtro=None):
    """
    Convierte un JSON de archivos con amparos en un array por archivo,
    con formato normalizado para cada amparo, permitiendo filtrar por tipo.

    Args:
        data (list): Lista de archivos con sus amparos, formato original.
        tipo_filtro (str, optional): Tipo de amparo a filtrar (ej: "Incendio" o "Responsabilidad Civil").

    Returns:
        list: Lista de listas de amparos normalizados por archivo.
    """
    resultado = []

    for idx, file in enumerate(data, start=1):
        file_array = []
        for amparo in file.get("amparos", []):
            tipos = amparo.get("tipo", [])

            # Si hay filtro, solo incluir los que contengan ese tipo
            if tipo_filtro and tipo_filtro not in tipos:
                continue

            file_array.append(
                {
                    "Archivo": "adicional",
                    "Amparo": amparo.get("amparo", ""),
                    "Deducible": amparo.get("deducible", ""),
                    "Tipo": ", ".join(tipos),
                    "file_name": f"amparos_adicional_{idx}.pdf",
                }
            )

        # Solo agregar si hay resultados para este archivo
        if file_array:
            resultado.append(file_array)

    return resultado


def mostrar_poliza(res):
    st.subheader(f"Archivo: {res['file_name']}")
    st.write(f"Asegurado: {res.get('data', {}).get('asegurado', '')}")
    st.text(f"Tipo de documento: {res['doc_type']}")

    data = res.get("data", {})

    # Detalle de Cobertura
    st.header("Detalle de Cobertura")
    if data.get("detalle_cobertura"):
        df_cobertura = pd.DataFrame(
            [
                {
                    "Interés Asegurado": item["interes_asegurado"],
                    "Valor Asegurado": item["valor_asegurado"],
                    "Tipos": ", ".join(item["tipo"]),
                }
                for item in data["detalle_cobertura"]
            ]
        )
        st.dataframe(df_cobertura.style.format({"Valor Asegurado": "{:,.0f}"}))

    # Riesgos por ubicación
    st.header("Riesgos por Ubicación")
    for riesgo in data.get("riesgos", []):
        with st.expander(riesgo["ubicacion"]):
            df_riesgo = pd.DataFrame(
                [
                    {
                        "Interés Asegurado": d["interes_asegurado"],
                        "Valor Asegurado": d["valor_asegurado"],
                        "Tipos": ", ".join(d["tipo"]),
                    }
                    for d in riesgo.get("detalle_cobertura", [])
                ]
            )
            st.dataframe(df_riesgo.style.format({"Valor Asegurado": "{:,.0f}"}))

    # Prima y tasas
    st.header("Prima y Tasas")
    st.write(f"Prima sin IVA: {data.get('prima_sin_iva')}")
    # st.write(f"IVA: {data.get('iva')}")
    # st.write(f"Prima con IVA: {data.get('prima_con_iva')}")
    st.write(f"Tasa: {data.get('tasa')}")

    # Daños Materiales
    st.header("Daños Materiales")
    if data.get("danos_materiales"):
        df_danos = pd.DataFrame(
            [
                {"Tipo de Daño": k.replace("_", " ").capitalize(), "Límite": v}
                for k, v in data["danos_materiales"].items()
            ]
        )
        st.dataframe(df_danos)

    # Manejo Global Comercial
    st.header("Manejo Global Comercial")
    if data.get("manejo_global_comercial"):
        df_manejo = pd.DataFrame(
            [
                {"Concepto": k.replace("_", " ").capitalize(), "Límite": v}
                for k, v in data["manejo_global_comercial"].items()
            ]
        )
        st.dataframe(df_manejo)

    # Transporte de Valores
    st.header("Transporte de Valores")
    if data.get("transporte_valores"):
        df_transporte = pd.DataFrame(
            [
                {"Concepto": k.replace("_", " ").capitalize(), "Límite": v}
                for k, v in data["transporte_valores"].items()
            ]
        )
        st.dataframe(df_transporte)

    # Responsabilidad Civil
    st.header("Responsabilidad Civil")
    if data.get("responsabilidad_civil"):
        df_rc = pd.DataFrame(
            [
                {"Concepto": k.replace("_", " ").capitalize(), "Detalle": v}
                for k, v in data["responsabilidad_civil"].items()
            ]
        )
        st.dataframe(df_rc)

    # Amparos
    st.header("Amparos")
    if data.get("amparos"):
        df_amparos = pd.DataFrame(
            [
                {
                    "Amparo": a["amparo"],
                    "Deducible": a["deducible"],
                    "Tipo": ", ".join(a["tipo"]),
                }
                for a in data["amparos"]
            ]
        )
        st.dataframe(df_amparos)
        st.write("---")


def mostrar_poliza_adicional(res):
    st.subheader(f"Archivo: {res['file_name']}")
    st.text(f"Tipo de documento: {res['doc_type']}")

    data = res.get("data", {})

    # Prima y tasas
    st.header("Prima y Tasas")
    st.write(f"Prima sin IVA: {data.get('prima_sin_iva'):,}")
    # st.write(f"IVA: {data.get('iva'):,}")
    # st.write(f"Prima con IVA: {data.get('prima_con_iva'):,}")
    st.write(f"Tasa: {data.get('tasa')}")

    # Amparos
    st.header("Amparos")
    if data.get("amparos"):
        df_amparos = pd.DataFrame(
            [
                {
                    "Amparo": a["amparo"],
                    "Deducible": a["deducible"],
                    "Tipo": ", ".join(a["tipo"]),
                }
                for a in data["amparos"]
            ]
        )
        st.dataframe(df_amparos)

    # Daños Materiales
    st.header("Daños Materiales")
    if data.get("danos_materiales"):
        df_danos = pd.DataFrame(
            [
                {"Tipo de Daño": k.replace("_", " ").capitalize(), "Límite": v}
                for k, v in data["danos_materiales"].items()
            ]
        )
        st.dataframe(df_danos)

    # Manejo Global Comercial
    st.header("Manejo Global Comercial")
    if data.get("manejo_global_comercial"):
        df_manejo = pd.DataFrame(
            [
                {"Concepto": k.replace("_", " ").capitalize(), "Límite": v}
                for k, v in data["manejo_global_comercial"].items()
            ]
        )
        st.dataframe(df_manejo)

    # Transporte de Valores
    st.header("Transporte de Valores")
    if data.get("transporte_valores"):
        df_transporte = pd.DataFrame(
            [
                {"Concepto": k.replace("_", " ").capitalize(), "Límite": v}
                for k, v in data["transporte_valores"].items()
            ]
        )
        st.dataframe(df_transporte)

    # Responsabilidad Civil
    st.header("Responsabilidad Civil")
    if data.get("responsabilidad_civil"):
        df_rc = pd.DataFrame(
            [
                {"Concepto": k.replace("_", " ").capitalize(), "Detalle": v}
                for k, v in data["responsabilidad_civil"].items()
            ]
        )
        st.dataframe(df_rc)
    st.write("---")


def generar_prompt_nombres(nombres: list[str]) -> str:
    """
    Genera un string que pide crear un nombre para un documento
    basado en una lista de nombres proporcionados.mbn
    """
    return (
        f"Genera un nombre claro, completo y profesional para un documento que represente: "
        f"{' | '.join(nombres)}. No utilices abreviaciones ni siglas."
    )


def generar_excel_analisis_polizas(
    riesgos_actuales: List[RiesgoDict],
    riesgos_renovacion: List[RiesgoDict],
    amparos_actuales: AmparosDict,
    amparos_renovacion: AmparosDict,
    amparos_adicionales: List[AmparosDict],
    output_path="reporte_polizas_riesgos.xlsx",
    titulo_excel: Optional[str] = None,
):
    """
    Genera un archivo Excel con análisis de pólizas replicando la funcionalidad
    del código original de app_structure.py usando los datos de x.py

    Args:
        output_path (str): Ruta donde se guardará el archivo Excel

    Returns:
        str: Ruta del archivo generado

    Parámetros adicionales:
        titulo_excel (Optional[str]): Título personalizado que se mostrará en la parte
        superior de cada hoja para diferenciar visualmente el contenido. Si no se
        especifica, se usará un título por defecto acorde a la hoja.
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ===== estilos =====
        header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        data_font = Font(name="Calibri", size=11)
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        center_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True
        )
        currency_alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True, shrink_to_fit=True
        )
        left_alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=True
        )

        # Se omite la creación de las hojas 'Análisis_Estructurado' y 'Polizas_Consolidadas'
        # según requerimiento. El resto del documento permanece intacto.

        # ===== hoja Amparos (tercera hoja) =====
        def crear_hoja_amparos():
            # Generar la tabla unificada de amparos basada en la imagen de referencia
            datos_amparos = []

            # Agregar secciones basadas en tipo de amparo
            tipos_encontrados = set()

            # Procesar amparos actuales
            for amp in amparos_actuales["amparos"]:
                tipos = amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                for t in tipos:
                    tipos_encontrados.add(t)

            # Procesar amparos renovación
            for amp in amparos_renovacion["amparos"]:
                tipos = amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                for t in tipos:
                    tipos_encontrados.add(t)

            # Procesar amparos adicionales
            for doc in amparos_adicionales:
                for amp in doc["amparos"]:
                    tipos = (
                        amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                    )
                    for t in tipos:
                        tipos_encontrados.add(t)

            # Ordenar tipos para crear secciones consistentes
            # Normalización canónica y orden preferente
            def _norm_basic(s: str) -> str:
                import unicodedata

                s2 = (
                    "".join(
                        c
                        for c in unicodedata.normalize("NFD", s or "")
                        if unicodedata.category(c) != "Mn"
                    )
                    .lower()
                    .strip()
                )
                return s2

            def _tipo_canon(s: str) -> str:
                t = _norm_basic(s)
                if t.startswith("responsabilidad civil"):
                    return "responsabilidad_civil_amparo"
                mapping = {
                    "incendio": "incendio",
                    "sustraccion": "sustraccion",
                    "sustracción": "sustraccion",
                    "equipo electronico": "equipo_electronico",
                    "equipo electrónico": "equipo_electronico",
                    "rotura de maquinaria": "rotura_de_maquinaria",
                    "manejo de dinero": "manejo",
                    "manejo": "manejo",
                    "transporte de valores": "transporte_de_valores",
                }
                return mapping.get(t, t.replace(" ", "_"))

            orden_preferente = [
                "incendio",
                "sustraccion",
                "equipo_electronico",
                "rotura_de_maquinaria",
                "manejo",
                "responsabilidad_civil_amparo",
                "transporte_de_valores",
            ]

            # Mantener orden de detección original de tipos
            tipos_detectados = []
            vistos_tipos = set()
            tipo_canon_map = {}

            # Recorrer fuentes en el mismo orden de recopilación previa
            for amp in amparos_actuales["amparos"]:
                tipos = amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                for t in tipos:
                    if t not in vistos_tipos:
                        vistos_tipos.add(t)
                        tipos_detectados.append(t)
                    tipo_canon_map[t] = _tipo_canon(t)

            for amp in amparos_renovacion["amparos"]:
                tipos = amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                for t in tipos:
                    if t not in vistos_tipos:
                        vistos_tipos.add(t)
                        tipos_detectados.append(t)
                    tipo_canon_map[t] = _tipo_canon(t)

            for doc in amparos_adicionales:
                for amp in doc["amparos"]:
                    tipos = (
                        amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                    )
                    for t in tipos:
                        if t not in vistos_tipos:
                            vistos_tipos.add(t)
                            tipos_detectados.append(t)
                        tipo_canon_map[t] = _tipo_canon(t)

            # Construir lista ordenada por la preferencia indicada
            tipos_ordenados = []
            for canon in orden_preferente:
                for t in tipos_detectados:
                    if tipo_canon_map.get(t) == canon and t not in tipos_ordenados:
                        tipos_ordenados.append(t)
            # Agregar cualquier tipo restante no contemplado en la preferencia
            for t in tipos_detectados:
                if t not in tipos_ordenados:
                    tipos_ordenados.append(t)

            # Crear estructura basada en la imagen de referencia
            amparos_unicos = []
            amparos_vistos = set()

            # Recopilar todos los amparos únicos
            for amp in amparos_actuales["amparos"]:
                if amp["amparo"] not in amparos_vistos:
                    amparos_unicos.append(amp)
                    amparos_vistos.add(amp["amparo"])

            for amp in amparos_renovacion["amparos"]:
                if amp["amparo"] not in amparos_vistos:
                    amparos_unicos.append(amp)
                    amparos_vistos.add(amp["amparo"])

            for doc in amparos_adicionales:
                for amp in doc["amparos"]:
                    if amp["amparo"] not in amparos_vistos:
                        amparos_unicos.append(amp)
                        amparos_vistos.add(amp["amparo"])

            # Organizar por tipo (cada amparo puede pertenecer a múltiples tipos)
            amparos_por_tipo = {}
            for amp in amparos_unicos:
                tipos_amp = (
                    amp["tipo"] if isinstance(amp["tipo"], list) else [amp["tipo"]]
                )
                for t in tipos_amp:
                    amparos_por_tipo.setdefault(t, []).append(amp)

            def _normalize(text: str) -> str:
                """Convierte a minúsculas y elimina acentos/espacios extra para comparar strings."""
                if not text:
                    return ""
                return (
                    "".join(
                        c
                        for c in unicodedata.normalize("NFD", text)
                        if unicodedata.category(c) != "Mn"
                    )
                    .lower()
                    .strip()
                )

            def _consolidar_deducibles(lista_amparos, nombre_amparo: str) -> str:
                """
                Une deducibles de amparos duplicados del mismo nombre (ignorando tildes y mayúsculas),
                sin repetir valores, conservando el orden original.
                """
                vals = [
                    (a.get("deducible") or "").strip()
                    for a in lista_amparos
                    if isinstance(a, dict)
                    and (
                        a.get("amparo") == nombre_amparo
                        or _normalize(a.get("amparo")) == _normalize(nombre_amparo)
                    )
                ]

                # Eliminar deducibles vacíos
                vals = [v for v in vals if v]

                # Quitar duplicados manteniendo orden
                seen, uniq = set(), []
                for v in vals:
                    if v not in seen:
                        seen.add(v)
                        uniq.append(v)

                return "\n".join(uniq)

            def _consolidar_deducibles_doc(doc: dict, nombre_amparo: str) -> str:
                """Une deducibles de un documento adicional para un amparo, sin repetir."""
                vals = [
                    (a.get("deducible") or "").strip()
                    for a in doc.get("amparos", [])
                    if a.get("amparo") == nombre_amparo
                ]
                vals = [v for v in vals if v]
                seen = set()
                uniq = []
                for v in vals:
                    if v not in seen:
                        seen.add(v)
                        uniq.append(v)
                return "\n".join(uniq)

            # Crear estructura de datos para Excel
            for tipo in tipos_ordenados:
                # Añadir fila de sección (encabezado del tipo)
                datos_amparos.append(
                    {
                        "RAMO": tipo.upper(),
                        "CONDICIONES ACTUALES": "",
                        "ZURICH": "",
                        "AXA": "",
                        "BBVA": "",
                        "_es_seccion": True,
                    }
                )

                # Añadir amparos de este tipo
                if tipo in amparos_por_tipo:
                    for amp in amparos_por_tipo[tipo]:
                        # Consolidar deducibles en actuales, renovación y por cada adicional
                        ded_actual = _consolidar_deducibles(
                            amparos_actuales.get("amparos", []), amp["amparo"]
                        )
                        ded_renovacion = _consolidar_deducibles(
                            amparos_renovacion.get("amparos", []), amp["amparo"]
                        )
                        datos_amparos.append(
                            {
                                "RAMO": amp["amparo"],
                                "CONDICIONES ACTUALES": ded_actual,
                                "CONDICIONES DE RENOVACIÓN": ded_renovacion,
                                **{
                                    doc.get("archivo", ""): _consolidar_deducibles_doc(
                                        doc, amp["amparo"]
                                    )
                                    for doc in amparos_adicionales
                                },
                                "_es_seccion": False,
                            }
                        )

            print(f"amparos_por_tipo: {amparos_por_tipo}")

            # Crear DataFrame
            # Columnas dinámicas: RAMO, actuales, renovación y un campo por cada archivo adicional
            columnas = [
                "RAMO",
                "CONDICIONES ACTUALES",
                "CONDICIONES DE RENOVACIÓN",
            ] + [doc["archivo"] for doc in amparos_adicionales]
            df_amparos = pd.DataFrame(datos_amparos)[columnas]

            # Escribir a Excel
            df_amparos.to_excel(writer, sheet_name="Amparos", index=False)
            ws = writer.sheets["Amparos"]

            # Formateo especial para la hoja de amparos

            # 1. Insertar filas de título y preparar cabeceras multinivel
            ws.insert_rows(1, 2)

            # 2. Título principal
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = titulo_excel if titulo_excel else "DEDUCIBLES"
            title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(
                start_row=1, start_column=1, end_row=1, end_column=len(columnas)
            )

            # 3. Encabezados de grupo (fila 2) y subencabezados (fila 3)
            # RAMO ocupa dos filas
            ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
            a2 = ws.cell(row=2, column=1)
            a2.value = "RAMO"
            a2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            a2.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            a2.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            a2.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

            # CONDICIONES ACTUALES (columna B) con subencabezado dinámico del archivo de actuales
            b2 = ws.cell(row=2, column=2)
            b2.value = "CONDICIONES ACTUALES"
            b2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            b2.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            b2.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            b2.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
            ws.cell(row=3, column=2).value = amparos_actuales.get(
                "archivo", "CONDICIONES ACTUALES"
            )

            # CONDICIONES DE RENOVACIÓN (columna C)
            c2 = ws.cell(row=2, column=3)
            c2.value = "CONDICIONES DE RENOVACIÓN"
            c2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c2.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            c2.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            c2.border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
            ws.cell(row=3, column=3).value = amparos_renovacion.get(
                "archivo", "CONDICIONES DE RENOVACIÓN"
            )

            # COTIZACIONES adicionales (desde la columna 4 en adelante)
            if len(columnas) > 3:
                ws.merge_cells(
                    start_row=2, start_column=4, end_row=2, end_column=len(columnas)
                )
                d2 = ws.cell(row=2, column=4)
                d2.value = "ARCHIVOS ADICIONALES"
                d2.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                d2.fill = PatternFill(
                    start_color="1F4E78", end_color="1F4E78", fill_type="solid"
                )
                d2.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                d2.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
                # Subencabezados con nombres reales de archivos
                for idx, nombre_archivo in enumerate(columnas[3:], start=4):
                    ws.cell(row=3, column=idx).value = nombre_archivo

            # 4. Formatear subencabezados de columna (fila 3)
            for col_num in range(1, len(columnas) + 1):
                cell = ws.cell(row=3, column=col_num)
                cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                fill_map = {
                    1: "1F4E78",  # RAMO
                    2: "1F4E78",  # Actual
                    3: "1F4E78",  # Renovación
                }
                color = fill_map.get(col_num, "1F4E78")
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
                cell.border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # 4. Formatear datos
            for row_num in range(4, len(df_amparos) + 4):
                fila_datos = datos_amparos[row_num - 4]
                es_seccion = fila_datos.get("_es_seccion", False)

                for col_num in range(1, len(columnas) + 1):
                    cell = ws.cell(row=row_num, column=col_num)

                    if es_seccion:
                        # Formato para filas de sección (tipo de amparo)
                        cell.font = Font(
                            name="Calibri", size=11, bold=True, color="FFFFFF"
                        )
                        cell.fill = PatternFill(
                            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
                        )
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

                        # Mergear toda la fila para la sección
                        if col_num == 1:
                            ws.merge_cells(
                                start_row=row_num,
                                start_column=1,
                                end_row=row_num,
                                end_column=len(columnas),
                            )
                    else:
                        # Formato para filas de datos normales
                        cell.font = Font(name="Calibri", size=10)
                        cell.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=True
                        )

                    # Bordes para todas las celdas
                    cell.border = Border(
                        left=Side(style="thin", color="000000"),
                        right=Side(style="thin", color="000000"),
                        top=Side(style="thin", color="000000"),
                        bottom=Side(style="thin", color="000000"),
                    )

            # 5. Ajustar anchos de columna
            ws.column_dimensions["A"].width = 50  # RAMO - más ancho para nombres largos
            ws.column_dimensions["B"].width = 30  # CONDICIONES ACTUALES
            ws.column_dimensions["C"].width = 30  # ZURICH
            ws.column_dimensions["D"].width = 30  # AXA
            ws.column_dimensions["E"].width = 30  # BBVA

            # 6. Ajustar altura de filas para mejor legibilidad
            for row_num in range(1, len(df_amparos) + 4):
                ws.row_dimensions[row_num].height = 20

        # Nota: la hoja de Amparos se generará después de Riesgos para priorizar su orden

        # ===== hoja Riesgos (cuarta hoja) =====
        def crear_hoja_riesgos():
            # 1) Normalización de intereses asegurados (para consistencia)
            def normalizar_interes(texto: str) -> str:
                t = (texto or "").strip().lower()
                reemplazos = {
                    "edificio": "Edificio",
                    "muebles y enseres": "Muebles y Enseres",
                    "maquinaria y equipo": "Maquinaria y Equipo",
                    "equipo electrico y electronico": "Equipo Eléctrico y Electrónico",
                    "equipo eléctrico y electrónico": "Equipo Eléctrico y Electrónico",
                    "mercancías": "Mercancías",
                    "mercancias": "Mercancías",
                    "dineros": "Dineros",
                    "equipo movil": "Equipo Móvil",
                    "equipo móvil": "Equipo Móvil",
                    "obras de arte": "Obras de Arte",
                    "responsabilidad civil extracontractual": "Responsabilidad Civil",  # caso fijo
                }
                # caso combinado en renovación
                if "+" in t and "muebles" in t and "obras" in t:
                    return "Muebles y Enseres"  # asignamos al rubro principal

                # regla general: todo lo que empiece con "responsabilidad civil"
                if t.startswith("responsabilidad civil"):
                    return "Responsabilidad Civil"

                return reemplazos.get(t, texto)

            # 2) Definir columnas a partir de riesgos_actuales (L36-375)
            intereses_cols = sorted(
                {
                    normalizar_interes(det["interes_asegurado"])  # noqa: E501
                    for r in riesgos_actuales
                    for det in r.get("detalle_cobertura", [])
                }
            )

            def pivot_riesgos(dataset):
                # Agregar por ubicación para evitar filas duplicadas
                agregados = {}
                for r in dataset:
                    ubic = r.get("ubicacion", "")
                    if ubic not in agregados:
                        agregados[ubic] = {col: 0 for col in intereses_cols}
                    for det in r.get("detalle_cobertura", []):
                        k = normalizar_interes(det.get("interes_asegurado", ""))
                        if k in agregados[ubic]:
                            valor_asegurado = det.get("valor_asegurado", 0)
                            agregados[ubic][k] += valor_asegurado

                filas = []
                for ubic, valores in agregados.items():
                    fila = {"Ubicación": ubic}
                    fila.update(valores)
                    filas.append(fila)

                # Totales
                tot = {"Ubicación": "TOTAL VALORES"}
                for col in intereses_cols:
                    tot[col] = sum(f[col] for f in filas)
                filas.append(tot)
                return pd.DataFrame(filas, columns=["Ubicación"] + intereses_cols)

            out_dir = pathlib.Path(__file__).parent / "riesgos_json"
            out_dir.mkdir(exist_ok=True)

            df_actual = pivot_riesgos(riesgos_actuales)
            df_renov = pivot_riesgos(riesgos_renovacion)

            # Escribir ambas tablas en la MISMA hoja, con los mismos encabezados
            # Reservamos la fila 1 para un título general de la hoja
            start_row = 1
            sheet_name = "Riesgos"

            # Tabla 1: Póliza actual
            df_actual.to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=start_row + 1
            )
            ws = writer.sheets[sheet_name]

            # Título principal de la hoja Riesgos (fila 1)
            riesgos_title_cell = ws.cell(row=1, column=1)
            riesgos_title_cell.value = titulo_excel if titulo_excel else "RIESGOS"
            riesgos_title_cell.font = Font(
                name="Calibri", size=16, bold=True, color="FFFFFF"
            )
            riesgos_title_cell.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            riesgos_title_cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.merge_cells(
                start_row=1,
                start_column=1,
                end_row=1,
                end_column=len(df_actual.columns),
            )
            ws.cell(row=start_row + 1, column=1, value="Póliza actual").font = Font(
                name="Arial", size=13, bold=True
            )
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=1,
                end_row=start_row + 1,
                end_column=len(df_actual.columns),
            )

            # Formato encabezados y celdas de la primera tabla
            header_row_1 = start_row + 2
            for col_idx in range(1, len(df_actual.columns) + 1):
                c = ws.cell(row=header_row_1, column=col_idx)
                c.font = header_font
                c.fill = header_fill
                c.border = border
                c.alignment = center_alignment
            for r in range(header_row_1 + 1, header_row_1 + 1 + len(df_actual)):
                for cidx in range(1, len(df_actual.columns) + 1):
                    c = ws.cell(row=r, column=cidx)
                    c.font = data_font
                    c.border = border
                    if cidx == 1:
                        c.alignment = left_alignment
                    else:
                        c.alignment = currency_alignment
                        if isinstance(c.value, (int, float)):
                            c.number_format = "$#,##0"

            # Tabla 2: Póliza de Renovación
            start_row_2 = header_row_1 + len(df_actual) + 2
            df_renov.to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=start_row_2 + 1
            )
            ws.cell(
                row=start_row_2 + 1, column=1, value="Póliza de Renovación"
            ).font = Font(name="Arial", size=13, bold=True)
            ws.merge_cells(
                start_row=start_row_2 + 1,
                start_column=1,
                end_row=start_row_2 + 1,
                end_column=len(df_renov.columns),
            )

            header_row_2 = start_row_2 + 2
            for col_idx in range(1, len(df_renov.columns) + 1):
                c = ws.cell(row=header_row_2, column=col_idx)
                c.font = header_font
                c.fill = header_fill
                c.border = border
                c.alignment = center_alignment
            for r in range(header_row_2 + 1, header_row_2 + 1 + len(df_renov)):
                for cidx in range(1, len(df_renov.columns) + 1):
                    c = ws.cell(row=r, column=cidx)
                    c.font = data_font
                    c.border = border
                    if cidx == 1:
                        c.alignment = left_alignment
                    else:
                        c.alignment = currency_alignment
                        if isinstance(c.value, (int, float)):
                            c.number_format = "$#,##0"

            # Ajuste de anchos para toda la hoja
            for cidx in range(1, len(df_actual.columns) + 1):
                letter = get_column_letter(cidx)
                if cidx == 1:
                    ws.column_dimensions[letter].width = 50
                else:
                    ws.column_dimensions[letter].width = 20

        crear_hoja_riesgos()

        # Generar hoja de Amparos después de Riesgos para que aparezca luego en el libro
        crear_hoja_amparos()

    with open(output_path, "wb") as f:
        f.write(output.getvalue())

    print(f"✅ Archivo Excel generado exitosamente: {output_path}")
    return output_path


def generar_prompt_unico(prompt, grouped_data):
    """
    Genera un único prompt que incluya actual, renovacion y todos los adicionales.
    """

    # Actual
    if "actual" in grouped_data:
        prompt += f"\namparos_actuales = {grouped_data['actual']}\n"

    # Renovacion
    if "renovacion" in grouped_data:
        prompt += f"\namparos_renovacion = {grouped_data['renovacion']}\n"

    # Adicionales: cada subarray separado
    if "adicional" in grouped_data:
        for idx, subarray in enumerate(grouped_data["adicional"], start=1):
            prompt += f"\namparos_adicional_{idx} = {subarray}\n"

    return prompt


def _sanitize_key(name: str) -> str:
    # quitar extensión y normalizar a formato seguro para keys
    base = re.sub(r"\.pdf$", "", name, flags=re.IGNORECASE)
    base = re.sub(r"[^\w]+", "_", base)  # reemplaza cualquier char no alfanum por _
    base = base.strip("_").lower()
    return f"deducible_{base}" if base else "deducible_adicional"


def agregar_deducibles_adicionales(response_schema: dict, data: dict) -> dict:
    """
    Agrega propiedades deducible_<file_name_sanitizado> por cada file_name
    presente en items de 'adicional'. Las marca como required y
    las inserta justo después de 'deducible_renovacion' en propertyOrdering.
    """
    schema = deepcopy(response_schema)
    items = schema.setdefault("items", {})
    properties = items.setdefault("properties", {})
    prop_order = items.setdefault("propertyOrdering", [])
    required = items.setdefault("required", [])

    # Recorrer todos los amparos en 'adicional' (que es una lista de listas)
    vistos = []
    for sublist in data.get("adicional", []):
        for d in sublist:
            if "file_name" in d:
                fn = d["file_name"]
                if fn not in vistos:
                    vistos.append(fn)

    if not vistos:
        return schema  # no hay adicionales, devolver schema tal cual

    # posición base: justo después de 'deducible_renovacion' si existe, si no, al final antes de 'tipo' si existe
    try:
        base_index = prop_order.index("deducible_renovacion") + 1
    except ValueError:
        try:
            tipo_idx = prop_order.index("tipo")
            base_index = tipo_idx
        except ValueError:
            base_index = len(prop_order)

    # insertar cada nuevo campo manteniendo orden relativo
    for i, file_name in enumerate(vistos):
        key = _sanitize_key(file_name)
        if key in properties:
            if key not in required:
                required.append(key)
            if key not in prop_order:
                prop_order.insert(base_index + i, key)
            continue

        properties[key] = {
            "type": "STRING",
            "description": f"Deducible que aplica para el documento adicional '{file_name}'. Si no se encuentra ningún amparo válido, devuelve una cadena vacía.",
        }

        if key not in required:
            required.append(key)

        insert_pos = base_index + i
        if insert_pos > len(prop_order):
            prop_order.append(key)
        else:
            prop_order.insert(insert_pos, key)

    return schema


def calcular_totales_riesgos(riesgos):
    """
    Calcula el total de 'valor_asegurado' por cada 'interes_asegurado'
    a partir de una lista de ubicaciones con sus 'detalle_cobertura'.

    Args:
        riesgos (list[dict]): Estructura con ubicaciones y detalle_cobertura.

    Returns:
        dict: {interes_asegurado: total_valor_asegurado, ..., 'TOTAL_GENERAL': total}
    """
    totales = {}

    for r in riesgos:
        for detalle in r.get("detalle_cobertura", []):
            interes = detalle.get("interes_asegurado")
            valor = detalle.get("valor_asegurado", 0) or 0
            if interes:
                totales[interes] = totales.get(interes, 0) + valor

    return totales


def actualizar_todos_los_valores(data, totales):
    """
    Recorre todo el JSON y reemplaza 'valor_asegurado' según el 'interes_asegurado'.
    """
    for _, items in data.items():
        for item in items:
            interes = item.get("interes_asegurado", "").upper()
            if interes in totales:
                item["valor_asegurado"] = totales[interes]
    return data


def extraer_adicionales(data, max_adicionales=3):
    """
    Devuelve los deducibles adicionales agrupados en arrays por cada deducible_amparos_adicional.

    Args:
        data (list): Lista de diccionarios de amparos.
        max_adicionales (int): Número máximo de deducibles adicionales a revisar por amparo.

    Returns:
        list: Lista de arrays, uno por cada deducible adicional.
    """
    # Creamos un array vacío por cada deducible adicional
    arrays_adicionales = [[] for _ in range(max_adicionales)]

    for a in data:
        for i in range(1, max_adicionales + 1):
            key = f"deducible_amparos_adicional_{i}"
            if key in a and a[key].lower() != "no aplica":
                arrays_adicionales[i - 1].append(
                    {
                        "amparo": a["amparo"],
                        "deducible": a[key],
                        "tipo": a.get("tipo", []),
                    }
                )

    # Eliminamos los arrays vacíos si no hay deducibles para ese índice
    return [arr for arr in arrays_adicionales if arr]


def limpiar_dineros(data):
    """
    Elimina los registros con 'Dineros' de todos los campos no permitidos
    y los mueve automáticamente a 'Incendio' y "Sustracción".
    Si esos campos no existen, los crea.
    """
    campos_permitidos = {"Incendio", "Sustracción"}
    nuevo_data = {}
    dinerios_a_mover = []

    # Paso 1: eliminar 'Dineros' de los campos no permitidos y guardarlos aparte
    for tipo, items in data.items():
        if tipo in campos_permitidos:
            nuevo_data[tipo] = list(items)  # copiamos los registros
        else:
            nuevo_data[tipo] = []
            for item in items:
                if item.get("interes_asegurado") == "Dineros":
                    dinerios_a_mover.append(item)  # guardar para mover
                else:
                    nuevo_data[tipo].append(item)

    # Paso 2: aseguramos que los campos permitidos existan
    for campo in campos_permitidos:
        if campo not in nuevo_data:
            nuevo_data[campo] = []

    # Paso 3: agregar los registros de 'Dineros' a los campos permitidos
    for item in dinerios_a_mover:
        for campo in campos_permitidos:
            nuevo_data[campo].append(item)

    return nuevo_data


def extraer_deducibles(data):
    # Las secciones que queremos transformar
    secciones = [
        "incendio",
        "sustraccion",
        "equipo_electronico",
        "rotura_de_maquinaria",
        "manejo",
        "transporte_de_valores",
        "maquinaria_y_equipo",
        "responsabilidad_civil_amparo",
    ]

    resultado = []

    for seccion in secciones:
        amparos = data["data"].get(seccion, {})
        for amparo, deducible in amparos.items():
            # Convertimos el nombre de la clave a un nombre más legible
            amparo_nombre = amparo.replace("_", " ").title()
            # Ajuste de tipo: capitalizamos palabras y reemplazamos guiones bajos
            tipo = []
            if seccion == "equipo_electronico":
                tipo = ["Equipo Electronico"]
            elif seccion == "rotura_de_maquinaria":
                tipo = ["Rotura de Maquinaria"]
            elif seccion == "transporte_de_valores":
                tipo = ["Transporte de Valores"]
            elif seccion == "manejo":
                tipo = ["Manejo de Dinero"]
            elif seccion == "sustraccion":
                tipo = ["Sustracción"]
            elif seccion == "maquinaria_y_equipo":
                tipo = ["Maquinaria y Equipo"]
            elif seccion == "responsabilidad_civil_amparo":
                tipo = ["Responsabilidad Civil"]
            else:
                tipo = ["Incendio"]

            resultado.append(
                {"amparo": amparo_nombre, "deducible": deducible, "tipo": tipo}
            )

    return resultado


class InvoiceOrchestrator:
    def __init__(self, api_key: str, model: str, nutrient_keys: list[str]):
        self.api_key = api_key
        self.model = model
        if not nutrient_keys:
            raise ValueError("Debes pasar al menos 1 API key NUTRIENT.")
        self.nutrient_keys_cycle = cycle(nutrient_keys)
        # Evitar toasts duplicados por el mismo proceso/archivo
        self._empties_toasted: set[str] = set()

    # Hace requests a la API con reintentos
    async def make_api_request(
        self, url: str, headers: Dict, data: Dict, process_id: str, retries: int = 5
    ) -> Optional[Dict]:
        ssl_context = ssl.create_default_context(cafile=certifi.where())
        connector = aiohttp.TCPConnector(ssl=ssl_context)
        async with aiohttp.ClientSession(connector=connector) as session:
            for i in range(retries):
                try:
                    async with session.post(
                        url, headers=headers, json=data
                    ) as response:
                        if response.status == 200:
                            return await response.json()
                        elif response.status in [429, 529, 503]:
                            sleep_time = 15 * (i + 1)  # Espera incremental en segundos
                            app_logger.warning(
                                f"API request failed with status {response.status}. Retrying in {sleep_time} seconds... res: {response}"
                            )
                            await asyncio.sleep(sleep_time)
                        else:
                            error_data = await response.json()
                            app_logger.error(f"Error data: {error_data}")

                            app_logger.error(
                                f"Error Code: {error_data.get("error").get('code')}"
                            )
                            app_logger.error(
                                f"Error Message: {error_data.get("error").get('message')}"
                            )

                            if (
                                error_data.get("error").get("code") == 400
                                and error_data.get("error").get("message")
                                == "The document has no pages."
                            ):
                                # Mostrar toast solo una vez por process_id
                                if process_id not in self._empties_toasted:
                                    st.toast(
                                        f"⚠️ El archivo {process_id} está vacío. No tiene páginas. Se omitirá.",
                                        icon="⚠️",
                                    )
                                    self._empties_toasted.add(process_id)
                                return
                            error_str = await response.text()
                            # app_logger.error(f"Error str: {error_str}")
                            app_logger.error(
                                f"API request failed with status {response.status} - {error_str}"
                            )
                            app_logger.error(
                                f"Headers: {headers}, url: {url}, process_id: {process_id}"
                            )
                            raise ValueError(
                                f"Request failed with status {response.status}"
                            )
                except aiohttp.ClientError as e:
                    raise ValueError(f"Request error: {str(e)}")
        raise ValueError("Max retries exceeded.")

    async def handle_pdf(self, item: QueueItem, prompt: str, responseSchema: dict):
        encoded_pdf = item["b64_str"]

        # URL del endpoint de la API de Gemini para generar contenido a partir del modelo especificado
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.model}:generateContent"

        # Encabezados HTTP: autenticación con la clave de API y tipo de contenido JSON
        headers = {"x-goog-api-key": self.api_key, "Content-Type": "application/json"}

        # Cuerpo de la petición: incluye el PDF codificado y el texto del prompt
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "inline_data": {
                                "mime_type": "application/pdf",
                                "data": encoded_pdf,
                            }
                        },
                        {"text": prompt},
                    ]
                }
            ],
            "generationConfig": {
                "responseMimeType": "application/json",  # Forzar que la respuesta sea JSON
                "responseSchema": responseSchema,  # Esquema JSON que debe respetar la respuesta
            },
        }

        # Realizar la petición asíncrona a la API con reintentos incluidos
        response = await self.make_api_request(url, headers, payload, item["file_name"])

        # Almacenar la respuesta de la API dentro del item bajo la clave "data"
        return response

    async def batch_processor(
        self, input_files: List[Dict], prompt: str, responseSchema: dict
    ):
        if not input_files:
            return

        base64_docs = []

        for file in input_files:
            base64_docs.append(file["base64"])

        # # Construir payload
        files_metadata_b64 = [
            {"inline_data": {"mime_type": "application/pdf", "data": doc}}
            for doc in base64_docs
        ]

        url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.model}:generateContent?key={self.api_key}"

        headers = {"Content-Type": "application/json"}

        payload = {
            "contents": [{"parts": files_metadata_b64 + [{"text": prompt}]}],
            "generationConfig": {
                "responseMimeType": "application/json",
                "responseSchema": responseSchema,
            },
        }

        response = await self.make_api_request(
            url=url, headers=headers, data=payload, process_id=uuid.uuid4().hex
        )

        response["doc_type"] = input_files[0]["doc_type"]

        return response

    def uploaded_file_to_base64(self, uploaded_file: UploadedFile) -> str:
        """
        Convierte un UploadedFile de Streamlit en un string Base64.

        Args:
            uploaded_file (UploadedFile): Archivo subido por el usuario.

        Returns:
            str: Contenido del archivo en Base64.
        """
        # Leer los bytes del archivo
        file_bytes = uploaded_file.read()

        # Codificar a Base64 y convertir a string UTF-8
        b64_str = base64.b64encode(file_bytes).decode("utf-8")

        return b64_str

    def obtener_nombres_archivos(self, archivos):
        """
        Recibe una lista de UploadedFile (Streamlit) y devuelve
        un string con todos los nombres de archivo separados por comas.
        """
        if not archivos:
            return ""
        return ", ".join([archivo.name for archivo in archivos])

    async def execute_toolset(self, item: QueueItem, tools: List[Dict]):
        file_name = item.get("file_name")
        doc_type = item.get("doc_type")

        results = await asyncio.gather(
            *[
                self.handle_pdf(
                    item=item, prompt=tool["prompt"], responseSchema=tool["data"]
                )
                for tool in tools
            ]
        )

        merged = {}
        for item_res in results:
            if not item_res:
                continue
            data = json.loads(
                item_res.get("candidates")[0].get("content").get("parts")[0].get("text")
            )
            merged.update(data)

        tokens = self.sumar_tokens_por_tipo([r for r in results if r])

        return {
            "file_name": file_name,
            "doc_type": doc_type,
            "data": merged,
            "tokens": tokens,
        }

    async def execute_multiple(self, item: QueueItem, tools: List[Dict]):
        file_name = item.get("file_name")
        doc_type = item.get("doc_type")

        base64_pdfs = item.get("b64_str")

        if isinstance(base64_pdfs, list):
            files_metadata_b64 = [
                {"inline_data": {"mime_type": "application/pdf", "data": doc}}
                for doc in base64_pdfs
            ]
        else:
            return

        url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.model}:generateContent?key={self.api_key}"

        headers = {"Content-Type": "application/json"}

        payloads = [
            {
                "contents": [
                    {"parts": files_metadata_b64 + [{"text": tool["prompt"]}]}
                ],
                "generationConfig": {
                    "responseMimeType": "application/json",
                    "responseSchema": tool["data"],
                },
            }
            for tool in tools
        ]

        results = await asyncio.gather(
            *[
                self.make_api_request(
                    url=url,
                    headers=headers,
                    data=payload,
                    process_id=item.get("file_name", "Dinardi"),
                )
                for payload in payloads
            ]
        )

        merged = {}
        for item_res in results:
            if not item_res:
                continue
            data = json.loads(
                item_res.get("candidates")[0].get("content").get("parts")[0].get("text")
            )
            merged.update(data)

        tokens = self.sumar_tokens_por_tipo([r for r in results if r])

        return {
            "file_name": file_name,
            "doc_type": doc_type,
            "data": merged,
            "tokens": tokens,
        }

    def sumar_tokens_por_tipo(self, data: List[Dict]) -> Dict[str, int]:
        """
        Suma los tokens por tipo en una lista de diccionarios que contienen 'usageMetadata'.
        Retorna un diccionario con el total de cada tipo de token.
        """
        total_tokens = {}

        for item in data:
            usage = item.get("usageMetadata", {})
            for key, value in usage.items():
                if key not in total_tokens:
                    total_tokens[key] = 0
                # A veces el valor puede no ser int (como un dict interno), entonces chequeamos
                if isinstance(value, int):
                    total_tokens[key] += value
                elif isinstance(value, dict):
                    # Sumamos los valores internos si es un diccionario
                    for sub_val in value.values():
                        if isinstance(sub_val, int):
                            total_tokens[key] += sub_val
        return total_tokens

    def unificar_segmentos(self, actual, renovacion):
        # Paso 1: construir mapa de interes -> tipos (segmentos)
        mapa_actual = {}
        for segmento, items in actual.items():
            for item in items:
                ia = item["interes_asegurado"]
                mapa_actual.setdefault(ia, set()).add(segmento)

        mapa_renov = {}
        for segmento, items in renovacion.items():
            for item in items:
                ia = item["interes_asegurado"]
                mapa_renov.setdefault(ia, set()).add(segmento)

        # Paso 2: obtener la union de segmentos para cada interes
        intereses = set(mapa_actual.keys()) | set(mapa_renov.keys())
        union_map = {}
        for ia in intereses:
            union_map[ia] = mapa_actual.get(ia, set()) | mapa_renov.get(ia, set())

        # Paso 3: reconstruir actual y renovacion con los segmentos unificados
        def aplicar_union(diccionario, mapa_union):
            nuevo_dicc = {seg: [] for seg in diccionario.keys()}
            for segmento, items in diccionario.items():
                for item in items:
                    ia = item["interes_asegurado"]
                    # expandir a todos los segmentos que le tocan
                    for seg in mapa_union[ia]:
                        nuevo_item = item.copy()
                        nuevo_item["tipo"] = list(mapa_union[ia])
                        nuevo_dicc.setdefault(seg, []).append(nuevo_item)
            return nuevo_dicc

        actual_unificado = aplicar_union(actual, union_map)
        renov_unificado = aplicar_union(renovacion, union_map)

        return actual_unificado, renov_unificado

    async def action_item_tool(self, prompt: str, responseSchema: dict):
        # URL del endpoint de la API de Gemini para generar contenido a partir del modelo especificado
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.model}:generateContent"

        # Encabezados HTTP: autenticación con la clave de API y tipo de contenido JSON
        headers = {"x-goog-api-key": self.api_key, "Content-Type": "application/json"}

        # Cuerpo de la petición: incluye el PDF codificado y el texto del prompt
        payload = {
            "contents": [
                {
                    "parts": [
                        {"text": prompt},
                    ]
                }
            ],
            "generationConfig": {
                "responseMimeType": "application/json",  # Forzar que la respuesta sea JSON
                "responseSchema": responseSchema,  # Esquema JSON que debe respetar la respuesta
            },
        }

        # Realizar la petición asíncrona a la API con reintentos incluidos
        response = await self.make_api_request(url, headers, payload, uuid.uuid4().hex)

        # Almacenar la respuesta de la API dentro del item bajo la clave "data"
        return response

    async def excel2pdf(
        self, input_file: str, output_folder: str, output_filename: str = None
    ):
        try:
            api_key = next(self.nutrient_keys_cycle)

            os.makedirs(output_folder, exist_ok=True)

            if output_filename is None:
                base = os.path.splitext(os.path.basename(input_file))[0]
                output_filename = f"{base}.pdf"

            output_path = os.path.join(output_folder, output_filename)

            url = "https://api.nutrient.io/processor/convert_to_pdf"

            headers = {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Authorization": f"Bearer {api_key}",
            }

            print(f"[DEBUG] Sending request to {url} with file {input_file}")

            # Leer archivo
            with open(input_file, "rb") as f:
                file_bytes = f.read()

            async with aiohttp.ClientSession() as session:
                async with session.post(url, headers=headers, data=file_bytes) as res:

                    print(f"[DEBUG] Status {res.status} para {input_file}")

                    # Key sin créditos → probar siguiente
                    if res.status == 402:
                        app_logger.error(
                            f"[ERROR] Error {res.status}: No créditos disponibles para {api_key}"
                        )
                        return await self.excel2pdf(
                            input_file, output_folder, output_filename
                        )

                    if res.status == 400:
                        text = await res.text()
                        app_logger.error(
                            f"[ERROR] Error {res.status}: {text}, input_file: {input_file}"
                        )
                        raise Exception(f"Error {res.status}: {text}")

                    if res.status != 200:
                        text = await res.text()
                        app_logger.error(
                            f"[ERROR] Error {res.status}: {text}, input_file: {input_file}"
                        )
                        raise Exception(f"Error {res.status}: {text}")

                    pdf_bytes = await res.read()

            with open(output_path, "wb") as f:
                f.write(pdf_bytes)

            return output_path

        except Exception as e:
            app_logger.exception(
                f"[EXCEPTION] Error en excel2pdf: {e}, input_file: {input_file}"
            )
            # return

    # Helper: convertir UploadedFile .xlsx a PDF y devolver base64 del PDF
    async def to_pdf_base64(self, uploaded_file: UploadedFile):
        app_logger.info(
            f"[DEBUG] Iniciando conversión para {uploaded_file.name}, mime: {uploaded_file.type}"
        )
        name_lower = uploaded_file.name.lower()
        mime = getattr(uploaded_file, "type", "") or ""

        is_xlsx = name_lower.endswith(".xlsx") or (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in mime
        )

        # Si no es Excel, convertir directamente a base64 con tu helper
        if not is_xlsx:
            app_logger.info(
                f"[DEBUG] No es Excel, usando base64 directo para {uploaded_file.name}"
            )
            return orchestrator.uploaded_file_to_base64(uploaded_file)

        temp_dir = None  # importante para poder borrarlo en el finally

        try:
            # Crear carpeta temporal única
            temp_id = uuid.uuid4().hex
            temp_dir = os.path.join("converted_pdfs", f"tmp_{temp_id}")
            os.makedirs(temp_dir, exist_ok=True)

            app_logger.info(
                f"[DEBUG] Es Excel, convirtiendo {uploaded_file.name} a PDF"
            )

            # Guardar el .xlsx temporalmente
            xlsx_path = os.path.join(temp_dir, uploaded_file.name)
            with open(xlsx_path, "wb") as f:
                f.write(uploaded_file.getvalue())

            # Convertir XLSX → PDF
            pdf_path = await orchestrator.excel2pdf(
                input_file=xlsx_path, output_folder=temp_dir
            )

            app_logger.info(f"[DEBUG] PDF convertido: {pdf_path}")

            # Leer PDF generado
            with open(pdf_path, "rb") as f:
                pdf_bytes = f.read()

            app_logger.info(
                f"[DEBUG] Es Excel, convirtiendo a PDF para {uploaded_file.name}"
            )

            # Base64 final
            return base64.b64encode(pdf_bytes).decode("utf-8")

        except Exception as e:
            app_logger.error(
                f"[DEBUG] Error en to_pdf_base64 para {uploaded_file.name}: {e}"
            )
            # Puedes loggear si quieres
            app_logger.error(f"[ERROR] No se pudo convertir Excel a PDF: {e}")
            # raise

        finally:
            # Borrar carpeta temporal si existe
            if temp_dir and os.path.exists(temp_dir):
                try:
                    shutil.rmtree(temp_dir)
                except Exception as cleanup_error:
                    app_logger.warning(
                        f"[WARN] No se pudo borrar carpeta temporal {temp_dir}: {cleanup_error}"
                    )


orchestrator = InvoiceOrchestrator(
    api_key=os.getenv("GEMINI_API_KEY"),
    model="gemini-2.5-flash",
    nutrient_keys=[os.getenv("NUTRIENT_KEY_1"), os.getenv("NUTRIENT_KEY_2")],
)

coberturas_predeterminadas = {
    "Incendio": [
        {"amparo": "Incendio y/o Rayo", "tipo": ["Incendio"]},
        {"amparo": "Explosión", "tipo": ["Incendio"]},
        {"amparo": "Terremoto, temblor", "tipo": ["Incendio"]},
        {"amparo": "Asonada, motín, conm. Civil/popular huelga", "tipo": ["Incendio"]},
        {"amparo": "Extensión de amparos", "tipo": ["Incendio"]},
        {"amparo": "Daños por agua", "tipo": ["Incendio"]},
        {"amparo": "Anegación", "tipo": ["Incendio"]},
        {"amparo": "Daños a caldera y AP. Generadores vapor", "tipo": ["Incendio"]},
        {"amparo": "Incendio y/o Rayo en aparatos electricos", "tipo": ["Incendio"]},
        {"amparo": "Remoción de Escombros", "tipo": ["Incendio"]},
        {"amparo": "Actos Terroristas", "tipo": ["Incendio"]},
    ],
    "Sustracción": [
        {"amparo": "Hurto Calificado", "tipo": ["Sustracción"]},
        {"amparo": "Asalto", "tipo": ["Sustracción"]},
        {"amparo": "Atraco", "tipo": ["Sustracción"]},
    ],
    "Equipo Electrónico": [
        {"amparo": "Pérdidas o daños accidentales", "tipo": ["Equipo Electronico"]},
        {"amparo": "Terremoto- Temblor", "tipo": ["Equipo Electronico"]},
        {
            "amparo": "Asonada, motín, conm. Civil/popular huelga",
            "tipo": ["Equipo Electronico"],
        },
        {
            "amparo": "Actos mal intecionados de terceros",
            "tipo": ["Equipo Electronico"],
        },
        {"amparo": "Hurto", "tipo": ["Equipo Electronico"]},
    ],
    "Rotura de Maquinaria": [
        {"amparo": "Excesos de tension electrica", "tipo": ["Rotura de Maquinaria"]},
        {
            "amparo": "Impericia o Negligencia de los empleados",
            "tipo": ["Rotura de Maquinaria"],
        },
        {
            "amparo": "Actos mal intencionados de los empleados",
            "tipo": ["Rotura de Maquinaria"],
        },
        {
            "amparo": "Explosion quimica interna, incendio interno, Caida de rayo",
            "tipo": ["Rotura de Maquinaria"],
        },
        {"amparo": "Obstrucciones Internas", "tipo": ["Rotura de Maquinaria"]},
        {
            "amparo": "Rotura debido a Fuerza centrifuga",
            "tipo": ["Rotura de Maquinaria"],
        },
    ],
    "Manejo de Dinero": [
        {"amparo": "Hurto calificado", "tipo": ["Manejo de Dinero"]},
        {"amparo": "Abuso de confianza", "tipo": ["Manejo de Dinero"]},
        {"amparo": "Desfalco, Estafa, Falsificación", "tipo": ["Manejo de Dinero"]},
        {
            "amparo": "Pérdidas por personas no identificadas",
            "tipo": ["Manejo de Dinero"],
        },
        {"amparo": "Perdidas por personal temporal", "tipo": ["Manejo de Dinero"]},
    ],
    "Responsabilidad Civil": [
        {
            "amparo": "Predio labores y operaciones, incluyendo Incendio y Explosión",
            "tipo": ["Responsabilidad Civil"],
        },
        {"amparo": "Gastos Médicos", "tipo": ["Responsabilidad Civil"]},
        {
            "amparo": "R.C. Contratistas y Subcontratistas",
            "tipo": ["Responsabilidad Civil"],
        },
        {
            "amparo": "R.C. Vehiculos Propios y No Propios",
            "tipo": ["Responsabilidad Civil"],
        },
        {
            "amparo": "R.C. Productos y trabajos terminados",
            "tipo": ["Responsabilidad Civil"],
        },
        {"amparo": "R.C. Parqueaderos", "tipo": ["Responsabilidad Civil"]},
        {
            "amparo": "Bienes bajo cuidado, tenencia y control",
            "tipo": ["Responsabilidad Civil"],
        },
        {"amparo": "Responsabilidad Civil Patronal", "tipo": ["Responsabilidad Civil"]},
    ],
    "Transporte de Valores": [
        {
            "amparo": "Daños ocasionados a los titulos valores con ocasión de su transporte. Falta de Entrega (Hurto Total). Averia Particular. Saqueo (Hurto Parcial)",
            "tipo": ["Transporte de Valores"],
        },
        {
            "amparo": "Huelga (Terrorismo para los titulos valores movilizados dentro del país)",
            "tipo": ["Transporte de Valores"],
        },
    ],
    # "Maquinaria y Equipo de Contratista": [
    #     {
    #         "amparo": "Minitractor Jhon Deere Golf & Turf modelo 2020",
    #         "tipo": ["Maquinaria y Equipo de Contratista"],
    #     },
    #     {
    #         "amparo": "Responsabilidad Civil",
    #         "tipo": ["Maquinaria y Equipo de Contratista", "Responsabilidad Civil"],
    #     },
    # ],
}

# Configuración de la página
st.set_page_config(
    page_title="Análisis de pólizas",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)


async def main():
    # Barra lateral
    with st.sidebar:
        # Campo 1: Póliza Actual
        st.subheader("1️⃣ Póliza Actual")
        archivo_poliza_actual = st.file_uploader(
            "Cargar archivo de póliza actual",
            type=["pdf", "docx", "txt", "jpg", "png", "xlsx"],
            key="poliza_actual",
            help="Sube el archivo de tu póliza actual para extraer sus datos",
            accept_multiple_files=True,
        )

        st.markdown("---")

        # Campo 2: Póliza de Renovación
        st.subheader("2️⃣ Póliza de Renovación")
        archivo_poliza_renovacion = st.file_uploader(
            "Cargar archivo de póliza de renovación",
            type=["pdf", "docx", "txt", "jpg", "png", "xlsx"],
            key="poliza_renovacion",
            help="Sube el archivo de la póliza de renovación para comparar",
            accept_multiple_files=True,
        )

        st.markdown("---")

        # Campo 3: Múltiples Documentos
        st.subheader("3️⃣ Documentos Adicionales")
        archivos_multiples = st.file_uploader(
            "Cargar múltiples documentos",
            type=["pdf", "docx", "txt", "jpg", "png", "xlsx"],
            accept_multiple_files=True,
            key="documentos_multiples",
            help="Sube múltiples documentos para extraer las primas de cada uno",
        )

        st.markdown("---")

        # campo 4: Conjunto de documentos
        st.subheader("4️⃣ Conjunto de Documentos adicionales")
        archivos_conjuntos_1 = st.file_uploader(
            "Cargar conjunto de documentos",
            type=["pdf", "docx", "txt", "jpg", "png", "xlsx"],
            accept_multiple_files=True,
            key="archivos_conjuntos",
            help="Sube un conjunto de documentos para extraer información de todos ellos",
        )

        st.markdown("---")

        # campo 5: Conjunto de documentos
        st.subheader("5️⃣ Conjunto de Documentos adicionales")
        archivos_conjuntos_2 = st.file_uploader(
            "Cargar conjunto de documentos",
            type=["pdf", "docx", "txt", "jpg", "png", "xlsx"],
            accept_multiple_files=True,
            key="archivos_conjuntos_2",
            help="Sube un conjunto de documentos para extraer información de todos ellos",
        )

        st.markdown("---")

        # campo 6: Conjunto de documentos
        st.subheader("6️⃣ Conjunto de Documentos adicionales")
        archivos_conjuntos_3 = st.file_uploader(
            "Cargar conjunto de documentos",
            type=["pdf", "docx", "txt", "jpg", "png", "xlsx"],
            accept_multiple_files=True,
            key="archivos_conjuntos_3",
            help="Sube un conjunto de documentos para extraer información de todos ellos",
        )

        st.markdown("---")

        debug = st.toggle("Debug Mode", value=False)

    # TODO: Hacer que si es xlsx, se convierta a pdf
    if st.sidebar.button(
        "🚀 Iniciar Proceso",
        type="primary",
        use_container_width=True,
        help="Haz clic para iniciar el procesamiento de todos los archivos cargados",
    ):
        tasks = []
        archivos_conjuntos_actuales_renovacion = []
        poliza_actual_item = None
        poliza_renovacion_item = None
        if archivo_poliza_actual:
            if len(archivo_poliza_actual) > 1:
                # Convertir cada archivo a PDF base64 si es .xlsx, sino usar base64 directo
                b64_list = [
                    await orchestrator.to_pdf_base64(archivo)
                    for archivo in archivo_poliza_actual
                ]
                archivos_conjuntos_actuales_renovacion.append(
                    QueueItem(
                        file_name=orchestrator.obtener_nombres_archivos(
                            archivo_poliza_actual
                        ),
                        file_extension="pdf",
                        b64_str=b64_list,
                        media_type="application/pdf",
                        process_id=uuid.uuid4().hex,
                        doc_type="actual",
                    )
                )
            else:
                archivo_poliza_actual_unico = archivo_poliza_actual[0]
                # Si es .xlsx convertir a PDF; de lo contrario usar base64 original
                is_xlsx_actual = (
                    archivo_poliza_actual_unico.name.lower().endswith(".xlsx")
                    or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    in archivo_poliza_actual_unico.type
                )
                if is_xlsx_actual:
                    b64_pdf = await orchestrator.to_pdf_base64(
                        archivo_poliza_actual_unico
                    )
                    poliza_actual_item = QueueItem(
                        file_name=archivo_poliza_actual_unico.name,
                        file_extension="pdf",
                        b64_str=b64_pdf,
                        media_type="application/pdf",
                        process_id=uuid.uuid4().hex,
                        doc_type="actual",
                    )
                else:
                    poliza_actual_item = QueueItem(
                        file_name=archivo_poliza_actual_unico.name,
                        file_extension=archivo_poliza_actual_unico.type.split("/")[-1],
                        b64_str=orchestrator.uploaded_file_to_base64(
                            archivo_poliza_actual_unico
                        ),
                        media_type=archivo_poliza_actual_unico.type,
                        process_id=uuid.uuid4().hex,
                        doc_type="actual",
                    )
                poliza_actual_process = orchestrator.execute_toolset(
                    item=poliza_actual_item, tools=tools_actuales
                )
                tasks.append(poliza_actual_process)

        if archivo_poliza_renovacion:
            if len(archivo_poliza_renovacion) > 1:
                b64_list = [
                    await orchestrator.to_pdf_base64(archivo)
                    for archivo in archivo_poliza_renovacion
                ]
                archivos_conjuntos_actuales_renovacion.append(
                    QueueItem(
                        file_name=orchestrator.obtener_nombres_archivos(
                            archivo_poliza_renovacion
                        ),
                        file_extension="pdf",
                        b64_str=b64_list,
                        media_type="application/pdf",
                        process_id=uuid.uuid4().hex,
                        doc_type="renovacion",
                    )
                )
            else:
                archivo_poliza_renovacion_unico = archivo_poliza_renovacion[0]
                is_xlsx_renov = (
                    archivo_poliza_renovacion_unico.name.lower().endswith(".xlsx")
                    or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    in archivo_poliza_renovacion_unico.type
                )
                if is_xlsx_renov:
                    b64_pdf = await orchestrator.to_pdf_base64(
                        archivo_poliza_renovacion_unico
                    )
                    poliza_renovacion_item = QueueItem(
                        file_name=archivo_poliza_renovacion_unico.name,
                        file_extension="pdf",
                        b64_str=b64_pdf,
                        media_type="application/pdf",
                        process_id=uuid.uuid4().hex,
                        doc_type="renovacion",
                    )
                else:
                    poliza_renovacion_item = QueueItem(
                        file_name=archivo_poliza_renovacion_unico.name,
                        file_extension=archivo_poliza_renovacion_unico.type.split("/")[
                            -1
                        ],
                        b64_str=orchestrator.uploaded_file_to_base64(
                            archivo_poliza_renovacion_unico
                        ),
                        media_type=archivo_poliza_renovacion_unico.type,
                        process_id=uuid.uuid4().hex,
                        doc_type="renovacion",
                    )
                poliza_renovacion_process = orchestrator.execute_toolset(
                    item=poliza_renovacion_item, tools=tools_actuales
                )
                tasks.append(poliza_renovacion_process)

        documentos_adicionales_items = []
        if archivos_multiples:
            # Crear un QueueItem para cada archivo Multiple
            for archivo in archivos_multiples:
                is_xlsx_multi = (
                    archivo.name.lower().endswith(".xlsx")
                    or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    in archivo.type
                )
                if is_xlsx_multi:
                    b64_pdf = await orchestrator.to_pdf_base64(archivo)
                    documentos_adicionales_items.append(
                        QueueItem(
                            file_name=archivo.name,
                            file_extension="pdf",
                            b64_str=b64_pdf,
                            media_type="application/pdf",
                            process_id=uuid.uuid4().hex,
                            doc_type="adicional",
                        )
                    )
                else:
                    documentos_adicionales_items.append(
                        QueueItem(
                            file_name=archivo.name,
                            file_extension=archivo.type.split("/")[-1],
                            b64_str=orchestrator.uploaded_file_to_base64(archivo),
                            media_type=archivo.type,
                            process_id=uuid.uuid4().hex,
                            doc_type="adicional",
                        )
                    )
        archivos_conjuntos_items = []
        if archivos_conjuntos_1:
            b64_list = [
                await orchestrator.to_pdf_base64(archivo)
                for archivo in archivos_conjuntos_1
            ]
            archivos_conjuntos_items.append(
                QueueItem(
                    file_name=orchestrator.obtener_nombres_archivos(
                        archivos_conjuntos_1
                    ),
                    file_extension="pdf",
                    b64_str=b64_list,
                    media_type="application/pdf",
                    process_id=uuid.uuid4().hex,
                    doc_type="conjunto",
                )
            )

        if archivos_conjuntos_2:
            b64_list = [
                await orchestrator.to_pdf_base64(archivo)
                for archivo in archivos_conjuntos_2
            ]
            archivos_conjuntos_items.append(
                QueueItem(
                    file_name=orchestrator.obtener_nombres_archivos(
                        archivos_conjuntos_2
                    ),
                    file_extension="pdf",
                    b64_str=b64_list,
                    media_type="application/pdf",
                    process_id=uuid.uuid4().hex,
                    doc_type="conjunto",
                )
            )

        if archivos_conjuntos_3:
            b64_list = [
                await orchestratorto_pdf_base64(archivo)
                for archivo in archivos_conjuntos_3
            ]
            archivos_conjuntos_items.append(
                QueueItem(
                    file_name=orchestrator.obtener_nombres_archivos(
                        archivos_conjuntos_3
                    ),
                    file_extension="pdf",
                    b64_str=b64_list,
                    media_type="application/pdf",
                    process_id=uuid.uuid4().hex,
                    doc_type="conjunto",
                )
            )

        # Crear lista con todos los archivos cargados
        all_items = []
        if archivo_poliza_actual and poliza_actual_item:
            all_items.append(poliza_actual_item)
        if archivo_poliza_renovacion and poliza_renovacion_item:
            all_items.append(poliza_renovacion_item)
        if documentos_adicionales_items:
            all_items.extend(documentos_adicionales_items)
        if archivos_conjuntos_items:
            all_items.extend(archivos_conjuntos_items)
        if archivos_conjuntos_actuales_renovacion:
            all_items.extend(archivos_conjuntos_actuales_renovacion)

        # Mostrar DataFrame minimalista
        if all_items:
            st.subheader("📄 Archivos Cargados")
            df_summary = pd.DataFrame(
                [
                    {"Nombre": item["file_name"], "Tipo": item["doc_type"]}
                    for item in all_items
                ]
            )
            st.dataframe(df_summary, use_container_width=True)

        if documentos_adicionales_items:
            for item in documentos_adicionales_items:
                tasks.append(orchestrator.execute_toolset(item, tools_adicionales))

        if archivos_conjuntos_items:
            for item in archivos_conjuntos_items:
                tasks.append(orchestrator.execute_multiple(item, tools_adicionales))

        if archivos_conjuntos_actuales_renovacion:
            for item in archivos_conjuntos_actuales_renovacion:
                tasks.append(orchestrator.execute_multiple(item, tools_actuales))

        results = await asyncio.gather(*tasks)

        if debug:
            with st.expander("results"):
                st.write(results)

        poliza_actual = None
        poliza_renovacion = None
        documentos_adicionales = []

        amparos_adicionales = []

        nombres_de_asegurados = []

        for item in results:
            if item["doc_type"] == "actual":
                poliza_actual = item
                nombres_de_asegurados.append(item.get("data", {}).get("asegurado", ""))
                # if debug:
                #     mostrar_poliza(poliza_actual)
            elif item["doc_type"] == "renovacion":
                poliza_renovacion = item
                nombres_de_asegurados.append(item.get("data", {}).get("asegurado", ""))
                # if debug:
                #     mostrar_poliza(poliza_renovacion)
            elif item["doc_type"] == "adicional" or item["doc_type"] == "conjunto":
                documentos_adicionales.append(item)
                nombres_de_asegurados.append(item.get("data", {}).get("asegurado", ""))

                # amparos_adicionales.append(
                #     {
                #         "archivo": item.get("file_name"),
                #         "amparos": item.get("data").get("amparos"),
                #     }
                # )

        excel_title_prompt = generar_prompt_nombres(nombres_de_asegurados)
        excel_title_schema = {
            "type": "OBJECT",
            "properties": {
                "name": {
                    "description": "Crea un nombre representativo para el asegurado basado en la información proporcionada",
                    "title": "Nombre del Asegurado",
                    "type": "string",
                },
            },
            "required": ["name"],
        }

        excel_name_json = await orchestrator.action_item_tool(
            prompt=excel_title_prompt, responseSchema=excel_title_schema
        )

        content_str = excel_name_json["candidates"][0]["content"]["parts"][0]["text"]
        content_json = json.loads(content_str)  # parsea el string JSON
        excel_name = content_json["name"]

        if debug:
            with st.expander("excel_name_json"):
                st.write(excel_name_json)

        if debug and nombres_de_asegurados:
            with st.expander("Nombres de asegurados"):
                st.write(nombres_de_asegurados)
            with st.expander("Excel name"):
                st.write(excel_name)

        riesgos_actuales = poliza_actual.get("data", []).get("riesgos", [])
        riesgos_renovacion = poliza_renovacion.get("data", []).get("riesgos", [])

        poliza_actual["data"]["detalle_cobertura"] = flatten_detalle_cobertura(
            riesgos_actuales
        )
        poliza_renovacion["data"]["detalle_cobertura"] = flatten_detalle_cobertura(
            riesgos_renovacion
        )

        poliza_actual["data"]["amparos"] = extraer_deducibles(poliza_actual)
        poliza_renovacion["data"]["amparos"] = extraer_deducibles(poliza_renovacion)

        for item in documentos_adicionales:
            deducibles = extraer_deducibles(item)
            item["data"]["amparos"] = deducibles
            amparos_adicionales.append(
                {
                    # "file_name": item.get("file_name"),
                    "amparos": deducibles,
                    "archivo": item.get("data", {}).get("proveedor_poliza") or item.get("file_name")
                }
            )

        if debug:
            with st.expander("Poliza actual"):
                st.write(poliza_actual)
            with st.expander("Poliza de renovacion"):
                st.write(poliza_renovacion)
            with st.expander("Documentos adicionales"):
                st.write(documentos_adicionales)
            with st.expander("Amparos adicionales"):
                st.write(amparos_adicionales)

        # Crear excel y poder descargalo.
        if poliza_actual or poliza_renovacion or documentos_adicionales:
            # st.subheader("📥 Descargar Resultados")
            # Archivo Excel principal
            main_excel_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            main_excel_path = main_excel_file.name
            main_excel_file.close()

            # Archivo Excel resumen
            summary_excel_file = tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False
            )
            summary_excel_path = summary_excel_file.name
            summary_excel_file.close()

            amparos_actuales = {
                "archivo": poliza_actual.get("file_name"),
                "amparos": poliza_actual.get("data", []).get("amparos", []),
            }

            amparos_renovacion = {
                "archivo": poliza_renovacion.get("file_name"),
                "amparos": poliza_renovacion.get("data", []).get("amparos", []),
            }

            # Clasificaciones por tipo para el resumen RC (no sobrescribir AmparosDict)
            amparos_actuales_por_tipo = clasificar_por_tipo(
                poliza_actual.get("data", {}).get("amparos", {})
            )
            amparos_renovacion_por_tipo = clasificar_por_tipo(
                poliza_renovacion.get("data", {}).get("amparos", {})
            )

            clasificacion_actual = clasificar_por_tipo(
                poliza_actual.get("data", {}).get("detalle_cobertura", {})
            )

            clasificacion_renovacion = clasificar_por_tipo(
                poliza_renovacion.get("data", {}).get("detalle_cobertura", {})
            )

            if debug:
                with st.expander("Riesgos actuales"):
                    st.write(riesgos_actuales)

                with st.expander("Riesgos renovacion"):
                    st.write(riesgos_renovacion)

                with st.expander("clasificacion_actual"):
                    st.write(clasificacion_actual)

                with st.expander("clasificacion_renovacion"):
                    st.write(clasificacion_renovacion)

                with st.expander("documentos_adicionales"):
                    st.write(documentos_adicionales)

                with st.expander("Poliza actual"):
                    st.write(poliza_actual)

                with st.expander("Poliza renovacion"):
                    st.write(poliza_renovacion)

            docs_adicionales_data = [
                {
                    "Archivo": doc.get("file_name"),
                    "Tipo de Documento": doc.get("doc_type"),
                    "Prima Sin IVA": doc.get("data", {}).get("prima_sin_iva", ""),
                    # "IVA": doc.get("data", {}).get("iva", ""),
                    # "Prima Con IVA": doc.get("data", {}).get("prima_con_iva", ""),
                    "tasa": doc.get("data", {}).get("tasa", ""),
                    "amparos": doc.get("data", {}).get("amparos", []),
                    "danos_materiales": doc.get("data", {}).get("danos_materiales", {}),
                    "manejo_global_comercial": doc.get("data", {}).get(
                        "manejo_global_comercial", {}
                    ),
                    "transporte_valores": doc.get("data", {}).get(
                        "transporte_valores", {}
                    ),
                    "responsabilidad_civil": doc.get("data", {}).get(
                        "responsabilidad_civil", {}
                    ),
                }
                for doc in documentos_adicionales
            ]

            if debug:
                with st.expander("docs_adicionales_data"):
                    st.write(docs_adicionales_data)
                with st.expander("amparos actuales"):
                    st.write(amparos_actuales)
                with st.expander("amparos renovacion"):
                    st.write(amparos_renovacion)
                with st.expander("amparos_adicionales"):
                    st.write(amparos_adicionales)
                with st.expander("Amparos actuales por tipo"):
                    st.write(amparos_actuales_por_tipo)
                with st.expander("Amapros renovacion por tipo"):
                    st.write(amparos_renovacion_por_tipo)

                amparos_actuales["archivo"] = poliza_actual["data"]["proveedor_poliza"]
                amparos_renovacion["archivo"] = poliza_renovacion["data"]["proveedor_poliza"]

            try:

                main_output_path = generar_excel_analisis_polizas(
                    riesgos_actuales=riesgos_actuales,
                    riesgos_renovacion=riesgos_renovacion,
                    amparos_actuales=amparos_actuales,
                    amparos_renovacion=amparos_renovacion,
                    amparos_adicionales=amparos_adicionales,
                    output_path=main_excel_path,
                    titulo_excel=excel_name,
                )

                actual_u, renovacion_u = orchestrator.unificar_segmentos(
                    clasificacion_actual, clasificacion_renovacion
                )

                totales_actuales = calcular_totales_riesgos(riesgos_actuales)
                totales_renovacion = calcular_totales_riesgos(riesgos_renovacion)

                actual_u_actualizado = actualizar_todos_los_valores(
                    actual_u, totales_actuales
                )
                renovacion_u_actualizado = actualizar_todos_los_valores(
                    renovacion_u, totales_renovacion
                )

                if debug:
                    with st.expander("actual_u_actualizado"):
                        st.write(actual_u_actualizado)
                    with st.expander("renovacion_u_actualizado"):
                        st.write(renovacion_u_actualizado)

                if debug:
                    st.write("Debugging")
                    with st.expander("actual_u_actualizado"):
                        st.write(limpiar_dineros(actual_u_actualizado))
                    with st.expander("renovacion_u_actualizado"):
                        st.write(limpiar_dineros(renovacion_u_actualizado))
                    with st.expander("docs_adicionales_data"):
                        st.write(docs_adicionales_data)
                    with st.expander("poliza_actual"):
                        st.write(poliza_actual.get("data", {}))
                    with st.expander("poliza_renovacion"):
                        st.write(poliza_renovacion.get("data", {}))
                    with st.expander("totales_actuales"):
                        st.write(totales_actuales)
                    with st.expander("totales_renovacion"):
                        st.write(totales_renovacion)

                summary_output_path = generar_tabla_excel_rc(
                    amparos_actuales=coberturas_predeterminadas,
                    amparos_renovacion=coberturas_predeterminadas,
                    clasificacion_actual=limpiar_dineros(actual_u_actualizado),
                    clasificacion_renovacion=limpiar_dineros(renovacion_u_actualizado),
                    docs_adicionales_data=docs_adicionales_data,
                    poliza_actual=poliza_actual.get("data", {}),
                    poliza_renovacion=poliza_renovacion.get("data", {}),
                    totales_actual=totales_actuales,
                    totales_renovacion=totales_renovacion,
                    titulo_excel=excel_name,
                    output_path=summary_excel_path,
                )

                integrar_hoja_en_libro(
                    ruta_libro_principal=main_output_path,
                    ruta_libro_origen=summary_output_path,
                    nombre_hoja_origen=None,
                    nombre_hoja_nueva=None,
                    crear_respaldo=False,
                )

                with open(main_output_path, "rb") as f:
                    excel_bytes = f.read()
                    b64_excel = base64.b64encode(excel_bytes).decode()
                    href = (
                        "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                        + b64_excel
                    )
                    st.markdown(
                        f"""
                            <a class="download-btn" href="{href}" download="reporte_polizas_riesgos.xlsx" target="_blank" role="button" aria-label="Descargar Excel de análisis">
                            📥 <span>Descargar Excel</span>
                            </a>
                            """,
                        unsafe_allow_html=True,
                    )

            except Exception as e:
                st.error(f"Error al generar el archivo Excel: {e}")
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

    else:
        st.info(
            "Por favor presiona el botón 'Iniciar Proceso' en la barra lateral para continuar."
        )


if __name__ == "__main__":
    asyncio.run(main())
