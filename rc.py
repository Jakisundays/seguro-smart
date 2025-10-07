# Clasificación de imports por bloques funcionales
# 1) Librerías estándar
import json
import os
import datetime
from copy import copy as _copy
from typing import List, Dict, Any, Optional

# 2) Ciencia de datos
import pandas as pd

# 3) openpyxl – manipulación de Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation


def clasificar_por_tipo(detalles):
    """
    Clasifica los elementos de 'detalle_cobertura' por tipo y retorna un diccionario
    con las llaves específicas solicitadas, cada una conteniendo un arreglo de detalles.
    Si un detalle pertenece a múltiples tipos, se incluye en todos los arreglos correspondientes.
    """
    objetivos = [
        "Incendio",
        "Sustracción",
        "Equipo Electrónico",
        "Rotura de Maquinaria",
        "Transporte de Valores",
        "Manejo de Dinero",
        "Responsabilidad Civil",
    ]
    resultado = {k: [] for k in objetivos}

    # Normalización para mapear variantes (acentos/casos) a las llaves objetivo
    normalizacion = {
        "incendio": "Incendio",
        "sustracción": "Sustracción",
        "sustraccion": "Sustracción",
        "equipo electrónico": "Equipo Electrónico",
        "equipo electronico": "Equipo Electrónico",
        "rotura de maquinaria": "Rotura de Maquinaria",
        "transporte de valores": "Transporte de Valores",
        "manejo de dinero": "Manejo de Dinero",
        "responsabilidad civil": "Responsabilidad Civil",
    }

    if not detalles:
        return resultado

    for detalle in detalles:
        for tipo in detalle.get("tipo", []):
            clave = normalizacion.get(str(tipo).strip().lower())
            if clave:
                resultado[clave].append(detalle)

    return resultado


def generar_tabla_excel_rc(
    amparos_actuales: Dict[str, List[Dict[str, Any]]],
    amparos_renovacion: Dict[str, List[Dict[str, Any]]],
    clasificacion_actual: Dict[str, List[Dict[str, Any]]],
    clasificacion_renovacion: Dict[str, List[Dict[str, Any]]],
    docs_adicionales_data: Optional[List[Dict[str, Any]]] = None,
    poliza_actual: Optional[Dict[str, Any]] = None,
    poliza_renovacion: Optional[Dict[str, Any]] = None,
    output_path: str = "Resumen_RC.xlsx",
) -> str:
    """
    Genera un archivo de Excel segmentado por tipo con:
      - Columna A: "Coberturas" (amparos únicos combinando Actual y Renovación)
      - Columna B: "Intereses Asegurados" (únicos combinando Actual y Renovación)
      - Columna C: "Valor Asegurado (Actual / Renovación)" para cada interés

    Notas:
      - Todo se agrupa por categoría/tipo. Cada tipo se imprime como sección.
      - Se evita duplicidad tanto en coberturas como en intereses (por nombre normalizado).
      - Cuando un interés existe en una sola póliza, el valor de la otra queda en blanco.
      - Parámetros opcionales `poliza_actual` y `poliza_renovacion` permiten agregar detalles
        en columnas de "PRIMA" para condiciones actuales y de renovación sin depender de variables globales.
    """
    # Validaciones
    if not isinstance(amparos_actuales, dict) or not isinstance(
        amparos_renovacion, dict
    ):
        raise ValueError(
            "'amparos_actuales' y 'amparos_renovacion' deben ser diccionarios clasificados por tipo."
        )
    if not isinstance(clasificacion_actual, dict) or not isinstance(
        clasificacion_renovacion, dict
    ):
        raise ValueError(
            "'clasificacion_actual' y 'clasificacion_renovacion' deben ser diccionarios clasificados por tipo."
        )

    def _norm(s: Any) -> str:
        return str(s).strip().lower()

    def _to_number(v: Any) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        try:
            # Limpia símbolos y separadores comunes
            digits = "".join(ch for ch in str(v) if ch.isdigit())
            return float(digits) if digits else 0.0
        except Exception:
            return 0.0

    def _fmt_money(v: float) -> str:
        try:
            return f"${v:,.0f}" if v else ""
        except Exception:
            return str(v)

    # 1) Coberturas únicas por tipo
    cob_por_tipo: Dict[str, Dict[str, str]] = {}
    for dataset in (amparos_actuales, amparos_renovacion):
        for tipo, items in (dataset or {}).items():
            if not isinstance(items, list):
                continue
            t = str(tipo)
            cob_por_tipo.setdefault(t, {})
            m = cob_por_tipo[t]
            for it in items:
                if not isinstance(it, dict):
                    continue
                nombre = it.get("amparo")
                if not nombre:
                    continue
                k = _norm(nombre)
                if k not in m:
                    m[k] = str(nombre).strip()

    # 2) Intereses por tipo (combinando Actual/Renovación y sus valores)
    intereses_por_tipo: Dict[str, Dict[str, Dict[str, Any]]] = {}

    def acumular_intereses(
        fuente: Dict[str, List[Dict[str, Any]]], es_actual: bool
    ) -> None:
        for tipo, items in (fuente or {}).items():
            if not isinstance(items, list):
                continue
            t = str(tipo)
            intereses_por_tipo.setdefault(t, {})
            mapa = intereses_por_tipo[t]
            for it in items:
                if not isinstance(it, dict):
                    continue
                nombre = it.get("interes_asegurado")
                if not nombre:
                    continue
                k = _norm(nombre)
                entry = mapa.setdefault(
                    k, {"nombre": str(nombre).strip(), "actual": 0.0, "renovacion": 0.0}
                )
                val = _to_number(it.get("valor_asegurado", 0))
                if es_actual:
                    entry["actual"] += val
                else:
                    entry["renovacion"] += val

    acumular_intereses(clasificacion_actual, es_actual=True)
    acumular_intereses(clasificacion_renovacion, es_actual=False)

    if not cob_por_tipo and not intereses_por_tipo:
        raise ValueError("No hay información suficiente para generar el Excel.")

    # Orden sugerido de tipos y fallback
    orden_preferente = [
        "Incendio",
        "Sustracción",
        "Equipo Electrónico",
        "Equipo Electronico",
        "Rotura de Maquinaria",
        "Manejo de Dinero",
        "Transporte de Valores",
        "Responsabilidad Civil",
    ]
    tipos_presentes = set(list(cob_por_tipo.keys()) + list(intereses_por_tipo.keys()))
    tipos_ordenados = [t for t in orden_preferente if t in tipos_presentes]
    tipos_ordenados += [t for t in tipos_presentes if t not in tipos_ordenados]

    # Crear Excel (tres columnas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # Encabezados
    headers = [
        "COBERTURAS",
        "INTERESES ASEGURADOS",
        "VALOR ASEGURADO ACTUAL",
        "VALOR ASEGURADO RENOVACIÓN",
        "TASA",
        "Prima",
        "TASA",
        "Prima",
    ]
    ws.append(headers)
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Encabezado base multi-fila: columnas simples con merge vertical; bloques de condiciones horizontales
    # Columnas con merge vertical (mantienen título en fila 1)
    for c in [1, 2, 3, 4]:
        for r in range(1, 3 + 1):
            cell = ws.cell(row=r, column=c)
            if r == 1:
                cell.value = headers[c - 1]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=c, end_row=3, end_column=c)

    # Bloque "CONDICIONES ACTUALES" abarca columnas 5 y 6
    ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=5)
    title_cell.value = "CONDICIONES ACTUALES"
    title_cell.fill = header_fill
    title_cell.font = header_font
    title_cell.alignment = header_align
    title_cell.border = thin_border
    for col in (5, 6):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border
    ws.cell(row=2, column=5).value = "TASA"
    ws.cell(row=2, column=6).value = "Prima"
    for col in (5, 6):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border
    for col in (5, 6):
        c = ws.cell(row=3, column=col)
        c.value = ""
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

    # Bloque "CONDICIONES DE RENOVACIÓN" abarca columnas 7 y 8
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=7)
    title_cell.value = "CONDICIONES DE RENOVACIÓN"
    title_cell.fill = header_fill
    title_cell.font = header_font
    title_cell.alignment = header_align
    title_cell.border = thin_border
    for col in (7, 8):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border
    ws.cell(row=2, column=7).value = "TASA"
    ws.cell(row=2, column=8).value = "Prima"
    for col in (7, 8):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border
    for col in (7, 8):
        c = ws.cell(row=3, column=col)
        c.value = ""
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = thin_border

    # Columnas dinámicas para COTIZACIONES (documentos adicionales)
    docs = docs_adicionales_data or []
    doc_headers = [
        str(d.get("Archivo", f"Documento {i+1}")) for i, d in enumerate(docs)
    ]
    num_docs = len(doc_headers)

    # Obtener primas por documento (soporta claves en mayúsculas/minúsculas y valores con/ sin IVA)
    doc_primas = []
    for d in docs:
        prima_val = None
        for k in ("Prima Con IVA", "prima_con_iva", "Prima Sin IVA", "prima_sin_iva"):
            if prima_val is None and d.get(k) is not None:
                prima_val = d.get(k)
        # Normalizar a número
        if isinstance(prima_val, (int, float)):
            val_num = float(prima_val)
        else:
            val_num = 0.0
            try:
                import re

                s = str(prima_val)
                s = re.sub(r"[^\d.,-]", "", s)
                if s.count(",") == 1 and s.count(".") >= 1:
                    s = s.replace(".", "").replace(",", ".")
                else:
                    s = s.replace(",", "")
                val_num = float(s) if s else 0.0
            except Exception:
                val_num = 0.0
        doc_primas.append(val_num)

    # Obtener tasas por documento
    doc_tasas = []
    for d in docs:
        tasa_val = d.get("tasa")
        # Normalizar a número
        if isinstance(tasa_val, (int, float)):
            val_num = float(tasa_val)
        else:
            val_num = None
            try:
                import re

                s = str(tasa_val)
                s = re.sub(r"[^\d.,-]", "", s)
                if s.count(",") == 1 and s.count(".") >= 1:
                    s = s.replace(".", "").replace(",", ".")
                else:
                    s = s.replace(",", "")
                val_num = float(s)
            except Exception:
                val_num = None
        doc_tasas.append(val_num)

    # Encabezado multi-fila para COTIZACIONES
    if num_docs > 0:
        first_doc_col = 9
        last_doc_col = 8 + (num_docs * 2)
        # Fila 1: título COTIZACIONES (merge I1:Ultima)
        ws.merge_cells(
            start_row=1, start_column=first_doc_col, end_row=1, end_column=last_doc_col
        )
        title_cell = ws.cell(row=1, column=first_doc_col)
        title_cell.value = "COTIZACIONES"
        title_cell.fill = header_fill
        title_cell.font = header_font
        title_cell.alignment = header_align
        title_cell.border = thin_border
        # Asegurar estilo/borde en todo el rango merged
        for col in range(first_doc_col, last_doc_col + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            c.border = thin_border
        # Fila 2: nombre de archivo por documento (merge 2 columnas por documento)
        for j, h in enumerate(doc_headers):
            start_col = first_doc_col + (j * 2)
            end_col = start_col + 1
            ws.merge_cells(
                start_row=2, start_column=start_col, end_row=2, end_column=end_col
            )
            c1 = ws.cell(row=2, column=start_col)
            c1.value = h
            for col in (start_col, end_col):
                c = ws.cell(row=2, column=col)
                c.fill = header_fill
                c.font = header_font
                c.alignment = header_align
                c.border = thin_border
        # Fila 3: textos TASA y PRIMA en columnas por documento (TASA a la izquierda)
        for j in range(num_docs):
            tasa_col = first_doc_col + (j * 2)
            prima_col = tasa_col + 1
            c1 = ws.cell(row=3, column=tasa_col)
            c1.value = "TASA"
            c1.fill = header_fill
            c1.font = header_font
            c1.alignment = header_align
            c1.border = thin_border
            c2 = ws.cell(row=3, column=prima_col)
            c2.value = "PRIMA"
            c2.fill = header_fill
            c2.font = header_font
            c2.alignment = header_align
            c2.border = thin_border

    section_fill = PatternFill(
        start_color="B7DEE8", end_color="B7DEE8", fill_type="solid"
    )
    section_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    total_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )

    # Iniciar datos en la fila 4 debido a encabezados en filas 1-3
    fila = 4
    for idx, tipo in enumerate(tipos_ordenados):
        # Espacio entre secciones
        if idx > 0:
            ws.append(["", "", "", "", "", "", "", ""])
            fila += 1
        # Título de sección (merge A:H)
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=8)
        cell = ws.cell(row=fila, column=1)
        cell.value = str(tipo).upper()
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = center_align
        cell.border = thin_border
        fila += 1

        # Filas en blanco disponibles dentro de la sección para detalles de documentos
        blank_rows_section: List[int] = []

        # Datos de la sección
        cob_list = sorted(
            list((cob_por_tipo.get(tipo) or {}).values()), key=lambda s: s.lower()
        )
        interes_map = intereses_por_tipo.get(tipo) or {}
        interes_list = sorted(
            [
                (v["nombre"], v.get("actual", 0.0), v.get("renovacion", 0.0))
                for v in interes_map.values()
            ],
            key=lambda t: t[0].lower(),
        )
        n = max(len(cob_list), len(interes_list), 1)
        section_rows_start = fila
        for i in range(n):
            cob = cob_list[i] if i < len(cob_list) else ""
            if i < len(interes_list):
                nombre_int, val_act, val_ren = interes_list[i]
                val_act_str = _fmt_money(val_act)
                val_ren_str = _fmt_money(val_ren)
            else:
                nombre_int, val_act_str, val_ren_str = "", "", ""
            ws.append([cob, nombre_int, val_act_str, val_ren_str, "", "", "", ""])
            # Bordes por fila
            for c in range(1, 9 + (num_docs * 2)):
                ws.cell(row=fila, column=c).border = thin_border
            fila += 1

        # Insertar detalles en columnas dinámicas (I en adelante) usando filas de la sección
        tipo_norm = str(tipo).strip().lower()
        docs = docs_adicionales_data or []
        for j, doc in enumerate(docs):
            tasa_col = 9 + (j * 2)
            prima_col = tasa_col + 1
            details: List[str] = []
            if tipo_norm == "incendio":
                dm = doc.get("danos_materiales", {}) or {}
                pairs = [
                    ("Incendio Máximo", dm.get("incendio_maximo")),
                    ("Terremoto Máximo", dm.get("terremoto_maximo")),
                    ("Terrorismo Máximo", dm.get("terrorismo_maximo")),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "equipo electronico":
                dm = doc.get("danos_materiales", {}) or {}
                pairs = [
                    ("Equipo electrónico", dm.get("equipo_electronico")),
                    (
                        "Equipos móviles y portátiles",
                        dm.get("equipos_moviles_portatiles"),
                    ),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "manejo de dinero":
                mgc = doc.get("manejo_global_comercial", {}) or {}
                pairs = [
                    ("Pérdidas máx anual", mgc.get("perdidas_maximo_anual")),
                    (
                        "Empleados no identificados",
                        mgc.get("empleados_no_identificados"),
                    ),
                    (
                        "Empleados temporales/firma",
                        mgc.get("empleados_temporales_firma"),
                    ),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "transporte de valores":
                tv = doc.get("transporte_valores", {}) or {}
                pairs = [
                    ("Límite máximo despacho", tv.get("limite_maximo_despacho")),
                    (
                        "Presupuesto anual movilizaciones",
                        tv.get("presupuesto_anual_movilizaciones"),
                    ),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "responsabilidad civil":
                rc = doc.get("responsabilidad_civil", {}) or {}
                ordered_keys = [
                    ("Vehículos propios/no propios", "vehiculos_propios_no_propios"),
                    ("Gastos urgencias médicas", "gastos_urgencias_medicas"),
                    ("Contratistas y subcontratistas", "contratistas_subcontratistas"),
                    ("Parqueaderos", "parqueaderos"),
                    ("Cruzada", "cruzada"),
                    ("Productos", "productos"),
                    ("Patronal", "patronal"),
                ]
                details = [
                    f"{label}: {rc.get(key)}"
                    for (label, key) in ordered_keys
                    if rc.get(key) is not None
                ]
            elif tipo_norm == "sustracción":
                dm = doc.get("danos_materiales", {}) or {}
                pairs = [
                    ("Sustracción Máximo", dm.get("sustraccion_maximo")),
                    ("Sustracción sin violencia", dm.get("sustraccion_sin_violencia")),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "rotura de maquinaria":
                dm = doc.get("danos_materiales", {}) or {}
                val = dm.get("rotura_maquinaria")
                details = [f"Rotura de maquinaria: {val}"] if val else []

            for k, text in enumerate(details):
                if k >= n:
                    break
                r = section_rows_start + k
                ccell = ws.cell(row=r, column=prima_col)
                ccell.value = text
                ccell.border = thin_border
                ccell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

        # Insertar detalles en "PRIMA" de CONDICIONES ACTUALES (col 6) y RENOVACIÓN (col 8)
        for pol, dest_col in (
            (poliza_actual or {}, 6),
            (poliza_renovacion or {}, 8),
        ):
            details = []
            if tipo_norm == "incendio":
                dm = pol.get("danos_materiales", {}) or {}
                pairs = [
                    ("Incendio Máximo", dm.get("incendio_maximo")),
                    ("Terremoto Máximo", dm.get("terremoto_maximo")),
                    ("Terrorismo Máximo", dm.get("terrorismo_maximo")),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "equipo electronico":
                dm = pol.get("danos_materiales", {}) or {}
                pairs = [
                    ("Equipo electrónico", dm.get("equipo_electronico")),
                    (
                        "Equipos móviles y portátiles",
                        dm.get("equipos_moviles_portatiles"),
                    ),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "manejo de dinero":
                mgc = pol.get("manejo_global_comercial", {}) or {}
                pairs = [
                    ("Pérdidas máx anual", mgc.get("perdidas_maximo_anual")),
                    (
                        "Empleados no identificados",
                        mgc.get("empleados_no_identificados"),
                    ),
                    (
                        "Empleados temporales/firma",
                        mgc.get("empleados_temporales_firma"),
                    ),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "transporte de valores":
                tv = pol.get("transporte_valores", {}) or {}
                pairs = [
                    ("Límite máximo despacho", tv.get("limite_maximo_despacho")),
                    (
                        "Presupuesto anual movilizaciones",
                        tv.get("presupuesto_anual_movilizaciones"),
                    ),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "responsabilidad civil":
                rc = pol.get("responsabilidad_civil", {}) or {}
                ordered_keys = [
                    ("Vehículos propios/no propios", "vehiculos_propios_no_propios"),
                    ("Gastos urgencias médicas", "gastos_urgencias_medicas"),
                    ("Contratistas y subcontratistas", "contratistas_subcontratistas"),
                    ("Parqueaderos", "parqueaderos"),
                    ("Cruzada", "cruzada"),
                    ("Productos", "productos"),
                    ("Patronal", "patronal"),
                ]
                details = [
                    f"{label}: {rc.get(key)}"
                    for (label, key) in ordered_keys
                    if rc.get(key) is not None
                ]
            elif tipo_norm == "sustracción":
                dm = pol.get("danos_materiales", {}) or {}
                pairs = [
                    ("Sustracción Máximo", dm.get("sustraccion_maximo")),
                    ("Sustracción sin violencia", dm.get("sustraccion_sin_violencia")),
                ]
                details = [f"{label}: {val}" for (label, val) in pairs if val]
            elif tipo_norm == "rotura de maquinaria":
                dm = pol.get("danos_materiales", {}) or {}
                val = dm.get("rotura_maquinaria")
                details = [f"Rotura de maquinaria: {val}"] if val else []
            for k, text in enumerate(details):
                if k >= n:
                    break
                r = section_rows_start + k
                ccell = ws.cell(row=r, column=dest_col)
                ccell.value = text
                ccell.border = thin_border
                ccell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

        # Fila de total por segmento (Intereses asegurados)
        total_act = sum(v[1] for v in interes_list)
        total_ren = sum(v[2] for v in interes_list)
        tipo_norm = str(tipo).strip().lower()
        # Extraer tasas y primas para INCENDIO desde poliza_actual y poliza_renovacion
        p_act = poliza_actual
        p_ren = poliza_renovacion
        tasa_act = p_act.get("tasa", "") if tipo_norm == "incendio" else ""
        prima_act = p_act.get("prima_sin_iva", "") if tipo_norm == "incendio" else ""
        tasa_ren = p_ren.get("tasa", "") if tipo_norm == "incendio" else ""
        prima_ren = p_ren.get("prima_sin_iva", "") if tipo_norm == "incendio" else ""
        ws.append(
            [
                "",
                f"TOTAL {str(tipo).upper()}",
                _fmt_money(total_act),
                _fmt_money(total_ren),
                str(tasa_act) if tasa_act != "" else "",
                _fmt_money(prima_act) if prima_act != "" else "",
                str(tasa_ren) if tasa_ren != "" else "",
                _fmt_money(prima_ren) if prima_ren != "" else "",
            ]
        )
        for c in range(1, 9):
            cell = ws.cell(row=fila, column=c)
            cell.border = thin_border
            cell.fill = total_fill
            if c in (2, 3, 4):
                cell.font = Font(bold=True)
        # Solo en TOTAL INCENDIO, colocar PRIMA y TASA de documentos adicionales en columnas I en adelante
        # tipo_norm ya definido arriba
        if tipo_norm == "incendio":
            for j, prima in enumerate(doc_primas):
                prima_col = 9 + (j * 2)
                tasa_col = prima_col + 1
                # PRIMA
                c_prima = ws.cell(row=fila, column=prima_col)
                c_prima.value = str(tasa_val)
                c_prima.border = thin_border
                c_prima.fill = total_fill
                c_prima.font = Font(bold=True)
                # TASA
                tasa_val = doc_tasas[j] if j < len(doc_tasas) else None
                c_tasa = ws.cell(row=fila, column=tasa_col)
                c_tasa.value = _fmt_money(prima) if prima is not None else ""
                c_tasa.border = thin_border
                c_tasa.fill = total_fill
                c_tasa.font = Font(bold=True)
        fila += 1

    # Congelar encabezado y ajustar anchos
    ws.freeze_panes = "A4"
    for col in range(1, 9 + (num_docs * 2)):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=col).value
            l = len(str(val)) if val is not None else 0
            if l > max_len:
                max_len = l
        from openpyxl.utils import get_column_letter as _gcl

        ws.column_dimensions[_gcl(col)].width = min(max_len + 4, 80)

    # Guardar
    try:
        wb.save(output_path)
    except Exception as exc:
        raise RuntimeError(
            f"No se pudo guardar el archivo Excel en '{output_path}': {exc}"
        )

    return output_path


def _safe_sheet_title(wb, desired_title: str) -> str:
    """
    Ensure the sheet title is unique within the workbook.
    If a conflict exists, append an incrementing suffix.
    """
    title = desired_title or "Sheet"
    if title not in wb.sheetnames:
        return title
    base = title
    i = 2
    while True:
        candidate = f"{base} ({i})"
        if candidate not in wb.sheetnames:
            return candidate
        i += 1


def _copy_sheet_contents(src: Worksheet, dst: Worksheet) -> None:
    """
    Copy values, styles, merges, dimensions and common sheet settings
    from src worksheet to dst worksheet, aiming to preserve appearance.
    """
    # Copy cell values and styles
    for row in src.iter_rows():
        for cell in row:
            dcell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            # Styles and formatting
            if cell.has_style:
                dcell.font = _copy(cell.font)
                dcell.border = _copy(cell.border)
                dcell.fill = _copy(cell.fill)
                dcell.number_format = cell.number_format
                dcell.protection = _copy(cell.protection)
                dcell.alignment = _copy(cell.alignment)
            # Hyperlinks and comments
            if cell.hyperlink:
                dcell.hyperlink = cell.hyperlink
            if cell.comment:
                dcell.comment = _copy(cell.comment)

    # Merged cells
    for merged_range in src.merged_cells.ranges:
        dst.merge_cells(str(merged_range))

    # Column dimensions (widths, hidden, etc.)
    for key, col_dim in src.column_dimensions.items():
        dd = dst.column_dimensions[key]
        # Width triggers customWidth internally in openpyxl; no direct setter for customWidth
        try:
            dd.width = col_dim.width
        except Exception:
            pass
        try:
            dd.min = col_dim.min
            dd.max = col_dim.max
        except Exception:
            pass
        try:
            dd.hidden = col_dim.hidden
            dd.bestFit = col_dim.bestFit
        except Exception:
            pass

    # Row dimensions (heights, hidden)
    for idx, row_dim in src.row_dimensions.items():
        dd = dst.row_dimensions[idx]
        dd.height = row_dim.height
        dd.hidden = row_dim.hidden

    # Freeze panes
    dst.freeze_panes = src.freeze_panes

    # Print options and page setup/margins
    try:
        dst.print_options.horizontalCentered = src.print_options.horizontalCentered
        dst.print_options.verticalCentered = src.print_options.verticalCentered
    except Exception:
        pass

    try:
        pm = src.page_margins
        dpm = dst.page_margins
        dpm.left = pm.left
        dpm.right = pm.right
        dpm.top = pm.top
        dpm.bottom = pm.bottom
        dpm.header = pm.header
        dpm.footer = pm.footer
    except Exception:
        pass

    try:
        ps = src.page_setup
        dps = dst.page_setup
        dps.orientation = ps.orientation
        dps.paperSize = ps.paperSize
        dps.fitToPage = ps.fitToPage
        dps.fitToHeight = ps.fitToHeight
        dps.fitToWidth = ps.fitToWidth
        dps.scale = ps.scale
    except Exception:
        pass

    # Sheet view (zoom, gridlines)
    try:
        dst.sheet_view.zoomScale = src.sheet_view.zoomScale
        dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    except Exception:
        pass

    # Auto filter
    try:
        if src.auto_filter and src.auto_filter.ref:
            dst.auto_filter.ref = src.auto_filter.ref
    except Exception:
        pass

    # Conditional formatting rules
    try:
        cf = src.conditional_formatting
        for rng in cf:
            for rule in cf[rng]:
                # rule objects can be reused; openpyxl will rebind as needed
                dst.conditional_formatting.add(rng, rule)
    except Exception:
        pass

    # Data validations
    try:
        if src.data_validations is not None:
            for dv in src.data_validations.dataValidation:
                ndv = DataValidation(
                    type=dv.type,
                    formula1=dv.formula1,
                    formula2=dv.formula2,
                    allow_blank=dv.allow_blank,
                    operator=dv.operator,
                    showErrorMessage=dv.showErrorMessage,
                    showInputMessage=dv.showInputMessage,
                    error=dv.error,
                    errorTitle=dv.errorTitle,
                    prompt=dv.prompt,
                    promptTitle=dv.promptTitle,
                )
                # Copy ranges
                for sqref in dv.sqref:
                    ndv.add(str(sqref))
                dst.add_data_validation(ndv)
    except Exception:
        pass


def integrar_hoja_en_libro(
    ruta_libro_principal: str,
    ruta_libro_origen: str,
    nombre_hoja_origen: Optional[str] = None,
    nombre_hoja_nueva: Optional[str] = None,
    crear_respaldo: bool = False,
) -> str:
    """
    Abre `ruta_libro_principal` (loli.xlsx) y agrega como nueva hoja el contenido
    de `ruta_libro_origen` (Resumen_RC.xlsx), preservando formatos y estructura.

    Retorna la ruta del archivo principal guardado.
    """
    # Cargar libros
    wb_dest = load_workbook(ruta_libro_principal)
    wb_src = load_workbook(ruta_libro_origen)

    # Selección de hoja de origen
    if nombre_hoja_origen:
        if nombre_hoja_origen not in wb_src.sheetnames:
            raise ValueError(
                f"La hoja '{nombre_hoja_origen}' no existe en el libro de origen."
            )
        ws_src = wb_src[nombre_hoja_origen]
    else:
        ws_src = wb_src.active

    # Determinar título de nueva hoja
    desired_title = nombre_hoja_nueva or ws_src.title
    new_title = _safe_sheet_title(wb_dest, desired_title)

    # Crear nueva hoja y copiar contenido
    ws_dst = wb_dest.create_sheet(title=new_title)
    _copy_sheet_contents(ws_src, ws_dst)

    # Respaldo antes de sobrescribir
    if crear_respaldo:
        base = os.path.basename(ruta_libro_principal)
        root, ext = os.path.splitext(base)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"{root}_backup_{timestamp}{ext}"
        backup_path = os.path.join(os.path.dirname(ruta_libro_principal), backup_name)
        wb_backup = load_workbook(ruta_libro_principal)
        wb_backup.save(backup_path)

    # Guardar cambios en el libro principal
    wb_dest.save(ruta_libro_principal)
    return ruta_libro_principal


# # poliza_actual
# poliza_actual = {
#     "detalle_cobertura": [
#         {
#             "interes_asegurado": "Edificios",
#             "valor_asegurado": 22615066519,
#             "tipo": ["Incendio", "Responsabilidad Civil"],
#         },
#         {
#             "interes_asegurado": "Equipo Electrónico",
#             "valor_asegurado": 1204502680,
#             "tipo": ["Equipo Electronico", "Sustracción", "Incendio"],
#         },
#         {
#             "interes_asegurado": "Muebles y Enseres",
#             "valor_asegurado": 1621213188,
#             "tipo": ["Incendio", "Sustracción"],
#         },
#         {
#             "interes_asegurado": "Maquinaria y Equipo",
#             "valor_asegurado": 2928485866,
#             "tipo": ["Rotura de Maquinaria", "Incendio", "Sustracción"],
#         },
#         {
#             "interes_asegurado": "Mercancías",
#             "valor_asegurado": 46200000,
#             "tipo": ["Incendio", "Sustracción"],
#         },
#         {
#             "interes_asegurado": "Dinero",
#             "valor_asegurado": 196000000,
#             "tipo": ["Manejo de Dinero", "Sustracción", "Transporte de Valores"],
#         },
#         {
#             "interes_asegurado": "Eq. Móviles y Portátiles",
#             "valor_asegurado": 82695800,
#             "tipo": ["Equipo Electronico", "Sustracción"],
#         },
#     ],
#     "total_valores_asegurados": 28694164053,
#     "amparos": [
#         {
#             "amparo": "Incendio y/o Impacto Directo De Rayo",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Explosión",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Extensión de Amparos",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Explosión de Calderas u Otros Aparatos Generadores de Vapor",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria", "Incendio"],
#         },
#         {
#             "amparo": "Rotura Accidental De Vidrios",
#             "deducible": "0,25 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Terremoto, Temblor De Tierra, Erupción Volcánica, Tsunami, Maremoto",
#             "deducible": "2% del valor asegurable del artículo mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Anegación",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daños Por Agua",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "AMIT Y Terrorismo",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "HMACC",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Productos Almacenados En Frigoríficos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daño Interno Equipos Eléctricos Y Electrónicos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Portadores Externos de Datos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Equipos Móviles Y Portátiles",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Incremento en Costos De Operación",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daños por Fallas En Equipos De Climatización",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria", "Equipo Electronico"],
#         },
#         {
#             "amparo": "Rotura De Maquinaria",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria"],
#         },
#         {
#             "amparo": "Pérdida de Contenido en Tanques",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio", "Sustracción"],
#         },
#         {
#             "amparo": "Deterioro de Bienes Refrigerados",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria", "Incendio"],
#         },
#         {
#             "amparo": "Hurto Calificado",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción"],
#         },
#         {
#             "amparo": "Hurto Simple para Equipo Eléctrico Y Electrónico Fijo de Oficina",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción", "Equipo Electronico"],
#         },
#         {
#             "amparo": "Bienes de Propiedad de Empleados del Asegurado",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Manejo de Dinero", "Sustracción"],
#         },
#         {
#             "amparo": "Traslado Temporal de Bienes",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción", "Incendio"],
#         },
#         {
#             "amparo": "Construcciones y Montajes Nuevos",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio", "Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Bienes a la Intemperie",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio", "Sustracción"],
#         },
#         {
#             "amparo": "Actos de Autoridad",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Remoción de Escombros",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Honorarios de Profesionales como Arquitectos, Interventores, Ingenieros y Consultores",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Gastos para la Preservación de Bienes",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Gastos para la Reproducción y/o Reemplazo de la Información",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Gastos para Demostrar la Pérdida",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Transporte De Valores",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Transporte de Valores"],
#         },
#         {
#             "amparo": "Asistencia",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Labores y Materiales",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Gastos de Extinción del siniestro",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daños o pérdidas de Mercancías a Granel",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio", "Sustracción"],
#         },
#         {
#             "amparo": "Manejo",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Empleados de Carácter Temporal y/o de Firmas Especializadas",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Empleados no Identificados",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Bienes de Propiedad de Terceros",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero", "Sustracción"],
#         },
#         {
#             "amparo": "Protección para Depósitos Bancarios",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Predios Labores y Operaciones",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Contratistas y Subcontratistas",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Responsabilidad Civil Patronal",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Vehículos Propios y no Propios",
#             "deducible": "En exceso del SOAT y RCE autos mínimo 100.000.000/100.000.000/200.000.000 COP",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Propietarios, Arrendatarios y Poseedores",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Gastos Médicos",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Responsabilidad Civil Cruzada",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Responsabilidad Civil Parqueaderos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#     ],
#     "riesgos": [
#         {
#             "ubicacion": "1 Calle 28 # 30-15 PALMIRA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 3061592832,
#                     "tipo": ["Incendio", "Responsabilidad Civil"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 141240000,
#                     "tipo": ["Rotura de Maquinaria", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 58850000,
#                     "tipo": ["Equipo Electronico", "Incendio", "Sustracción"],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "2 Carrera 18 #8-31 FLORIDA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 505019752,
#                     "tipo": ["Incendio", "Responsabilidad Civil"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 100145560,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 77255146,
#                     "tipo": ["Rotura de Maquinaria", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 59352677,
#                     "tipo": ["Equipo Electronico", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 11000000,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": [
#                         "Manejo de Dinero",
#                         "Sustracción",
#                         "Transporte de Valores",
#                     ],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "3 Calle 6 # 12-50 PRADERA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 586566851,
#                     "tipo": ["Incendio", "Responsabilidad Civil"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 123035974,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 77255146,
#                     "tipo": ["Rotura de Maquinaria", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 55386331,
#                     "tipo": ["Equipo Electronico", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 11000000,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": [
#                         "Manejo de Dinero",
#                         "Sustracción",
#                         "Transporte de Valores",
#                     ],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "4 Carrera 7 #. 11-56 CANDELARIA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 806887083,
#                     "tipo": ["Incendio", "Responsabilidad Civil"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 91561655,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 38627573,
#                     "tipo": ["Rotura de Maquinaria", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 46705241,
#                     "tipo": ["Equipo Electronico", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 2200000,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": [
#                         "Manejo de Dinero",
#                         "Sustracción",
#                         "Transporte de Valores",
#                     ],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "5 Carrera 31 # 28-00 y Calle 28 #31-30 Esq PALMIRA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 17655000000,
#                     "tipo": ["Incendio", "Responsabilidad Civil"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 921591000,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "OBRAS DE ARTE",
#                     "valor_asegurado": 384879000,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 2594108000,
#                     "tipo": ["Rotura de Maquinaria", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 977178797,
#                     "tipo": ["Equipo Electronico", "Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO MOVIL",
#                     "valor_asegurado": 73562500,
#                     "tipo": ["Equipo Electronico", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 22000000,
#                     "tipo": ["Incendio", "Sustracción"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 130000000,
#                     "tipo": [
#                         "Manejo de Dinero",
#                         "Sustracción",
#                         "Transporte de Valores",
#                     ],
#                 },
#             ],
#         },
#     ],
#     "danos_materiales": {
#         "incendio_maximo": "28.694.164.053,00",
#         "terremoto_maximo": "28.694.164.053,00",
#         "terrorismo_maximo": "28.694.164.053,00",
#         "sustraccion_maximo": "5.996.401.734,00",
#         "dinero_fuera_caja_fuerte": "1.000.000.000,00",
#         "dinero_dentro_caja_fuerte": "130.000.000,00",
#         "sustraccion_sin_violencia": "1.204.502.680,00",
#         "equipo_electronico": "1.204.502.680,00",
#         "equipos_moviles_portatiles": "82.695.800,00",
#         "rotura_maquinaria": "2.928.485.866,00",
#     },
#     "manejo_global_comercial": {
#         "perdidas_maximo_anual": "20.000.000,00",
#         "empleados_no_identificados": "20.000.000,00",
#         "empleados_temporales_firma": "20.000.000,00",
#     },
#     "transporte_valores": {
#         "limite_maximo_despacho": "20.000.000,00",
#         "presupuesto_anual_movilizaciones": "1.000.000.000,00",
#     },
#     "responsabilidad_civil": {
#         "vehiculos_propios_no_propios": "2.000.000.000,00",
#         "gastos_urgencias_medicas": "100.000.000,00",
#         "contratistas_subcontratistas": "2.000.000.000,00",
#         "parqueaderos": "200.000.000,00",
#         "cruzada": "1.000.000.000,00",
#         "productos": "N/A",
#         "patronal": "2.000.000.000,00",
#     },
#     "tasa": 0.10797434778,
#     "iva": 5882492.62,
#     "prima_con_iva": 36842980.11,
#     "prima_sin_iva": 30960487.49,
#     "file_name": "POLIZA ACTUAL EMPRESARIAL 128663630-0 CAMARA DE COMERCIO DE PALMIRA 2024-2025 (1).pdf",
# }

# # poliza_renovacion
# poliza_renovacion = {
#     "detalle_cobertura": [
#         {
#             "interes_asegurado": "Edificios",
#             "valor_asegurado": 24876573170,
#             "tipo": ["Incendio"],
#         },
#         {
#             "interes_asegurado": "Equipo Electrónico",
#             "valor_asegurado": 1245453104,
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "interes_asegurado": "Muebles y Enseres",
#             "valor_asegurado": 1783334507,
#             "tipo": ["Incendio"],
#         },
#         {
#             "interes_asegurado": "Maquinaria y Equipo",
#             "valor_asegurado": 3221334452,
#             "tipo": ["Rotura de Maquinaria"],
#         },
#         {
#             "interes_asegurado": "Mercancías",
#             "valor_asegurado": 46200000,
#             "tipo": ["Incendio"],
#         },
#         {
#             "interes_asegurado": "Dinero",
#             "valor_asegurado": 196000000,
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "interes_asegurado": "Eq. Móviles y Portátiles",
#             "valor_asegurado": 94156337,
#             "tipo": ["Equipo Electronico"],
#         },
#     ],
#     "total_valores_asegurados": 31463051570,
#     "amparos": [
#         {
#             "amparo": "Incendio y/o Impacto Directo De Rayo",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Explosión",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Extensión de Amparos",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Explosión de Calderas u Otros Aparatos Generadores de Vapor",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria"],
#         },
#         {
#             "amparo": "Rotura Accidental De Vidrios",
#             "deducible": "0,25 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Terremoto, Temblor De Tierra, Erupción Volcánica, Tsunami, Maremoto",
#             "deducible": "2% del valor asegurable del artículo mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Anegación",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daños Por Agua",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "AMIT Y Terrorismo",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "HMACC",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Productos Almacenados En Frigoríficos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daño Interno Equipos Eléctricos Y Electrónicos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Portadores Externos de Datos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Equipos Móviles Y Portátiles",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Incremento en Costos De Operación",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daños por Fallas En Equipos De Climatización",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria"],
#         },
#         {
#             "amparo": "Rotura De Maquinaria",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Rotura de Maquinaria"],
#         },
#         {
#             "amparo": "Pérdida de Contenido en Tanques",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Deterioro de Bienes Refrigerados",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Hurto Calificado",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción"],
#         },
#         {
#             "amparo": "Hurto Simple para Equipo Eléctrico Y Electrónico Fijo de Oficina",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción", "Equipo Electronico"],
#         },
#         {
#             "amparo": "Bienes de Propiedad de Empleados del Asegurado",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción"],
#         },
#         {
#             "amparo": "Traslado Temporal de Bienes",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Sustracción"],
#         },
#         {
#             "amparo": "Construcciones y Montajes Nuevos",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Bienes a la Intemperie",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Actos de Autoridad",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Remoción de Escombros",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Honorarios de Profesionales como Arquitectos, Interventores, Ingenieros y Consultores",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Gastos para la Preservación de Bienes",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Gastos para la Reproducción y/o Reemplazo de la Información",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Equipo Electronico"],
#         },
#         {
#             "amparo": "Gastos para Demostrar la Pérdida",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Transporte De Valores",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Transporte de Valores"],
#         },
#         {
#             "amparo": "Asistencia",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Labores y Materiales",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Gastos de Extinción del siniestro",
#             "deducible": "Aplica el de la cobertura afectada",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Daños o pérdidas de Mercancías a Granel",
#             "deducible": "5% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Incendio"],
#         },
#         {
#             "amparo": "Manejo",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Empleados de Carácter Temporal y/o de Firmas Especializadas",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Empleados no Identificados",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Bienes de Propiedad de Terceros",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Protección para Depósitos Bancarios",
#             "deducible": "10% del valor de la pérdida mínimo 2 SMMLV",
#             "tipo": ["Manejo de Dinero"],
#         },
#         {
#             "amparo": "Predios Labores y Operaciones",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Contratistas y Subcontratistas",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Responsabilidad Civil Patronal",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Vehículos Propios y no Propios",
#             "deducible": "En exceso del SOAT y RCE autos mínimo 100.000.000/100.000.000/200.000.000 COP",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Propietarios, Arrendatarios y Poseedores",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Gastos Médicos",
#             "deducible": "N/A",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Responsabilidad Civil Cruzada",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#         {
#             "amparo": "Responsabilidad Civil Parqueaderos",
#             "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#             "tipo": ["Responsabilidad Civil"],
#         },
#     ],
#     "riesgos": [
#         {
#             "ubicacion": "CLL 28 #30-15 PALMIRA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 24876573170,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 1783334507,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 3221334452,
#                     "tipo": ["Rotura de Maquinaria"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCÍAS",
#                     "valor_asegurado": 46200000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "DINERO",
#                     "valor_asegurado": 196000000,
#                     "tipo": ["Manejo de Dinero"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRONICO",
#                     "valor_asegurado": 1245453104,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO MOVIL Y PORTATIL",
#                     "valor_asegurado": 94156337,
#                     "tipo": ["Equipo Electronico"],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "CALLE 28 # 18-44 FLORIDA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 1367732116,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 155364000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 64735000,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": ["Manejo de Dinero"],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "CALLE 28 A # 18-64 FLORIDA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 84980661,
#                     "tipo": ["Rotura de Maquinaria"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRONICO",
#                     "valor_asegurado": 69072144,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 11000000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": ["Manejo de Dinero"],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "CALLE 6 # 12-50 PRADERA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO Y MEJORAS",
#                     "valor_asegurado": 597000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 135339571,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 188880661,
#                     "tipo": ["Rotura de Maquinaria"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRICO Y ELECTRONICO",
#                     "valor_asegurado": 64709165,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 11000000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": ["Manejo de Dinero"],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "CARRERA 38 # 32-34 CANDELARIA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO",
#                     "valor_asegurado": 381575791,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 100717820,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 42490830,
#                     "tipo": ["Rotura de Maquinaria"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRONICO",
#                     "valor_asegurado": 35159965,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 11000000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 22000000,
#                     "tipo": ["Manejo de Dinero"],
#                 },
#             ],
#         },
#         {
#             "ubicacion": "CARRERA 31 # 28-00 Y CALLE 28 #31-30 ESA PALMIRA",
#             "detalle_cobertura": [
#                 {
#                     "interes_asegurado": "EDIFICIO Y OBRAS DE ARTE",
#                     "valor_asegurado": 2437117000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MUEBLES Y ENSERES",
#                     "valor_asegurado": 1013750100,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "MAQUINARIA Y EQUIPO",
#                     "valor_asegurado": 283850000,
#                     "tipo": ["Rotura de Maquinaria"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO ELECTRONICO",
#                     "valor_asegurado": 991776830,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "EQUIPO MOVIL",
#                     "valor_asegurado": 94156337,
#                     "tipo": ["Equipo Electronico"],
#                 },
#                 {
#                     "interes_asegurado": "MERCANCIAS",
#                     "valor_asegurado": 23100000,
#                     "tipo": ["Incendio"],
#                 },
#                 {
#                     "interes_asegurado": "DINEROS",
#                     "valor_asegurado": 130000000,
#                     "tipo": ["Manejo de Dinero"],
#                 },
#             ],
#         },
#     ],
#     "danos_materiales": {
#         "incendio_maximo": "31.463.051.570,00",
#         "terremoto_maximo": "31.463.051.570,00",
#         "terrorismo_maximo": "31.463.051.570,00",
#         "sustraccion_maximo": "6.492.322.063,00",
#         "dinero_fuera_caja_fuerte": "null",
#         "dinero_dentro_caja_fuerte": "null",
#         "sustraccion_sin_violencia": "null",
#         "equipo_electronico": "1.245.453.104,00",
#         "equipos_moviles_portatiles": "94.156.337,00",
#         "rotura_maquinaria": "3.221.334.452,00",
#     },
#     "manejo_global_comercial": {
#         "perdidas_maximo_anual": "20.000.000,00",
#         "empleados_no_identificados": "20.000.000,00",
#         "empleados_temporales_firma": "20.000.000,00",
#     },
#     "transporte_valores": {
#         "limite_maximo_despacho": "20.000.000,00",
#         "presupuesto_anual_movilizaciones": "1.000.000.000,00",
#     },
#     "responsabilidad_civil": {
#         "vehiculos_propios_no_propios": "2.000.000.000,00",
#         "gastos_urgencias_medicas": "100.000.000,00",
#         "contratistas_subcontratistas": "2.000.000.000,00",
#         "parqueaderos": "200.000.000,00",
#         "cruzada": "1.000.000.000,00",
#         "productos": "null",
#         "patronal": "2.000.000.000,00",
#     },
#     "tasa": 0,
#     "iva": 0,
#     "prima_con_iva": 0,
#     "prima_sin_iva": 40653811.84,
#     "file_name": "PR336-186756401-26358662 COT CyC PALMIRA.pdf",
# }
# # docs_adicionales_data
# docs_adicionales_data = [
#     {
#         "Archivo": "Slip Cotización 2500008845 Pyme Segura 10+AXA COLPATRIA.pdf",
#         "Tipo de Documento": "Adicional",
#         "Prima Sin IVA": 10636637,
#         "IVA": 2020961,
#         "Prima Con IVA": 12657598,
#         "tasa": 0.002,
#         "amparos": [
#             {
#                 "amparo": "Todo riesgo incendio",
#                 "deducible": "10% de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Terremoto, Temblor, Erupción Volcánica y Maremoto, Marejada, Tsunami",
#                 "deducible": "2% valor asegurable del ítem afectado, mínimo 2 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Hmac-Amit/terrorismo",
#                 "deducible": "15% de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Sustracción con violencia para dinero en efectivo y títulos valores",
#                 "deducible": "10% de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Sustracción", "Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Sustracción con violencia",
#                 "deducible": "10% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Sustracción sin violencia - Contenido",
#                 "deducible": "20% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Equipo eléctrico y electrónico",
#                 "deducible": "10% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Rotura de maquinarias",
#                 "deducible": "10% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Pérdida arrendamiento",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos alojamiento temporal",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Índice variable edificio",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Índice variable maquinaria",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Manejo básico (Manejo global comercial)",
#                 "deducible": "10% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Manejo demás (Manejo global comercial)",
#                 "deducible": "15% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Transporte de valores",
#                 "deducible": "5% de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Transporte de Valores"],
#             },
#             {
#                 "amparo": "Predios, labores y operaciones (Responsabilidad Civil)",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Gastos médicos (Responsabilidad Civil)",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Gastos de defensa (Responsabilidad Civil)",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Parqueaderos (Responsabilidad Civil)",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Viajes al exterior (Responsabilidad Civil)",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad civil patronal (por accidentes de trabajo)",
#                 "deducible": "En exceso Seguridad Social",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad civil por vehículos propios y no propios",
#                 "deducible": "En exceso de $100.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad civil contratistas y subcontratistas (Incluye RCE Cruzada)",
#                 "deducible": "En exceso de $10.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad civil por contaminación accidental",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "RCE bienes bajo cuidado, tenencia y control",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "RC productos y/o trabajos terminados",
#                 "deducible": "$1.500.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#         ],
#         "danos_materiales": {
#             "incendio_maximo": "$ 5.304.471.886",
#             "terremoto_maximo": "$ 5.304.471.886",
#             "terrorismo_maximo": "$ 5.304.471.886",
#             "sustraccion_maximo": "$ 2.542.798.438",
#             "dinero_fuera_caja_fuerte": "$ 17.500.000",
#             "dinero_dentro_caja_fuerte": "$ 17.500.000",
#             "sustraccion_sin_violencia": "No especificado",
#             "equipo_electronico": "$ 454.777.139",
#             "equipos_moviles_portatiles": "No especificado",
#             "rotura_maquinaria": "$ 66.770.519",
#         },
#         "manejo_global_comercial": {
#             "perdidas_maximo_anual": "$ 10.000.000",
#             "empleados_no_identificados": "$ 5.000.000",
#             "empleados_temporales_firma": "$ 5.000.000",
#         },
#         "transporte_valores": {
#             "limite_maximo_despacho": "$ 20.000.000",
#             "presupuesto_anual_movilizaciones": "$ 1.050.000.000",
#         },
#         "responsabilidad_civil": {
#             "vehiculos_propios_no_propios": "$ 1.500.000.000",
#             "gastos_urgencias_medicas": "$ 1.500.000.000",
#             "contratistas_subcontratistas": "$ 1.500.000.000",
#             "parqueaderos": "$ 1.500.000.000",
#             "cruzada": "$ 1.500.000.000",
#             "productos": "$ 1.500.000.000",
#             "patronal": "$ 1.500.000.000",
#         },
#     },
#     {
#         "Archivo": "SLIP TRDM OFICINAS_CAMARA DE COMERCIO.pdf",
#         "Tipo de Documento": "Adicional",
#         "Prima Sin IVA": 28289403,
#         "IVA": 0,
#         "Prima Con IVA": 28289403,
#         "tasa": 0.09,
#         "amparos": [
#             {
#                 "amparo": "Incendio y/o rayo o sus efectos inmediatos como calor y humo",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Explosión",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Extensión de Amparos (Huracán, tifón, tornado, ciclón, granizo, vientos fuertes, caída de aeronaves, choque de vehículos terrestres)",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Inundación, anegación, avalancha, enlodamiento y daños por agua",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Terremoto, temblor, erupción volcánica, maremoto, marejada y/o tsunami",
#                 "deducible": "2% del valor asegurable del artículo afectado por el siniestro, mínimo 1 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Huelga, asonada, motín, conmoción civil o popular, actos mal intencionados de terceros, sabotaje y terrorismo (HMACC y AMIT)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Sustracción con violencia (bienes diferentes a equipos eléctricos y electrónicos)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Sustracción con violencia (dineros dentro de caja fuerte)",
#                 "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#                 "tipo": ["Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Sustracción con violencia (dineros fuera de caja fuerte)",
#                 "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#                 "tipo": ["Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Rotura de maquinaria",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Equipo Eléctrico y Electrónico (Daño interno y hurto calificado)",
#                 "deducible": "Hurto calificado: 10% del valor de la pérdida, mínimo 1 SMMLV. Demás eventos: 20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Hurto simple (equipos fijos de oficina)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Equipos de cómputo móviles y portátiles",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Amparo automático de nuevos bienes o propiedades",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Remoción de escombros",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Propiedad personal de empleados",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Gastos para la extinción del siniestro",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos para evitar la propagación del siniestro",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos para la preservación de bienes",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos para acelerar la reparación o reemplazo",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos de viaje y estadía",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Honorarios de auditores, revisores y contadores",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos por concepto de horas extras, trabajo nocturno, trabajo en días feriados y flete expreso",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos por reparaciones provisionales o construcciones transitorias y/o temporales",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Honorarios profesionales",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Incendio y/o rayo en aparatos eléctricos",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Portadores externos de datos",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Gastos para la reposición de documentos",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Rotura accidental de vidrios",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Traslados temporales de maquinaria y equipos (excluyendo transporte, cargue y descargue)",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Nuevas construcciones y montajes menores",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Amparo automático de ferias y eventos",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Daños a calderas y/o aparatos generadores de vapor",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Daños a cimentaciones",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Daños por fallas en equipos de climatización",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Pérdida de contenidos en tanques",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Flete aéreo",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos adicionales derivados del siniestro",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos para demostración de la ocurrencia y cuantía del siniestro",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos para la obtención de licencias y permisos",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Incremento en el costo de operación para equipos electrónicos procesadores",
#                 "deducible": "3 días calendario",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Bienes bajo cuidado, tenencia y control",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Amparo automático para equipos de reemplazo temporal",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Amparo automático de máquinas y equipos en demostración",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Obras de arte o murales decorativos",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Cobertura para aceites, lubricantes y refrigerantes",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Reparación estética del inmueble",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos para el reacondicionamiento de jardines naturales u ornamentales y/o urbanismo",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Bienes de propiedad de visitantes",
#                 "deducible": "15% de la pérdida, mínimo 5 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#         ],
#         "danos_materiales": {
#             "incendio_maximo": "$31.432.470.033 por evento y en el agregado anual",
#             "terremoto_maximo": "$31.432.470.033 por evento y en el agregado anual",
#             "terrorismo_maximo": "$31.432.470.033 por evento y en el agregado anual",
#             "sustraccion_maximo": "$4.627.502.059 por evento y en el agregado anual",
#             "dinero_fuera_caja_fuerte": "$10.000.000 por evento y en el agregado anual",
#             "dinero_dentro_caja_fuerte": "$30.000.000 por evento y en el agregado anual",
#             "sustraccion_sin_violencia": "$984.208.430 por evento y en el agregado anual",
#             "equipo_electronico": "$1.309.227.904 por evento y en el agregado anual",
#             "equipos_moviles_portatiles": "$20.000.000 por evento y $50.000.000 en el agregado anual",
#             "rotura_maquinaria": "$3.221.334.452 por evento y en el agregado anual",
#         },
#         "manejo_global_comercial": {
#             "perdidas_maximo_anual": "No especificado en el documento",
#             "empleados_no_identificados": "No especificado en el documento",
#             "empleados_temporales_firma": "No especificado en el documento",
#         },
#         "transporte_valores": {
#             "limite_maximo_despacho": "No especificado en el documento",
#             "presupuesto_anual_movilizaciones": "No especificado en el documento",
#         },
#         "responsabilidad_civil": {
#             "vehiculos_propios_no_propios": "No especificado en el documento",
#             "gastos_urgencias_medicas": "No especificado en el documento",
#             "contratistas_subcontratistas": "No especificado en el documento",
#             "parqueaderos": "No especificado en el documento",
#             "cruzada": "No especificado en el documento",
#             "productos": "No especificado en el documento",
#             "patronal": "No especificado en el documento",
#         },
#     },
#     {
#         "prima_sin_iva": 10636637,
#         "iva": 2020961,
#         "prima_con_iva": 12657598,
#         "tasa": 0.09,
#         "amparos": [
#             {
#                 "amparo": "Todo riesgo incendio",
#                 "deducible": "10% de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Terremoto, temblor, erupción volcánica y maremoto, marejada, tsunami",
#                 "deducible": "2% del valor asegurable del artículo afectado, mínimo 1 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Huelga, motín, asonada, conmoción civil y/o popular y actos mal intencionados de terceros, incluido actos terroristas y terrorismo (HMACC - AMIT)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Sustracción con violencia",
#                 "deducible": "10% del valor de la pérdida mínimo 1 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Sustracción sin violencia - Contenido",
#                 "deducible": "20% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Sustracción"],
#             },
#             {
#                 "amparo": "Equipo electrónico (Daño interno y hurto calificado)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Equipo electrónico (Hurto simple)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Equipo electrónico (Equipos de cómputo móviles y portátiles)",
#                 "deducible": "10% del valor de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Equipo Electronico"],
#             },
#             {
#                 "amparo": "Rotura de Maquinaria",
#                 "deducible": "20% del valor de la pérdida, mínimo 3 SMMLV",
#                 "tipo": ["Rotura de Maquinaria"],
#             },
#             {
#                 "amparo": "Rotura de vidrios",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Pérdida arrendamiento",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Gastos alojamiento temporal",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Incendio"],
#             },
#             {
#                 "amparo": "Manejo Global Comercial (Hurto, Abuso de confianza, Falsedad, Estafa)",
#                 "deducible": "10% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Manejo Global Comercial (Empleados no identificados / Temporales)",
#                 "deducible": "15% de la pérdida, mínimo 2 SMMLV",
#                 "tipo": ["Manejo de Dinero"],
#             },
#             {
#                 "amparo": "Transporte de Valores (Pérdida o daño material)",
#                 "deducible": "5% de la pérdida, mínimo 1 SMMLV",
#                 "tipo": ["Transporte de Valores"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Predios, labores y operaciones",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Gastos médicos",
#                 "deducible": "Sin deducible",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Parqueaderos",
#                 "deducible": "$1.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Vehículos propios y no propios",
#                 "deducible": "En exceso de $100.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Contratistas y Subcontratistas (Incluye Cruzada)",
#                 "deducible": "En exceso de $10.000.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Productos y/o trabajos terminados",
#                 "deducible": "$1.500.000",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#             {
#                 "amparo": "Responsabilidad Civil - Patronal",
#                 "deducible": "En exceso Seguridad Social",
#                 "tipo": ["Responsabilidad Civil"],
#             },
#         ],
#         "danos_materiales": {
#             "incendio_maximo": "$31.432.470.033",
#             "terremoto_maximo": "$31.432.470.033",
#             "terrorismo_maximo": "$31.432.470.033",
#             "sustraccion_maximo": "$4.627.502.059",
#             "dinero_fuera_caja_fuerte": "$10.000.000",
#             "dinero_dentro_caja_fuerte": "$30.000.000",
#             "sustraccion_sin_violencia": "$984.208.430",
#             "equipo_electronico": "$1.309.227.904",
#             "equipos_moviles_portatiles": "$50.000.000",
#             "rotura_maquinaria": "$3.221.334.452",
#         },
#         "manejo_global_comercial": {
#             "perdidas_maximo_anual": "$10.000.000",
#             "empleados_no_identificados": "$5.000.000",
#             "empleados_temporales_firma": "$5.000.000",
#         },
#         "transporte_valores": {
#             "limite_maximo_despacho": "$20.000.000",
#             "presupuesto_anual_movilizaciones": "$1.050.000.000",
#         },
#         "responsabilidad_civil": {
#             "vehiculos_propios_no_propios": "Límite único $1.500.000.000 por evento y vigencia",
#             "gastos_urgencias_medicas": "Límite único $1.500.000.000 por evento y vigencia",
#             "contratistas_subcontratistas": "Límite único $1.500.000.000 por evento y vigencia",
#             "parqueaderos": "Límite único $1.500.000.000 por evento y vigencia",
#             "cruzada": "Límite único $1.500.000.000 por evento y vigencia",
#             "productos": "Límite único $1.500.000.000 por evento y vigencia",
#             "patronal": "Límite único $1.500.000.000 por evento y vigencia",
#         },
#         "file_name": "Slip Cotización 2500008845 Pyme Segura 10+AXA COLPATRIA.pdf, SLIP TRDM OFICINAS_CAMARA DE COMERCIO.pdf",
#         "Archivo": "Slip Cotización 2500008845 Pyme Segura 10+AXA COLPATRIA.pdf, SLIP TRDM OFICINAS_CAMARA DE COMERCIO.pdf",
#         "Prima Sin IVA": 10636637,
#         "IVA": 2020961,
#         "Prima Con IVA": 12657598,
#         "Tipo de Documento": "Adicional",
#     },
# ]


# if __name__ == "__main__":
#     amparos_actuales = clasificar_por_tipo(poliza_actual["amparos"])
#     amparos_renovacion = clasificar_por_tipo(poliza_renovacion["amparos"])
#     clasificacion_actual = clasificar_por_tipo(poliza_actual["detalle_cobertura"])
#     clasificacion_renovacion = clasificar_por_tipo(
#         poliza_renovacion["detalle_cobertura"]
#     )

#     try:
#         ruta_excel = generar_tabla_excel_rc(
#             amparos_actuales,
#             amparos_renovacion,
#             clasificacion_actual,
#             clasificacion_renovacion,
#             docs_adicionales_data,
#             poliza_actual,
#             poliza_renovacion,
#             output_path="Resumen_RC.xlsx",
#         )
#         print(f"Tabla de Excel generada correctamente: {ruta_excel}")
#     except Exception as e:
#         print(f"Error al generar la tabla de Excel: {e}")
